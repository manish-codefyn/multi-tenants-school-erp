import logging
import csv
import io
import re
import uuid
import json
import openpyxl
from django.shortcuts import render
from django.views.generic import View, ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Count, Q, Sum, Avg
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.core.exceptions import PermissionDenied
from django.contrib.auth import get_user_model
from django.utils.translation import gettext_lazy as _

from apps.core.permissions.mixins import PermissionRequiredMixin, RoleRequiredMixin, TenantAccessMixin
from apps.core.utils.tenant import get_current_tenant
from apps.core.utils.audit import audit_log
from apps.core.views import (
    BaseView, BaseListView, BaseDetailView, BaseCreateView, 
    BaseUpdateView, BaseDeleteView, BaseTemplateView, ExportMixin
)

logger = logging.getLogger(__name__)

from .models import (
    Department, Designation, Staff, StaffAddress, StaffDocument,
    StaffAttendance, LeaveType, LeaveApplication, LeaveBalance,
    SalaryStructure, Payroll, Promotion, EmploymentHistory,
    TrainingProgram, TrainingParticipation, PerformanceReview,
    Recruitment, JobApplication, Holiday, WorkSchedule, TaxConfig, PFESIConfig,
    Qualification
)
from .forms import (
    DepartmentForm, DesignationForm, StaffForm, StaffAddressForm,
    StaffDocumentForm, AttendanceForm, LeaveTypeForm, LeaveApplicationForm,
    LeaveBalanceForm, SalaryStructureForm, PayrollForm, PromotionForm,
    EmploymentHistoryForm, TrainingProgramForm, TrainingParticipationForm,
    PerformanceReviewForm, RecruitmentForm, JobApplicationForm, PublicJobApplicationForm, StaffImportForm,
    QualificationForm
)


User = get_user_model()
from .idcard import StaffIDCardGenerator


class HRDashboardView(BaseTemplateView):
    template_name = 'hr/dashboard.html'
    permission_required = 'hr.view_dashboard'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        today = timezone.now().date()
        
        # Staff Statistics
        staff_queryset = Staff.objects.filter(tenant=tenant, is_active=True)
        context['total_staff'] = staff_queryset.count()
        context['total_teachers'] = staff_queryset.filter(user__role='teacher').count()
        context['teaching_staff'] = staff_queryset.filter(designation__category='TEACHING').count()
        context['non_teaching_staff'] = staff_queryset.filter(designation__category='NON_TEACHING').count()
        context['administrative_staff'] = staff_queryset.filter(designation__category='ADMINISTRATIVE').count()
        
        # Department Statistics
        departments = Department.objects.filter(tenant=tenant).annotate(
            staff_members_count=Count('staff_members', filter=Q(staff_members__is_active=True))
        )
        context['departments'] = departments
        context['total_departments'] = departments.count()
        
        # Today's StaffAttendance
        context['present_today'] = StaffAttendance.objects.filter(
            tenant=tenant, 
            date=today, 
            status__in=['PRESENT', 'LATE']
        ).count()
        context['absent_today'] = StaffAttendance.objects.filter(
            tenant=tenant, 
            date=today, 
            status='ABSENT'
        ).count()
        context['on_leave_today'] = StaffAttendance.objects.filter(
            tenant=tenant, 
            date=today, 
            status='LEAVE'
        ).count()
        
        # Leave Statistics
        context['pending_leaves'] = LeaveApplication.objects.filter(
            tenant=tenant,
            status='PENDING'
        ).count()
        
        # Payroll Statistics
        current_month = today.replace(day=1)
        context['payroll_processed'] = Payroll.objects.filter(
            tenant=tenant,
            salary_month__year=current_month.year,
            salary_month__month=current_month.month,
            status='PAID'
        ).count()
        
        # Recent Activity
        context['recent_staff'] = staff_queryset.select_related(
            'user', 'department', 'designation'
        ).order_by('-joining_date')[:5]
        
        context['recent_leaves'] = LeaveApplication.objects.filter(
            tenant=tenant
        ).select_related('staff', 'leave_type').order_by('-applied_date')[:5]
        
        # Performance Overview
        recent_reviews = PerformanceReview.objects.filter(
            tenant=tenant,
            review_date__gte=timezone.now() - timezone.timedelta(days=90)
        )
        if recent_reviews.exists():
            context['avg_performance'] = recent_reviews.aggregate(
                avg=Avg('overall_rating')
            )['avg']
        
        # Audit Trail
        audit_log(
            user=self.request.user,
            action='VIEW_DASHBOARD',
            resource_type='HR Dashboard',
            details={'tenant_id': str(tenant.id)},
            severity='INFO'
        )
        
        return context


class StaffListView(BaseListView):
    model = Staff
    template_name = 'hr/staff/list.html'
    context_object_name = 'staff_members'
    permission_required = 'hr.view_staff'
    paginate_by = 20
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant, 
            is_active=True
        ).select_related(
            'user', 'department', 'designation'
        ).prefetch_related('addresses')
        
        # Apply filters
        self.apply_filters(queryset)
        
        return queryset
    
    def apply_filters(self, queryset):
        filters = Q()
        
        # Role filter
        role = self.request.GET.get('role')
        if role and role != 'all':
            filters &= Q(user__role=role)
        
        # Department filter
        department_id = self.request.GET.get('department')
        if department_id and department_id != 'all':
            filters &= Q(department_id=department_id)
        
        # Employment type filter
        emp_type = self.request.GET.get('employment_type')
        if emp_type and emp_type != 'all':
            filters &= Q(employment_type=emp_type)
        
        # Employment status filter
        emp_status = self.request.GET.get('employment_status')
        if emp_status and emp_status != 'all':
            filters &= Q(employment_status=emp_status)
        
        # Search
        search = self.request.GET.get('search', '').strip()
        if search:
            filters &= (
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__email__icontains=search) |
                Q(employee_id__icontains=search) |
                Q(personal_phone__icontains=search)
            )
        
        return queryset.filter(filters)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['departments'] = Department.objects.filter(tenant=tenant)
        context['designations'] = Designation.objects.filter(tenant=tenant)
        context['filters'] = {
            'role': self.request.GET.get('role', ''),
            'department': self.request.GET.get('department', ''),
            'employment_type': self.request.GET.get('employment_type', ''),
            'employment_status': self.request.GET.get('employment_status', ''),
            'search': self.request.GET.get('search', ''),
        }
        
        return context


class StaffDetailView(BaseDetailView):
    model = Staff
    template_name = 'hr/staff/detail.html'
    context_object_name = 'staff'
    permission_required = 'hr.view_staff'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        staff = self.object
        
        # Add related data
        context['addresses'] = staff.addresses.all()
        context['documents'] = staff.documents.all()
        context['leave_applications'] = staff.leave_applications.all().order_by('-start_date')[:10]
        context['attendance_records'] = staff.attendances.all().order_by('-date')[:30]
        context['promotions'] = staff.promotions.all().order_by('-effective_date')
        context['performance_reviews'] = staff.performance_reviews.all().order_by('-review_date')
        
        # Leave balances
        context['leave_balances'] = LeaveBalance.objects.filter(
            staff=staff,
            year=timezone.now().year
        ).select_related('leave_type')
        
        # Recent payroll
        context['recent_payroll'] = Payroll.objects.filter(
            staff=staff
        ).order_by('-salary_month')[:6]
        
        # Training participation
        context['trainings'] = TrainingParticipation.objects.filter(
            staff=staff
        ).select_related('training').order_by('-training__start_date')
        
        return context


class StaffCreateView(BaseCreateView):
    model = Staff
    form_class = StaffForm
    template_name = 'hr/staff/form.html'
    permission_required = 'hr.add_staff'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:staff_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Staff member {form.instance.full_name} created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_STAFF',
            resource_type='Staff',
            resource_id=str(self.object.id),
            details={
                'staff_id': str(self.object.id),
                'employee_id': self.object.employee_id,
                'name': self.object.full_name
            },
            severity='INFO'
        )
        
        return response


class StaffUpdateView(BaseUpdateView):
    model = Staff
    form_class = StaffForm
    template_name = 'hr/staff/form.html'
    permission_required = 'hr.change_staff'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:staff_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Staff member {form.instance.full_name} updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_STAFF',
            resource_type='Staff',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class StaffDeleteView(BaseDeleteView):
    model = Staff
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_staff'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:staff_list')
    
    def delete(self, request, *args, **kwargs):
        staff = self.get_object()
        
        # Perform soft delete
        staff.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Staff member {staff.full_name} has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_STAFF',
            resource_type='Staff',
            resource_id=str(staff.id),
            details={
                'employee_id': staff.employee_id,
                'name': staff.full_name,
                'reason': request.POST.get('deletion_reason')
            },
            severity='WARNING'
        )
        

        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class StaffIDCardView(BaseDetailView):
    """Generate ID Card for staff"""
    model = Staff
    permission_required = 'hr.view_staff'
    roles_required = ['admin', 'staff','super_admin'] 
    
    def get(self, request, *args, **kwargs):
        staff = self.get_object()
        
        # Check permissions: can only view own ID card unless admin/hr
        if not (request.user.role in ['admin','staff','super_admin','hr'] or staff.user == request.user):
            raise PermissionDenied("You do not have permission to view ID card for this staff member.")
            
        generator = StaffIDCardGenerator(staff)
        return generator.get_id_card_response()



class StaffImportView(BaseView):
    """View to import staff from CSV/Excel"""
    permission_required = 'hr.add_staff'
    roles_required = ['admin', 'hr_manager']
    template_name = 'hr/staff/import.html'
    
    def get(self, request, *args, **kwargs):
        form = StaffImportForm()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request, *args, **kwargs):
        form = StaffImportForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})
            
        file = request.FILES['file']
        tenant = get_current_tenant()
        created_count = 0
        updated_count = 0
        errors = []
        
        try:
            # Read file based on extension
            if file.name.endswith('.csv'):
                data_list = []
                # Use utf-8-sig to handle BOM from Excel
                decoded_file = file.read().decode('utf-8-sig').splitlines()
                reader = csv.DictReader(decoded_file)
                # Normalize headers
                if reader.fieldnames:
                    reader.fieldnames = [str(name).strip().lower().replace(' ', '_') for name in reader.fieldnames]
                
                for row in reader:
                    data_list.append(row)
            else:
                # Excel
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                # Normalize headers
                headers = []
                for cell in ws[1]:
                    value = cell.value
                    if value:
                        headers.append(str(value).strip().lower().replace(' ', '_'))
                    else:
                        headers.append('') # Handle empty header cells if any
                
                data_list = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Only zip if we have headers, and handle row length
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_dict[headers[i]] = value
                    data_list.append(row_dict)
            
            # Process data
            for index, row in enumerate(data_list):
                row_num = index + 2
                email = row.get('email')
                first_name = row.get('first_name')
                # Handle flexible names
                if not first_name and row.get('name'):
                    name_parts = row['name'].strip().split(' ', 1)
                    first_name = name_parts[0]
                    if len(name_parts) > 1 and not row.get('last_name'):
                        row['last_name'] = name_parts[1]
                
                if not email or not first_name:
                    errors.append(f"Row {row_num}: Missing required keys. Found: {list(row.keys())}")
                    continue
                    
                # Create/Get User
                user, user_created = User.objects.get_or_create(
                    email=email.strip(),
                    defaults={
                        'password': 'Staff@123',
                        'first_name': first_name.strip(),
                        'last_name': row.get('last_name', '').strip(),
                        'role': 'staff',
                        'is_active': True,
                        'tenant': tenant
                    }
                )
                
                if user_created:
                    user.set_password('Staff@123') # Default password
                    user.save()
                
                # Validate IDs and Lookup by Name
                dept_id = row.get('department_id')
                desig_id = row.get('designation_id')
                
                # Check Department (ID or Name)
                if dept_id:
                    try:
                         uuid.UUID(str(dept_id))
                    except (ValueError, TypeError):
                        dept_id = None
                
                if not dept_id and row.get('department'):
                    dept_obj = Department.objects.filter(tenant=tenant, name__iexact=row['department'].strip()).first()
                    if dept_obj:
                        dept_id = dept_obj.id

                # Check Designation (ID or Name)
                if desig_id:
                    try:
                        uuid.UUID(str(desig_id))
                    except (ValueError, TypeError):
                        desig_id = None
                        
                if not desig_id and row.get('designation'):
                    desig_obj = Designation.objects.filter(tenant=tenant, title__iexact=row['designation'].strip()).first()
                    if desig_obj:
                        desig_id = desig_obj.id

                # Create/Update Staff
                staff, staff_created = Staff.objects.update_or_create(
                    user=user,
                    defaults={
                        'department_id': dept_id,
                        'designation_id': desig_id,
                        'joining_date': row.get('joining_date', timezone.now().date()),
                        'date_of_birth': row.get('date_of_birth', '2000-01-01'),
                        'personal_phone': row.get('phone', ''),
                        'tenant': tenant
                    }
                )
                
                if staff_created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            messages.success(request, f"Import complete. Created: {created_count}, Updated: {updated_count}")
            if errors:
                messages.warning(request, f"Encounters {len(errors)} errors: " + "; ".join(errors[:5]))
                
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            
        return redirect('hr:staff_list')


class StaffImportSampleView(BaseView):
    """Download sample import file"""
    permission_required = 'hr.add_staff'
    roles_required = ['admin', 'hr_manager']
    
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="staff_import_sample.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['first_name', 'last_name', 'email', 'phone', 'department_id', 'designation_id', 'joining_date', 'date_of_birth'])
        
        # Get real examples if available
        tenant = get_current_tenant()
        dept = Department.objects.filter(tenant=tenant).first()
        desig = Designation.objects.filter(tenant=tenant).first()
        
        dept_id = str(dept.id) if dept else 'UUID-OR-EMPTY'
        desig_id = str(desig.id) if desig else 'UUID-OR-EMPTY'
        
        writer.writerow(['John', 'Doe', 'john@example.com', '9876543210', dept_id, desig_id, '2023-01-01', '1990-01-01'])
        
        return response


class StaffExportView(ExportMixin, BaseListView):
    """Export staff list"""
    permission_required = 'hr.view_staff'
    model = Staff
    export_filename = 'staff_export'
    
    def get(self, request, *args, **kwargs):
        return self.export(request, *args, **kwargs)
        
    def get_queryset(self):
        return super().get_queryset().select_related('user', 'department', 'designation')

    def get_export_headers(self):
        return ['first_name', 'last_name', 'email', 'phone', 'department_id', 'designation_id', 'joining_date', 'date_of_birth']
    
    def get_export_row(self, staff):
        return [
            staff.user.first_name,
            staff.user.last_name,
            staff.user.email,
            staff.personal_phone,
            str(staff.department.id) if staff.department else '',
            str(staff.designation.id) if staff.designation else '',
            staff.joining_date,
            staff.date_of_birth
        ]


# ==================== DEPARTMENT VIEWS ====================

class DepartmentListView(BaseListView):
    model = Department
    template_name = 'hr/department/list.html'
    context_object_name = 'departments'
    permission_required = 'hr.view_department'


class DepartmentCreateView(BaseCreateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'hr/department/form.html'
    permission_required = 'hr.add_department'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_success_url(self):
        return reverse_lazy('hr:department_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        return super().form_valid(form)


class DepartmentUpdateView(BaseUpdateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'hr/department/form.html'
    permission_required = 'hr.change_department'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_success_url(self):
        return reverse_lazy('hr:department_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)


class DepartmentDeleteView(BaseDeleteView):
    model = Department
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_department'
    roles_required = ['admin', 'principal']
    
    def get_success_url(self):
        return reverse_lazy('hr:department_list')
    
    def delete(self, request, *args, **kwargs):
        department = self.get_object()
        
        # Check if department has staff
        if department.staff_members.filter(is_active=True).exists():
            messages.error(
                request,
                f"Cannot delete department '{department.name}' because it has active staff members."
            )
            return JsonResponse({'success': False, 'error': 'Department has active staff'})
            
        return super().delete(request, *args, **kwargs)


# ==================== StaffAttendance VIEWS ====================

class AttendanceListView(BaseListView):
    model = StaffAttendance
    template_name = 'hr/attendance/list.html'
    context_object_name = 'attendances'
    permission_required = 'hr.view_attendance'
    paginate_by = 50
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('staff', 'staff__user')
        
        # Date filter
        date_filter = self.request.GET.get('date')
        if date_filter:
            try:
                date_obj = timezone.datetime.strptime(date_filter, '%Y-%m-%d').date()
                queryset = queryset.filter(date=date_obj)
            except ValueError:
                pass
        
        # Status filter
        status = self.request.GET.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Department filter
        dept_id = self.request.GET.get('department')
        if dept_id and dept_id != 'all':
            queryset = queryset.filter(staff__department_id=dept_id)
        
        return queryset.order_by('-date', 'staff__user__first_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['departments'] = Department.objects.filter(tenant=tenant)
        context['today'] = timezone.now().date()
        context['filters'] = {
            'date': self.request.GET.get('date', ''),
            'status': self.request.GET.get('status', ''),
            'department': self.request.GET.get('department', ''),
        }
        
        return context


class AttendanceBulkCreateView(BaseTemplateView):
    template_name = 'hr/attendance/bulk_create.html'
    permission_required = 'hr.add_attendance'
    roles_required = ['admin', 'hr_manager', 'attendance_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        date = self.request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
        
        # Get all active staff for the selected date
        staff_list = Staff.objects.filter(
            tenant=tenant,
            is_active=True,
            employment_status='ACTIVE'
        ).select_related('user', 'department')
        
        context['staff_list'] = staff_list
        context['selected_date'] = date
        context['attendance_statuses'] = StaffAttendance.ATTENDANCE_STATUS_CHOICES
        
        return context
    
    def post(self, request, *args, **kwargs):
        if not request.user.has_perm('hr.add_attendance'):
            return HttpResponseForbidden("You don't have permission to mark StaffAttendance")
        
        date = request.POST.get('date')
        if not date:
            messages.error(request, "Date is required")
            return redirect('hr:attendance_mark')
        
        try:
            attendance_date = timezone.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid date format")
            return redirect('hr:attendance_mark')
        
        tenant = get_current_tenant()
        created_count = 0
        updated_count = 0
        
        for staff in Staff.objects.filter(tenant=tenant, is_active=True):
            status_key = f"status_{staff.id}"
            check_in_key = f"check_in_{staff.id}"
            check_out_key = f"check_out_{staff.id}"
            remarks_key = f"remarks_{staff.id}"
            
            status = request.POST.get(status_key)
            if status:
                # Parse check-in/out times
                check_in_str = request.POST.get(check_in_key)
                check_out_str = request.POST.get(check_out_key)
                
                check_in_time = None
                if check_in_str:
                    try:
                        check_in_time = timezone.datetime.strptime(check_in_str, '%H:%M').time()
                    except ValueError:
                        pass
                        
                check_out_time = None
                if check_out_str:
                    try:
                        check_out_time = timezone.datetime.strptime(check_out_str, '%H:%M').time()
                    except ValueError:
                        pass

                # Create or update StaffAttendance record
                attendance_record, created = StaffAttendance.objects.update_or_create(
                    tenant=tenant,
                    staff=staff,
                    date=attendance_date,
                    defaults={
                        'status': status,
                        'check_in': check_in_time,
                        'check_out': check_out_time,
                        'remarks': request.POST.get(remarks_key, ''),
                        'marked_by': request.user,
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        
        audit_log(
            user=request.user,
            action='BULK_ATTENDANCE',
            resource_type='StaffAttendance',
            details={
                'date': date,
                'created': created_count,
                'updated': updated_count
            },
            severity='INFO'
        )
        
        messages.success(
            request,
            f"StaffAttendance marked for {created_count + updated_count} staff members."
        )
        
        return redirect('hr:attendance_list')


# ==================== LEAVE VIEWS ====================

class LeaveApplicationListView(BaseListView):
    model = LeaveApplication
    template_name = 'hr/leave/application_list.html'
    context_object_name = 'leave_applications'
    permission_required = 'hr.view_leaveapplication'
    paginate_by = 20
    
    def get_queryset(self):
        tenant = get_current_tenant()
        user = self.request.user
        
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('staff', 'leave_type', 'staff__user')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Filter by leave type
        leave_type = self.request.GET.get('leave_type')
        if leave_type and leave_type != 'all':
            queryset = queryset.filter(leave_type_id=leave_type)
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date and end_date:
            try:
                start = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(
                    Q(start_date__lte=end) & Q(end_date__gte=start)
                )
            except ValueError:
                pass
        
        # HR/Admin can see all, others see only their own
        if not (user.has_perm('hr.view_all_leaves') or user.role in ['admin', 'hr_manager']):
            queryset = queryset.filter(staff__user=user)
        
        return queryset.order_by('-applied_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['leave_types'] = LeaveType.objects.filter(tenant=tenant, is_active=True)
        context['can_approve'] = self.request.user.has_perm('hr.approve_leaveapplication')
        context['filters'] = {
            'status': self.request.GET.get('status', ''),
            'leave_type': self.request.GET.get('leave_type', ''),
            'start_date': self.request.GET.get('start_date', ''),
            'end_date': self.request.GET.get('end_date', ''),
        }
        
        return context


class LeaveApplicationCreateView(BaseCreateView):
    model = LeaveApplication
    form_class = LeaveApplicationForm
    template_name = 'hr/leave/application_form.html'
    permission_required = 'hr.add_leaveapplication'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        
        # Pre-select staff for non-HR users
        if not self.request.user.has_perm('hr.apply_leave_for_others'):
            staff = get_object_or_404(
                Staff.objects.filter(tenant=get_current_tenant()),
                user=self.request.user
            )
            kwargs['initial'] = {'staff': staff}
        
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:leave_list')
    
    def form_valid(self, form):
        form.instance.applied_date = timezone.now()
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Leave application submitted successfully. Reference: {self.object.id}"
        )
        
        audit_log(
            user=self.request.user,
            action='APPLY_LEAVE',
            resource_type='Leave Application',
            resource_id=str(self.object.id),
            details={
                'staff': self.object.staff.full_name,
                'leave_type': self.object.leave_type.name,
                'dates': f"{self.object.start_date} to {self.object.end_date}"
            },
            severity='INFO'
        )
        
        return response


class LeaveApplicationApproveView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = LeaveApplication
    fields = []  # No form fields needed for approval
    permission_required = 'hr.approve_leaveapplication'
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        leave_app = self.get_object()
        
        if leave_app.status != 'PENDING':
            return JsonResponse({
                'success': False,
                'error': f'Leave application is already {leave_app.status.lower()}'
            })
        
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '')
        
        if action == 'approve':
            leave_app.approve(user=request.user, remarks=remarks)
            message = "Leave application approved successfully."
        elif action == 'reject':
            reason = request.POST.get('rejection_reason', '')
            if not reason:
                return JsonResponse({
                    'success': False,
                    'error': 'Rejection reason is required'
                })
            leave_app.reject(user=request.user, reason=reason)
            message = "Leave application rejected."
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})
        
        audit_log(
            user=request.user,
            action=f'{action.upper()}_LEAVE',
            resource_type='Leave Application',
            resource_id=str(leave_app.id),
            details={
                'staff': leave_app.staff.full_name,
                'action': action,
                'remarks': remarks
            },
            severity='INFO' if action == 'approve' else 'WARNING'
        )
        
        messages.success(request, message)
        return JsonResponse({'success': True, 'message': message})


# ==================== PAYROLL VIEWS ====================

class PayrollListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = Payroll
    template_name = 'hr/payroll/list.html'
    context_object_name = 'payrolls'
    permission_required = 'hr.view_payroll'
    roles_required = ['admin', 'hr_manager', 'accountant']
    paginate_by = 25
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('staff', 'staff__user')
        
        # Filter by month
        month = self.request.GET.get('month')
        if month:
            try:
                month_date = timezone.datetime.strptime(month, '%Y-%m').date()
                queryset = queryset.filter(salary_month__year=month_date.year,
                                         salary_month__month=month_date.month)
            except ValueError:
                pass
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Filter by department
        dept_id = self.request.GET.get('department')
        if dept_id and dept_id != 'all':
            queryset = queryset.filter(staff__department_id=dept_id)
        
        return queryset.order_by('-salary_month', 'staff__user__first_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['departments'] = Department.objects.filter(tenant=tenant)
        context['current_month'] = timezone.now().strftime('%Y-%m')
        context['filters'] = {
            'month': self.request.GET.get('month', ''),
            'status': self.request.GET.get('status', ''),
            'department': self.request.GET.get('department', ''),
        }
        
        # Summary statistics
        if self.object_list:
            context['total_amount'] = self.object_list.aggregate(
                total=Sum('net_salary')
            )['total'] or 0
            context['processed_count'] = self.object_list.filter(
                status__in=['PROCESSED', 'APPROVED', 'PAID']
            ).count()
        
        return context


class PayrollGenerateView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'hr/payroll/generate.html'
    permission_required = 'hr.generate_payroll'
    roles_required = ['admin', 'hr_manager', 'accountant']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        month = self.request.GET.get('month', timezone.now().strftime('%Y-%m'))
        
        # Get all active staff
        staff_list = Staff.objects.filter(
            tenant=tenant,
            is_active=True,
            employment_status='ACTIVE'
        ).select_related('user', 'department', 'designation')
        
        context['staff_list'] = staff_list
        context['selected_month'] = month
        
        return context
    
    def post(self, request, *args, **kwargs):
        month = request.POST.get('month')
        if not month:
            messages.error(request, "Month is required")
            return redirect('hr:payroll_generate')
        
        try:
            month_date = timezone.datetime.strptime(month, '%Y-%m').date()
        except ValueError:
            messages.error(request, "Invalid month format")
            return redirect('hr:payroll_generate')
        
        tenant = get_current_tenant()
        created_count = 0
        
        for staff in Staff.objects.filter(tenant=tenant, is_active=True, employment_status='ACTIVE'):
            # Check if payroll already exists for this month
            if Payroll.objects.filter(tenant=tenant, staff=staff, salary_month=month_date).exists():
                continue
            
            # Create payrol entry
            # defaults for required fields
            allowances = {}
            deductions = {}
            total_earnings = staff.basic_salary # Simplified initial value
            total_deductions = 0
            net_salary = total_earnings - total_deductions
            
            payroll = Payroll.objects.create(
                tenant=tenant,
                staff=staff,
                salary_month=month_date,
                pay_date=timezone.now().date(),
                basic_salary=staff.basic_salary,
                allowances=allowances,
                deductions=deductions,
                total_earnings=total_earnings,
                total_deductions=total_deductions,
                net_salary=net_salary,
                working_days=22,  # Default, should be configurable
                present_days=22,  # Default, assume full attendance for now
                status='DRAFT',
                created_by=request.user,
                updated_by=request.user,
                processed_by=request.user 
            )
            
            # Recalculate based on actual logic if needed
            payroll.calculate_salary()
            
            created_count += 1
        
        audit_log(
            user=request.user,
            action='GENERATE_PAYROLL',
            resource_type='Payroll',
            details={'month': month, 'entries_created': created_count},
            severity='INFO'
        )
        
        messages.success(
            request,
            f"Payroll generated for {created_count} staff members for {month}."
        )
        
        return redirect('hr:payroll_list')


# ==================== API VIEWS FOR AJAX ====================

class StaffAutocompleteView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'hr.view_staff'
    
    def get(self, request, *args, **kwargs):
        tenant = get_current_tenant()
        query = request.GET.get('q', '').strip()
        
        if len(query) < 2:
            return JsonResponse({'results': []})
        
        filters = (
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(employee_id__icontains=query)
        )

        staff_list = Staff.objects.filter(
            tenant=tenant,
            is_active=True
        ).filter(filters).select_related('user')[:10]

        results = []
        for staff in staff_list:
            results.append({
                'id': str(staff.id),
                'text': f"{staff.full_name} ({staff.employee_id}) - {staff.designation.title}",
                'employee_id': staff.employee_id,
                'name': staff.full_name,
                'email': staff.user.email,
                'designation': staff.designation.title,
                'department': staff.department.name
            })
        
        return JsonResponse({'results': results})


class LeaveBalanceCheckView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'hr.view_leavebalance'
    
    def get(self, request, *args, **kwargs):
        staff_id = request.GET.get('staff_id')
        leave_type_id = request.GET.get('leave_type_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not all([staff_id, leave_type_id, start_date, end_date]):
            return JsonResponse({'error': 'Missing parameters'}, status=400)
        
        try:
            staff = Staff.objects.get(id=staff_id, tenant=get_current_tenant())
            leave_type = LeaveType.objects.get(id=leave_type_id)
            start = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            end = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
        except (Staff.DoesNotExist, LeaveType.DoesNotExist, ValueError):
            return JsonResponse({'error': 'Invalid parameters'}, status=400)
        
        # Calculate total days
        total_days = (end - start).days + 1
        
        # Get leave balance for current year
        current_year = timezone.now().year
        balance, created = LeaveBalance.objects.get_or_create(
            staff=staff,
            leave_type=leave_type,
            year=current_year,
            defaults={'total_entitled': leave_type.max_days_per_year}
        )
        
        available_days = balance.available_days
        
        return JsonResponse({
            'available_days': available_days,
            'requested_days': total_days,
            'can_apply': available_days >= total_days,
            'message': f"Available: {available_days} days, Requested: {total_days} days"
        })


# ==================== EXPORT VIEWS ====================

class StaffExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'hr.export_staff'
    roles_required = ['admin', 'hr_manager']
    
    def get(self, request, *args, **kwargs):
        import csv
        from django.http import HttpResponse
        
        tenant = get_current_tenant()
        
        # Create the HttpResponse object with the appropriate CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="staff_list_{}.csv"'.format(
            timezone.now().strftime('%Y%m%d_%H%M%S')
        )
        
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow([
            'Employee ID', 'Name', 'Email', 'Phone', 'Department',
            'Designation', 'Employment Type', 'Joining Date', 'Basic Salary'
        ])
        
        # Write data
        staff_list = Staff.objects.filter(
            tenant=tenant,
            is_active=True
        ).select_related('user', 'department', 'designation')
        
        for staff in staff_list:
            writer.writerow([
                staff.employee_id,
                staff.full_name,
                staff.user.email,
                staff.personal_phone,
                staff.department.name,
                staff.designation.title,
                staff.get_employment_type_display(),
                staff.joining_date.strftime('%Y-%m-%d'),
                str(staff.basic_salary)
            ])
        
        audit_log(
            user=request.user,
            action='EXPORT_STAFF',
            resource_type='Staff',
            details={'format': 'csv', 'count': staff_list.count()},
            severity='INFO'
        )
        
        return response


# ==================== DASHBOARD WIDGETS ====================

class HRDashboardWidgetsView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'hr.view_dashboard'
    
    def get(self, request, *args, **kwargs):
        tenant = get_current_tenant()
        widget_type = request.GET.get('widget', 'StaffAttendance')
        
        data = {}
        
        if widget_type == 'StaffAttendance':
            today = timezone.now().date()
            StaffAttendance = StaffAttendance.objects.filter(
                tenant=tenant,
                date=today
            )
            
            data = {
                'present': StaffAttendance.filter(status__in=['PRESENT', 'LATE']).count(),
                'absent': StaffAttendance.filter(status='ABSENT').count(),
                'leave': StaffAttendance.filter(status='LEAVE').count(),
                'total_staff': Staff.objects.filter(tenant=tenant, is_active=True).count()
            }
        
        elif widget_type == 'leave':
            pending_leaves = LeaveApplication.objects.filter(
                tenant=tenant,
                status='PENDING'
            ).count()
            
            current_month = timezone.now().replace(day=1)
            month_leaves = LeaveApplication.objects.filter(
                tenant=tenant,
                start_date__gte=current_month,
                status='APPROVED'
            ).count()
            
            data = {
                'pending': pending_leaves,
                'month_approved': month_leaves
            }
        
        elif widget_type == 'staff_distribution':
            departments = Department.objects.filter(tenant=tenant).annotate(
                staff_count=Count('staff_members', filter=Q(staff_members__is_active=True))
            )
            
            data = {
                'labels': [dept.name for dept in departments],
                'data': [dept.staff_count for dept in departments]
            }
        
        return JsonResponse(data)

# ==================== STAFF SUB-PAGES ====================


class StaffAttendanceView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Staff-specific StaffAttendance view"""
    model = Staff
    template_name = 'hr/staff/StaffAttendance.html'
    context_object_name = 'staff'
    permission_required = 'hr.view_attendance'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        staff = self.object
        
        # Get StaffAttendance records for the last 30 days
        thirty_days_ago = timezone.now().date() - timezone.timedelta(days=30)
        context['attendance_records'] = staff.attendances.filter(
            date__gte=thirty_days_ago
        ).order_by('-date')
        
        # StaffAttendance summary
        context['present_count'] = context['attendance_records'].filter(status='PRESENT').count()
        context['absent_count'] = context['attendance_records'].filter(status='ABSENT').count()
        context['leave_count'] = context['attendance_records'].filter(status='LEAVE').count()
        context['late_count'] = context['attendance_records'].filter(status='LATE').count()
        
        # Monthly StaffAttendance
        current_month = timezone.now().date().replace(day=1)
        context['monthly_attendance'] = staff.attendances.filter(
            date__gte=current_month
        ).order_by('date')
        
        return context


class StaffLeavesView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Staff-specific leaves view"""
    model = Staff
    template_name = 'hr/staff/leaves.html'
    context_object_name = 'staff'
    permission_required = 'hr.view_leaveapplication'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        staff = self.object
        
        # Get all leave applications
        context['leave_applications'] = staff.leave_applications.all().select_related(
            'leave_type'
        ).order_by('-start_date')
        
        # Get current year leave balances
        current_year = timezone.now().year
        context['leave_balances'] = LeaveBalance.objects.filter(
            staff=staff,
            year=current_year
        ).select_related('leave_type')
        
        # Leave summary
        context['total_leaves'] = context['leave_applications'].count()
        context['approved_leaves'] = context['leave_applications'].filter(status='APPROVED').count()
        context['pending_leaves'] = context['leave_applications'].filter(status='PENDING').count()
        
        # Current month leaves
        current_month = timezone.now().date().replace(day=1)
        context['current_month_leaves'] = context['leave_applications'].filter(
            start_date__gte=current_month,
            status='APPROVED'
        )
        
        return context


class StaffDocumentsView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Staff-specific documents view"""
    model = Staff
    template_name = 'hr/staff/documents.html'
    context_object_name = 'staff'
    permission_required = 'hr.view_staffdocument'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        staff = self.object
        
        # Get all documents
        context['documents'] = staff.documents.all().order_by('-created_at')
        
        # Group by document type
        document_types = {}
        for doc in context['documents']:
            doc_type = doc.get_document_type_display()
            if doc_type not in document_types:
                document_types[doc_type] = []
            document_types[doc_type].append(doc)
        
        context['document_types'] = document_types
        
        # Document statistics
        context['total_documents'] = context['documents'].count()
        context['verified_documents'] = context['documents'].filter(is_verified=True).count()
        context['expired_documents'] = context['documents'].filter(
            expiry_date__lt=timezone.now().date()
        ).count()
        
        return context


class StaffSalaryView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Staff-specific salary view"""
    model = Staff
    template_name = 'hr/staff/salary.html'
    context_object_name = 'staff'
    permission_required = 'hr.view_payroll'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        staff = self.object
        
        # Get payroll records
        context['payroll_records'] = Payroll.objects.filter(
            staff=staff
        ).order_by('-salary_month')
        
        # Get salary structure
        try:
            context['salary_structure'] = SalaryStructure.objects.get(
                staff=staff,
                is_active=True
            )
        except SalaryStructure.DoesNotExist:
            context['salary_structure'] = None
        
        # Salary statistics
        if context['payroll_records'].exists():
            total_salary = sum([p.net_salary for p in context['payroll_records']])
            context['total_salary_paid'] = total_salary
            context['avg_monthly_salary'] = total_salary / context['payroll_records'].count()
            
            # Current month payroll
            current_month = timezone.now().date().replace(day=1)
            try:
                context['current_month_payroll'] = Payroll.objects.get(
                    staff=staff,
                    salary_month=current_month
                )
            except Payroll.DoesNotExist:
                context['current_month_payroll'] = None
        
        return context


class StaffPerformanceView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Staff-specific performance view"""
    model = Staff
    template_name = 'hr/staff/performance.html'
    context_object_name = 'staff'
    permission_required = 'hr.view_performancereview'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        staff = self.object
        
        # Get performance reviews
        context['performance_reviews'] = PerformanceReview.objects.filter(
            staff=staff
        ).select_related('reviewed_by').order_by('-review_date')
        
        # Performance statistics
        if context['performance_reviews'].exists():
            reviews = context['performance_reviews']
            context['avg_rating'] = reviews.aggregate(
                avg=Avg('overall_rating')
            )['avg']
            
            # Get latest review
            context['latest_review'] = reviews.first()
            
            # Rating distribution
            rating_counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
            for review in reviews:
                rating = round(review.overall_rating)
                if rating in rating_counts:
                    rating_counts[rating] += 1
            
            context['rating_distribution'] = rating_counts
        
        # Get training participations
        context['trainings'] = TrainingParticipation.objects.filter(
            staff=staff
        ).select_related('training').order_by('-training__start_date')
        
        # Training statistics
        context['completed_trainings'] = context['trainings'].filter(status='COMPLETED').count()
        context['certified_trainings'] = context['trainings'].filter(certificate_issued=True).count()
        
        return context


# ==================== DESIGNATION VIEWS ====================

class DesignationListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = Designation
    template_name = 'hr/designation/list.html'
    context_object_name = 'designations'
    permission_required = 'hr.view_designation'
    paginate_by = 20


class DesignationCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Designation
    form_class = DesignationForm
    template_name = 'hr/designation/form.html'
    permission_required = 'hr.add_designation'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:designation_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Designation '{form.instance.title}' created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_DESIGNATION',
            resource_type='Designation',
            resource_id=str(self.object.id),
            details={'title': self.object.title, 'code': self.object.code},
            severity='INFO'
        )
        
        return response


class DesignationDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    model = Designation
    template_name = 'hr/designation/detail.html'
    context_object_name = 'designation'
    permission_required = 'hr.view_designation'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        designation = self.object
        
        # Get staff with this designation
        context['staff_members'] = Staff.objects.filter(
            designation=designation,
            is_active=True
        ).select_related('user', 'department')
        
        # Salary statistics
        if context['staff_members'].exists():
            salaries = [s.basic_salary for s in context['staff_members']]
            context['avg_salary'] = sum(salaries) / len(salaries)
            context['min_salary'] = min(salaries)
            context['max_salary'] = max(salaries)
        
        return context


class DesignationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = Designation
    form_class = DesignationForm
    template_name = 'hr/designation/form.html'
    permission_required = 'hr.change_designation'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:designation_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Designation '{form.instance.title}' updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_DESIGNATION',
            resource_type='Designation',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class DesignationDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = Designation
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_designation'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:designation_list')
    
    def delete(self, request, *args, **kwargs):
        designation = self.get_object()
        
        # Check if designation has staff
        if designation.staff_members.filter(is_active=True).exists():
            messages.error(
                request,
                f"Cannot delete designation '{designation.title}' because it has active staff members."
            )
            return JsonResponse({'success': False, 'error': 'Designation has active staff'})
        
        # Perform soft delete
        designation.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Designation '{designation.title}' has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_DESIGNATION',
            resource_type='Designation',
            resource_id=str(designation.id),
            details={'title': designation.title, 'code': designation.code},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


# ==================== DEPARTMENT DETAIL VIEW ====================

class DepartmentDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Department detail view"""
    model = Department
    template_name = 'hr/department/detail.html'
    context_object_name = 'department'
    permission_required = 'hr.view_department'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        department = self.object
        
        # Get department staff
        context['staff_members'] = department.staff_members.filter(
            is_active=True
        ).select_related('user', 'designation')
        
        # Staff statistics by designation
        designations = Designation.objects.all()
        designation_stats = []
        for designation in designations:
            count = context['staff_members'].filter(designation=designation).count()
            if count > 0:
                designation_stats.append({
                    'designation': designation.title,
                    'count': count,
                    'percentage': (count / context['staff_members'].count()) * 100
                })
        
        context['designation_stats'] = designation_stats
        
        # Recent department activity
        context['recent_leaves'] = LeaveApplication.objects.filter(
            staff__department=department
        ).select_related('staff', 'leave_type').order_by('-applied_date')[:5]
        
        # Department performance
        department_reviews = PerformanceReview.objects.filter(
            staff__department=department
        )
        if department_reviews.exists():
            context['avg_department_rating'] = department_reviews.aggregate(
                avg=Avg('overall_rating')
            )['avg']
        
        return context

# ==================== DEPARTMENT VIEWS ====================

class DepartmentListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = Department
    template_name = 'hr/department/list.html'
    context_object_name = 'departments'
    permission_required = 'hr.view_department'


class DepartmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'hr/department/form.html'
    permission_required = 'hr.add_department'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:department_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Department '{form.instance.name}' created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_DEPARTMENT',
            resource_type='Department',
            resource_id=str(self.object.id),
            details={'name': self.object.name, 'code': self.object.code},
            severity='INFO'
        )
        
        return response


class DepartmentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'hr/department/form.html'
    permission_required = 'hr.change_department'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:department_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Department '{form.instance.name}' updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_DEPARTMENT',
            resource_type='Department',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class DepartmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = Department
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_department'
    roles_required = ['admin', 'principal']
    
    def get_success_url(self):
        return reverse_lazy('hr:department_list')
    
    def delete(self, request, *args, **kwargs):
        department = self.get_object()
        
        # Check if department has staff
        if department.staff_members.filter(is_active=True).exists():
            messages.error(
                request,
                f"Cannot delete department '{department.name}' because it has active staff members."
            )
            return JsonResponse({'success': False, 'error': 'Department has active staff'})
        
        # Perform soft delete
        department.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Department '{department.name}' has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_DEPARTMENT',
            resource_type='Department',
            resource_id=str(department.id),
            details={'name': department.name, 'code': department.code},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })
# ==================== StaffAttendance SUB-VIEWS ====================

class AttendanceDailyView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Daily StaffAttendance view"""
    template_name = 'hr/attendance/daily.html'
    permission_required = 'hr.view_attendance'
    roles_required = ['admin', 'hr_manager', 'attendance_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        date = self.request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
        
        try:
            selected_date = timezone.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
        
        # Get StaffAttendance for selected date
        attendance_records = StaffAttendance.objects.filter(
            tenant=tenant,
            date=selected_date
        ).select_related('staff', 'staff__user', 'staff__department')
        
        # Get all active staff
        all_staff = Staff.objects.filter(
            tenant=tenant,
            is_active=True,
            employment_status='ACTIVE'
        ).select_related('user', 'department')
        
        # Create a dict of StaffAttendance records for easy lookup
        attendance_dict = {record.staff_id: record for record in attendance_records}
        
        # Prepare staff list with StaffAttendance status
        staff_list = []
        for staff in all_staff:
            record = attendance_dict.get(staff.id)
            staff_list.append({
                'staff': staff,
                'StaffAttendance': record,
                'has_record': record is not None
            })
        
        # Group by department
        departments = {}
        for item in staff_list:
            dept_name = item['staff'].department.name if item['staff'].department else 'No Department'
            if dept_name not in departments:
                departments[dept_name] = []
            departments[dept_name].append(item)
        
        context['departments'] = departments
        context['selected_date'] = selected_date
        context['total_staff'] = len(staff_list)
        context['present_count'] = len([r for r in attendance_records if r.status in ['PRESENT', 'LATE']])
        context['absent_count'] = len([r for r in attendance_records if r.status == 'ABSENT'])
        context['leave_count'] = len([r for r in attendance_records if r.status == 'LEAVE'])
        
        return context


class AttendanceMonthlyView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Monthly StaffAttendance summary"""
    template_name = 'hr/StaffAttendance/monthly.html'
    permission_required = 'hr.view_attendance'
    roles_required = ['admin', 'hr_manager', 'attendance_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        month = self.request.GET.get('month', timezone.now().strftime('%Y-%m'))
        
        try:
            month_date = timezone.datetime.strptime(month, '%Y-%m').date()
        except ValueError:
            month_date = timezone.now().date().replace(day=1)
        
        # Calculate date range for the month
        import calendar
        year = month_date.year
        month_num = month_date.month
        _, last_day = calendar.monthrange(year, month_num)
        start_date = timezone.datetime(year, month_num, 1).date()
        end_date = timezone.datetime(year, month_num, last_day).date()
        
        # Get all StaffAttendance for the month
        attendance_records = StaffAttendance.objects.filter(
            tenant=tenant,
            date__range=[start_date, end_date]
        ).select_related('staff', 'staff__user', 'staff__department')
        
        # Get all active staff
        all_staff = Staff.objects.filter(
            tenant=tenant,
            is_active=True,
            employment_status='ACTIVE'
        ).select_related('user', 'department')
        
        # Prepare monthly summary
        monthly_summary = []
        for staff in all_staff:
            staff_records = attendance_records.filter(staff=staff)
            
            if staff_records.exists():
                present_days = staff_records.filter(status__in=['PRESENT', 'LATE']).count()
                absent_days = staff_records.filter(status='ABSENT').count()
                leave_days = staff_records.filter(status='LEAVE').count()
                total_days = staff_records.count()
                
                monthly_summary.append({
                    'staff': staff,
                    'present_days': present_days,
                    'absent_days': absent_days,
                    'leave_days': leave_days,
                    'total_days': total_days,
                    'attendance_percentage': (present_days / total_days * 100) if total_days > 0 else 0
                })
        
        context['monthly_summary'] = monthly_summary
        context['selected_month'] = month_date
        context['start_date'] = start_date
        context['end_date'] = end_date
        
        # Overall statistics
        if monthly_summary:
            context['avg_attendance'] = sum(item['attendance_percentage'] for item in monthly_summary) / len(monthly_summary)
            context['best_attendance'] = max(monthly_summary, key=lambda x: x['attendance_percentage'])
            context['worst_attendance'] = min(monthly_summary, key=lambda x: x['attendance_percentage'])
        
        return context


class AttendanceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = StaffAttendance
    form_class = AttendanceForm
    template_name = 'hr/attendance/form.html'
    permission_required = 'hr.change_attendance'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:attendance_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"StaffAttendance record for {form.instance.staff.full_name} on {form.instance.date} updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_ATTENDANCE',
            resource_type='StaffAttendance',
            resource_id=str(self.object.id),
            details={
                'staff': self.object.staff.full_name,
                'date': str(self.object.date),
                'status': self.object.status
            },
            severity='INFO'
        )
        
        return response


class AttendanceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = StaffAttendance
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_attendance'
    
    def get_success_url(self):
        return reverse_lazy('hr:attendance_list')
    
    def delete(self, request, *args, **kwargs):
        StaffAttendance = self.get_object()
        
        # Perform soft delete
        StaffAttendance.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"StaffAttendance record for {StaffAttendance.staff.full_name} on {StaffAttendance.date} has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_ATTENDANCE',
            resource_type='StaffAttendance',
            resource_id=str(StaffAttendance.id),
            details={
                'staff': StaffAttendance.staff.full_name,
                'date': str(StaffAttendance.date)
            },
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class AttendanceDailySummaryView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """API view for daily StaffAttendance summary"""
    permission_required = 'hr.view_attendance'
    
    def get(self, request, *args, **kwargs):
        tenant = get_current_tenant()
        date = request.GET.get('date', timezone.now().strftime('%Y-%m-%d'))
        
        try:
            selected_date = timezone.datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
        
        StaffAttendance = StaffAttendance.objects.filter(
            tenant=tenant,
            date=selected_date
        )
        
        # Group by department
        departments = Department.objects.filter(tenant=tenant)
        dept_summary = []
        
        for dept in departments:
            dept_attendance = StaffAttendance.filter(staff__department=dept)
            if dept_attendance.exists():
                present = dept_attendance.filter(status__in=['PRESENT', 'LATE']).count()
                absent = dept_attendance.filter(status='ABSENT').count()
                leave = dept_attendance.filter(status='LEAVE').count()
                total = dept.staff_members.filter(is_active=True).count()
                
                dept_summary.append({
                    'department': dept.name,
                    'present': present,
                    'absent': absent,
                    'leave': leave,
                    'total': total,
                    'present_percentage': (present / total * 100) if total > 0 else 0
                })
        
        # Overall summary
        total_staff = Staff.objects.filter(tenant=tenant, is_active=True).count()
        present = StaffAttendance.filter(status__in=['PRESENT', 'LATE']).count()
        absent = StaffAttendance.filter(status='ABSENT').count()
        leave = StaffAttendance.filter(status='LEAVE').count()
        
        return JsonResponse({
            'date': str(selected_date),
            'overall': {
                'total_staff': total_staff,
                'present': present,
                'absent': absent,
                'leave': leave,
                'present_percentage': (present / total_staff * 100) if total_staff > 0 else 0
            },
            'departments': dept_summary
        })


# ==================== LEAVE MANAGEMENT SUB-VIEWS ====================

class LeaveTypeListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = LeaveType
    template_name = 'hr/leave/type_list.html'
    context_object_name = 'leave_types'
    permission_required = 'hr.view_leavetype'
    
    def get_queryset(self):
        tenant = get_current_tenant()
        return super().get_queryset().filter(tenant=tenant).order_by('name')


class LeaveTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = LeaveType
    form_class = LeaveTypeForm
    template_name = 'hr/leave/type_form.html'
    permission_required = 'hr.add_leavetype'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:leavetype_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Leave type '{form.instance.name}' created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_LEAVETYPE',
            resource_type='Leave Type',
            resource_id=str(self.object.id),
            details={'name': self.object.name, 'code': self.object.code},
            severity='INFO'
        )
        
        return response


class LeaveTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = LeaveType
    form_class = LeaveTypeForm
    template_name = 'hr/leave/type_form.html'
    permission_required = 'hr.change_leavetype'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:leavetype_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Leave type '{form.instance.name}' updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_LEAVETYPE',
            resource_type='Leave Type',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class LeaveTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = LeaveType
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_leavetype'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:leavetype_list')
    
    def delete(self, request, *args, **kwargs):
        leave_type = self.get_object()
        
        # Check if leave type has applications
        if leave_type.applications.exists():
            messages.error(
                request,
                f"Cannot delete leave type '{leave_type.name}' because it has associated leave applications."
            )
            return JsonResponse({'success': False, 'error': 'Leave type has applications'})
        
        # Check if leave type has balances
        if leave_type.balances.exists():
            messages.error(
                request,
                f"Cannot delete leave type '{leave_type.name}' because it has associated leave balances."
            )
            return JsonResponse({'success': False, 'error': 'Leave type has balances'})
        
        # Perform soft delete
        leave_type.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Leave type '{leave_type.name}' has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_LEAVETYPE',
            resource_type='Leave Type',
            resource_id=str(leave_type.id),
            details={'name': leave_type.name, 'code': leave_type.code},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class LeaveApplicationDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Leave application detail"""
    model = LeaveApplication
    template_name = 'hr/leave/application_detail.html'
    context_object_name = 'leave_application'
    permission_required = 'hr.view_leaveapplication'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        leave_app = self.object
        
        # Add related data
        context['staff'] = leave_app.staff
        context['leave_type'] = leave_app.leave_type
        
        # Check if user can approve
        context['can_approve'] = self.request.user.has_perm('hr.approve_leaveapplication')
        
        # Get leave balance
        current_year = timezone.now().year
        try:
            context['leave_balance'] = LeaveBalance.objects.get(
                staff=leave_app.staff,
                leave_type=leave_app.leave_type,
                year=current_year
            )
        except LeaveBalance.DoesNotExist:
            context['leave_balance'] = None
        
        return context


class LeaveApplicationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = LeaveApplication
    form_class = LeaveApplicationForm
    template_name = 'hr/leave/application_form.html'
    permission_required = 'hr.change_leaveapplication'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:leave_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Leave application updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_LEAVE_APPLICATION',
            resource_type='Leave Application',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class LeaveApplicationCancelView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Cancel leave application"""
    permission_required = 'hr.change_leaveapplication'
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        leave_app = get_object_or_404(
            LeaveApplication.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        if leave_app.status not in ['PENDING', 'APPROVED']:
            return JsonResponse({
                'success': False,
                'error': f'Cannot cancel leave application with status: {leave_app.status}'
            })
        
        # Only allow cancellation if leave hasn't started yet
        if leave_app.start_date <= timezone.now().date():
            return JsonResponse({
                'success': False,
                'error': 'Cannot cancel leave that has already started'
            })
        
        # Cancel the leave application
        leave_app.status = 'CANCELLED'
        leave_app.save()
        
        messages.success(
            request,
            f"Leave application has been cancelled."
        )
        
        audit_log(
            user=request.user,
            action='CANCEL_LEAVE',
            resource_type='Leave Application',
            resource_id=str(leave_app.id),
            details={
                'staff': leave_app.staff.full_name,
                'leave_type': leave_app.leave_type.name,
                'dates': f"{leave_app.start_date} to {leave_app.end_date}"
            },
            severity='INFO'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Leave application cancelled successfully.'
        })


class LeaveApplicationDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = LeaveApplication
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_leaveapplication'
    
    def get_success_url(self):
        return reverse_lazy('hr:leave_list')
    
    def delete(self, request, *args, **kwargs):
        leave_app = self.get_object()
        
        # Only allow deletion of pending leaves
        if leave_app.status != 'PENDING':
            messages.error(
                request,
                f"Cannot delete leave application with status: {leave_app.status}"
            )
            return JsonResponse({
                'success': False,
                'error': f'Cannot delete leave application with status: {leave_app.status}'
            })
        
        # Perform soft delete
        leave_app.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Leave application has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_LEAVE_APPLICATION',
            resource_type='Leave Application',
            resource_id=str(leave_app.id),
            details={
                'staff': leave_app.staff.full_name,
                'leave_type': leave_app.leave_type.name
            },
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class LeaveBalanceListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    """Leave balance list"""
    model = LeaveBalance
    template_name = 'hr/leave/balance_list.html'
    context_object_name = 'leave_balances'
    permission_required = 'hr.view_leavebalance'
    paginate_by = 50
    
    def get_queryset(self):
        tenant = get_current_tenant()
        year = self.request.GET.get('year', timezone.now().year)
        
        queryset = super().get_queryset().filter(
            tenant=tenant,
            year=year
        ).select_related('staff', 'leave_type', 'staff__user', 'staff__department')
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by department
        dept_id = self.request.GET.get('department')
        if dept_id:
            queryset = queryset.filter(staff__department_id=dept_id)
        
        # Filter by leave type
        leave_type_id = self.request.GET.get('leave_type')
        if leave_type_id:
            queryset = queryset.filter(leave_type_id=leave_type_id)
        
        return queryset.order_by('staff__user__first_name', 'leave_type__name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['departments'] = Department.objects.filter(tenant=tenant)
        context['leave_types'] = LeaveType.objects.filter(tenant=tenant, is_active=True)
        context['staff_list'] = Staff.objects.filter(tenant=tenant, is_active=True)
        context['current_year'] = int(self.request.GET.get('year', timezone.now().year))
        context['years'] = range(timezone.now().year - 5, timezone.now().year + 2)
        
        context['filters'] = {
            'year': self.request.GET.get('year', ''),
            'staff': self.request.GET.get('staff', ''),
            'department': self.request.GET.get('department', ''),
            'leave_type': self.request.GET.get('leave_type', ''),
        }
        
        # Summary statistics
        if self.object_list:
            total_entitled = sum(b.total_entitled for b in self.object_list)
            total_used = sum(b.used_days for b in self.object_list)
            total_available = sum(b.available_days for b in self.object_list)
            
            context['summary'] = {
                'total_entitled': total_entitled,
                'total_used': total_used,
                'total_available': total_available,
                'utilization_rate': (total_used / total_entitled * 100) if total_entitled > 0 else 0
            }
        
        return context


# ==================== PAYROLL SUB-VIEWS ====================

class PayrollProcessView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Process payroll"""
    template_name = 'hr/payroll/process.html'
    permission_required = 'hr.process_payroll'
    roles_required = ['admin', 'hr_manager', 'accountant']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        month = self.request.GET.get('month', timezone.now().strftime('%Y-%m'))
        
        try:
            month_date = timezone.datetime.strptime(month, '%Y-%m').date()
        except ValueError:
            month_date = timezone.now().date().replace(day=1)
        
        # Get draft payroll for the month
        draft_payroll = Payroll.objects.filter(
            tenant=tenant,
            salary_month=month_date,
            status='DRAFT'
        ).select_related('staff', 'staff__user', 'staff__department')
        
        context['draft_payroll'] = draft_payroll
        context['selected_month'] = month_date
        context['total_amount'] = sum(p.net_salary for p in draft_payroll) if draft_payroll else 0
        context['total_records'] = draft_payroll.count()
        
        return context
    
    def post(self, request, *args, **kwargs):
        month = request.POST.get('month')
        if not month:
            return JsonResponse({'success': False, 'error': 'Month is required'})
        
        try:
            month_date = timezone.datetime.strptime(month, '%Y-%m').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid month format'})
        
        tenant = get_current_tenant()
        payroll_ids = request.POST.getlist('payroll_ids[]')
        
        if not payroll_ids:
            return JsonResponse({'success': False, 'error': 'No payroll records selected'})
        
        updated_count = 0
        total_amount = 0
        
        for payroll_id in payroll_ids:
            try:
                payroll = Payroll.objects.get(
                    id=payroll_id,
                    tenant=tenant,
                    salary_month=month_date,
                    status='DRAFT'
                )
                
                # Process the payroll
                payroll.status = 'PROCESSED'
                payroll.processed_by = request.user
                payroll.save()
                
                updated_count += 1
                total_amount += payroll.net_salary
                
            except Payroll.DoesNotExist:
                continue
        
        audit_log(
            user=request.user,
            action='PROCESS_PAYROLL',
            resource_type='Payroll',
            details={
                'month': month,
                'records_processed': updated_count,
                'total_amount': total_amount
            },
            severity='INFO'
        )
        
        messages.success(
            request,
            f"Successfully processed {updated_count} payroll records for {month} with total amount: ₹{total_amount:,.2f}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Processed {updated_count} payroll records.',
            'processed_count': updated_count,
            'total_amount': total_amount
        })


class PayrollDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Payroll detail view"""
    model = Payroll
    template_name = 'hr/payroll/detail.html'
    context_object_name = 'payroll'
    permission_required = 'hr.view_payroll'


class PayrollUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = Payroll
    form_class = PayrollForm
    template_name = 'hr/payroll/form.html'
    permission_required = 'hr.change_payroll'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:payroll_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Payroll record for {form.instance.staff.full_name} updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_PAYROLL',
            resource_type='Payroll',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class PayrollDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = Payroll
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_payroll'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:payroll_list')
    
    def delete(self, request, *args, **kwargs):
        payroll = self.get_object()
        
        # Only allow deletion of draft payroll
        if payroll.status != 'DRAFT':
            messages.error(
                request,
                f"Cannot delete payroll record with status: {payroll.status}"
            )
            return JsonResponse({
                'success': False,
                'error': f'Cannot delete payroll record with status: {payroll.status}'
            })
        
        # Perform soft delete
        payroll.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Payroll record for {payroll.staff.full_name} has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_PAYROLL',
            resource_type='Payroll',
            resource_id=str(payroll.id),
            details={
                'staff': payroll.staff.full_name,
                'month': payroll.salary_month.strftime('%B %Y')
            },
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class PayrollApproveView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Approve payroll"""
    permission_required = 'hr.approve_payroll'
    roles_required = ['admin', 'hr_manager', 'principal']
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        payroll = get_object_or_404(
            Payroll.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        if payroll.status not in ['PROCESSED', 'APPROVED']:
            return JsonResponse({
                'success': False,
                'error': f'Cannot approve payroll with status: {payroll.status}'
            })
        
        # Approve the payroll
        payroll.status = 'APPROVED'
        payroll.approved_by = request.user
        payroll.save()
        
        messages.success(
            request,
            f"Payroll for {payroll.staff.full_name} has been approved."
        )
        
        audit_log(
            user=request.user,
            action='APPROVE_PAYROLL',
            resource_type='Payroll',
            resource_id=str(payroll.id),
            details={
                'staff': payroll.staff.full_name,
                'month': payroll.salary_month.strftime('%B %Y'),
                'amount': float(payroll.net_salary)
            },
            severity='INFO'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Payroll approved successfully.'
        })


class PayrollPayslipView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """Generate payslip"""
    model = Payroll
    template_name = 'hr/payroll/payslip.html'
    context_object_name = 'payroll'
    permission_required = 'hr.view_payroll'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        payroll = self.object
        
        # Calculate deductions and earnings
        context['earnings'] = {
            'Basic Salary': payroll.basic_salary,
            'House Rent Allowance': payroll.allowances.get('hra', 0),
            'Travel Allowance': payroll.allowances.get('ta', 0),
            'Medical Allowance': payroll.allowances.get('ma', 0),
            'Special Allowance': payroll.allowances.get('sa', 0),
        }
        
        context['deductions'] = {
            'Provident Fund': payroll.deductions.get('pf', 0),
            'Professional Tax': payroll.deductions.get('pt', 0),
            'Tax Deducted at Source': payroll.deductions.get('tds', 0),
            'Other Deductions': payroll.deductions.get('other', 0),
        }
        
        # StaffAttendance summary
        context['attendance_summary'] = {
            'Working Days': payroll.working_days,
            'Present Days': payroll.present_days,
            'Leave Days': payroll.leave_days,
            'Absent Days': payroll.absent_days,
        }
        
        return context


class SalaryStructureListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = SalaryStructure
    template_name = 'hr/payroll/structure_list.html'
    context_object_name = 'salary_structures'
    permission_required = 'hr.view_salarystructure'
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant,
            is_active=True
        ).select_related('staff', 'staff__user')
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by effective date
        effective_from = self.request.GET.get('effective_from')
        if effective_from:
            try:
                date_obj = timezone.datetime.strptime(effective_from, '%Y-%m-%d').date()
                queryset = queryset.filter(effective_from__gte=date_obj)
            except ValueError:
                pass
        
        return queryset.order_by('-effective_from')


class SalaryStructureCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = SalaryStructure
    form_class = SalaryStructureForm
    template_name = 'hr/payroll/structure_form.html'
    permission_required = 'hr.add_salarystructure'
    roles_required = ['admin', 'hr_manager', 'accountant']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:salarystructure_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Salary structure for {form.instance.staff.full_name} created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_SALARY_STRUCTURE',
            resource_type='Salary Structure',
            resource_id=str(self.object.id),
            details={'staff': self.object.staff.full_name},
            severity='INFO'
        )
        
        return response


class SalaryStructureUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = SalaryStructure
    form_class = SalaryStructureForm
    template_name = 'hr/payroll/structure_form.html'
    permission_required = 'hr.change_salarystructure'
    roles_required = ['admin', 'hr_manager', 'accountant']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:salarystructure_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Salary structure for {form.instance.staff.full_name} updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_SALARY_STRUCTURE',
            resource_type='Salary Structure',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class SalaryStructureDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = SalaryStructure
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_salarystructure'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:salarystructure_list')
    
    def delete(self, request, *args, **kwargs):
        structure = self.get_object()
        
        # Perform soft delete
        structure.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Salary structure for {structure.staff.full_name} has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_SALARY_STRUCTURE',
            resource_type='Salary Structure',
            resource_id=str(structure.id),
            details={'staff': structure.staff.full_name},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class PayrollMonthlyReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Monthly payroll report"""
    template_name = 'hr/payroll/report_monthly.html'
    permission_required = 'hr.view_payroll_report'
    roles_required = ['admin', 'hr_manager', 'accountant']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        month = self.request.GET.get('month', timezone.now().strftime('%Y-%m'))
        
        try:
            month_date = timezone.datetime.strptime(month, '%Y-%m').date()
        except ValueError:
            month_date = timezone.now().date().replace(day=1)
        
        # Get payroll for the month
        payroll_records = Payroll.objects.filter(
            tenant=tenant,
            salary_month=month_date,
            status='PAID'
        ).select_related('staff', 'staff__department')
        
        # Group by department
        departments = {}
        for payroll in payroll_records:
            dept_name = payroll.staff.department.name if payroll.staff.department else 'No Department'
            if dept_name not in departments:
                departments[dept_name] = {
                    'records': [],
                    'total_salary': 0,
                    'count': 0
                }
            
            departments[dept_name]['records'].append(payroll)
            departments[dept_name]['total_salary'] += float(payroll.net_salary)
            departments[dept_name]['count'] += 1
        
        context['departments'] = departments
        context['selected_month'] = month_date
        context['total_records'] = payroll_records.count()
        context['total_salary'] = sum(float(p.net_salary) for p in payroll_records)
        
        # Payment method distribution
        payment_methods = {}
        for payroll in payroll_records:
            method = payroll.payment_method
            if method not in payment_methods:
                payment_methods[method] = 0
            payment_methods[method] += 1
        
        context['payment_methods'] = payment_methods
        
        return context


class PayrollAnnualReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Annual payroll report"""
    template_name = 'hr/payroll/report_annual.html'
    permission_required = 'hr.view_payroll_report'
    roles_required = ['admin', 'hr_manager', 'accountant']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        year = int(self.request.GET.get('year', timezone.now().year))
        
        # Get payroll for the year
        payroll_records = Payroll.objects.filter(
            tenant=tenant,
            salary_month__year=year,
            status='PAID'
        ).select_related('staff', 'staff__department')
        
        # Monthly breakdown
        monthly_data = {}
        for month in range(1, 13):
            month_payroll = payroll_records.filter(salary_month__month=month)
            if month_payroll.exists():
                total_salary = sum(float(p.net_salary) for p in month_payroll)
                monthly_data[month] = {
                    'count': month_payroll.count(),
                    'total_salary': total_salary,
                    'avg_salary': total_salary / month_payroll.count() if month_payroll.count() > 0 else 0
                }
        
        context['monthly_data'] = monthly_data
        context['selected_year'] = year
        context['total_records'] = payroll_records.count()
        context['total_salary'] = sum(float(p.net_salary) for p in payroll_records)
        context['avg_monthly_salary'] = context['total_salary'] / 12 if payroll_records.count() > 0 else 0
        
        # Department breakdown
        departments = Department.objects.filter(tenant=tenant)
        dept_data = []
        for dept in departments:
            dept_payroll = payroll_records.filter(staff__department=dept)
            if dept_payroll.exists():
                total_salary = sum(float(p.net_salary) for p in dept_payroll)
                dept_data.append({
                    'department': dept.name,
                    'count': dept_payroll.count(),
                    'total_salary': total_salary,
                    'percentage': (total_salary / context['total_salary'] * 100) if context['total_salary'] > 0 else 0
                })
        
        context['department_data'] = dept_data
        
        return context


class PayrollDepartmentReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Department-wise payroll report"""
    template_name = 'hr/payroll/report_department.html'
    permission_required = 'hr.view_payroll_report'
    roles_required = ['admin', 'hr_manager', 'accountant']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        dept_id = self.request.GET.get('department')
        year = int(self.request.GET.get('year', timezone.now().year))
        
        # Get department
        department = None
        if dept_id:
            try:
                department = Department.objects.get(id=dept_id, tenant=tenant)
            except Department.DoesNotExist:
                pass
        
        if not department:
            # Get all departments
            departments = Department.objects.filter(tenant=tenant)
            department_data = []
            
            for dept in departments:
                payroll_records = Payroll.objects.filter(
                    tenant=tenant,
                    staff__department=dept,
                    salary_month__year=year,
                    status='PAID'
                )
                
                if payroll_records.exists():
                    total_salary = sum(float(p.net_salary) for p in payroll_records)
                    department_data.append({
                        'department': dept,
                        'record_count': payroll_records.count(),
                        'total_salary': total_salary,
                        'avg_salary': total_salary / payroll_records.count() if payroll_records.count() > 0 else 0
                    })
            
            context['department_data'] = department_data
            context['selected_year'] = year
            return context
        
        # Get department-specific data
        payroll_records = Payroll.objects.filter(
            tenant=tenant,
            staff__department=department,
            salary_month__year=year,
            status='PAID'
        ).select_related('staff', 'staff__user')
        
        # Monthly breakdown for the department
        monthly_data = {}
        for month in range(1, 13):
            month_payroll = payroll_records.filter(salary_month__month=month)
            if month_payroll.exists():
                total_salary = sum(float(p.net_salary) for p in month_payroll)
                monthly_data[month] = {
                    'count': month_payroll.count(),
                    'total_salary': total_salary,
                    'avg_salary': total_salary / month_payroll.count() if month_payroll.count() > 0 else 0
                }
        
        # Staff-wise breakdown
        staff_data = []
        staff_ids = payroll_records.values_list('staff_id', flat=True).distinct()
        
        for staff_id in staff_ids:
            staff_payroll = payroll_records.filter(staff_id=staff_id)
            if staff_payroll.exists():
                total_salary = sum(float(p.net_salary) for p in staff_payroll)
                staff = staff_payroll.first().staff
                staff_data.append({
                    'staff': staff,
                    'record_count': staff_payroll.count(),
                    'total_salary': total_salary,
                    'avg_salary': total_salary / staff_payroll.count() if staff_payroll.count() > 0 else 0
                })
        
        context['department'] = department
        context['monthly_data'] = monthly_data
        context['staff_data'] = staff_data
        context['selected_year'] = year
        context['total_records'] = payroll_records.count()
        context['total_salary'] = sum(float(p.net_salary) for p in payroll_records)
        
        return context


class PayrollSummaryView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """API view for payroll summary"""
    permission_required = 'hr.view_payroll'
    
    def get(self, request, *args, **kwargs):
        tenant = get_current_tenant()
        month = request.GET.get('month', timezone.now().strftime('%Y-%m'))
        
        try:
            month_date = timezone.datetime.strptime(month, '%Y-%m').date()
        except ValueError:
            month_date = timezone.now().date().replace(day=1)
        
        payroll_records = Payroll.objects.filter(
            tenant=tenant,
            salary_month=month_date
        ).select_related('staff__department')
        
        # Status distribution
        status_counts = {}
        total_amount = 0
        
        for record in payroll_records:
            status = record.status
            if status not in status_counts:
                status_counts[status] = {'count': 0, 'amount': 0}
            
            status_counts[status]['count'] += 1
            status_counts[status]['amount'] += float(record.net_salary)
            total_amount += float(record.net_salary)
        
        # Department distribution
        dept_data = []
        departments = Department.objects.filter(tenant=tenant)
        
        for dept in departments:
            dept_payroll = payroll_records.filter(staff__department=dept)
            if dept_payroll.exists():
                dept_amount = sum(float(p.net_salary) for p in dept_payroll)
                dept_data.append({
                    'department': dept.name,
                    'count': dept_payroll.count(),
                    'amount': dept_amount,
                    'percentage': (dept_amount / total_amount * 100) if total_amount > 0 else 0
                })
        
        return JsonResponse({
            'month': str(month_date),
            'total_records': payroll_records.count(),
            'total_amount': total_amount,
            'status_distribution': status_counts,
            'department_distribution': dept_data
        })


# ==================== RECRUITMENT VIEWS ====================

class RecruitmentListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = Recruitment
    template_name = 'hr/recruitment/list.html'
    context_object_name = 'recruitments'
    permission_required = 'hr.view_recruitment'
    paginate_by = 20
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('department', 'designation', 'hiring_manager')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Filter by department
        dept_id = self.request.GET.get('department')
        if dept_id:
            queryset = queryset.filter(department_id=dept_id)
        
        # Filter by employment type
        emp_type = self.request.GET.get('employment_type')
        if emp_type:
            queryset = queryset.filter(employment_type=emp_type)
        
        # Search
        search = self.request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(position_title__icontains=search) |
                Q(job_description__icontains=search) |
                Q(department__name__icontains=search)
            )
        
        return queryset.order_by('-posting_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['departments'] = Department.objects.filter(tenant=tenant)
        context['status_choices'] = Recruitment.POSITION_STATUS_CHOICES
        context['employment_types'] = Staff.EMPLOYMENT_TYPE_CHOICES
        
        context['filters'] = {
            'status': self.request.GET.get('status', ''),
            'department': self.request.GET.get('department', ''),
            'employment_type': self.request.GET.get('employment_type', ''),
            'search': self.request.GET.get('search', ''),
        }
        
        # Statistics
        if self.object_list:
            context['total_positions'] = sum(r.no_of_openings for r in self.object_list)
            context['open_positions'] = sum(r.no_of_openings for r in self.object_list if r.is_open)
        
        return context


class RecruitmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Recruitment
    form_class = RecruitmentForm
    template_name = 'hr/recruitment/form.html'
    permission_required = 'hr.add_recruitment'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        
        # Set initial hiring manager
        kwargs['initial'] = {'hiring_manager': self.request.user}
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:recruitment_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        form.instance.hiring_manager = self.request.user
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Recruitment position '{form.instance.position_title}' created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_RECRUITMENT',
            resource_type='Recruitment',
            resource_id=str(self.object.id),
            details={'position': self.object.position_title, 'openings': self.object.no_of_openings},
            severity='INFO'
        )
        
        return response


class RecruitmentDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    model = Recruitment
    template_name = 'hr/recruitment/detail.html'
    context_object_name = 'recruitment'
    permission_required = 'hr.view_recruitment'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        recruitment = self.object
        
        # Get applications for this recruitment
        context['applications'] = recruitment.applications.all().order_by('-applied_date')
        
        # Application statistics
        apps = context['applications']
        context['total_applications'] = apps.count()
        context['shortlisted'] = apps.filter(status='SHORTLISTED').count()
        context['interview'] = apps.filter(status='INTERVIEW').count()
        context['offered'] = apps.filter(status='OFFERED').count()
        context['accepted'] = apps.filter(status='ACCEPTED').count()
        
        # Check if user can manage applications
        context['can_manage'] = self.request.user.has_perm('hr.manage_recruitment')
        
        return context


class RecruitmentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = Recruitment
    form_class = RecruitmentForm
    template_name = 'hr/recruitment/form.html'
    permission_required = 'hr.change_recruitment'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:recruitment_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Recruitment position '{form.instance.position_title}' updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_RECRUITMENT',
            resource_type='Recruitment',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class RecruitmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = Recruitment
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_recruitment'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:recruitment_list')
    
    def delete(self, request, *args, **kwargs):
        recruitment = self.get_object()
        
        # Check if recruitment has applications
        if recruitment.applications.exists():
            messages.error(
                request,
                f"Cannot delete recruitment position '{recruitment.position_title}' because it has applications."
            )
            return JsonResponse({'success': False, 'error': 'Recruitment has applications'})
        
        # Perform soft delete
        recruitment.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Recruitment position '{recruitment.position_title}' has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_RECRUITMENT',
            resource_type='Recruitment',
            resource_id=str(recruitment.id),
            details={'position': recruitment.position_title},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class RecruitmentCloseView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Close recruitment position"""
    permission_required = 'hr.change_recruitment'
    roles_required = ['admin', 'hr_manager']
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        recruitment = get_object_or_404(
            Recruitment.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        if recruitment.status in ['FILLED', 'CANCELLED', 'ON_HOLD']:
            return JsonResponse({
                'success': False,
                'error': f'Recruitment is already {recruitment.status.lower()}'
            })
        
        action = request.POST.get('action')
        reason = request.POST.get('reason', '')
        
        if action == 'close':
            recruitment.status = 'FILLED'
            message = "Recruitment position marked as filled."
        elif action == 'cancel':
            recruitment.status = 'CANCELLED'
            message = "Recruitment position cancelled."
        elif action == 'hold':
            recruitment.status = 'ON_HOLD'
            message = "Recruitment position put on hold."
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})
        
        recruitment.save()
        
        audit_log(
            user=request.user,
            action=f'{action.upper()}_RECRUITMENT',
            resource_type='Recruitment',
            resource_id=str(recruitment.id),
            details={
                'position': recruitment.position_title,
                'action': action,
                'reason': reason
            },
            severity='INFO'
        )
        
        messages.success(request, message)
        return JsonResponse({'success': True, 'message': message})


# ==================== JOB APPLICATION VIEWS ====================

class JobApplicationListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = JobApplication
    template_name = 'hr/recruitment/application_list.html'
    context_object_name = 'job_applications'
    permission_required = 'hr.view_jobapplication'
    paginate_by = 20
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('recruitment', 'recruitment__department')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Filter by recruitment
        recruitment_id = self.request.GET.get('recruitment')
        if recruitment_id:
            queryset = queryset.filter(recruitment_id=recruitment_id)
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date and end_date:
            try:
                start = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(applied_date__range=[start, end])
            except ValueError:
                pass
        
        # Search
        search = self.request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(applicant_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        
        return queryset.order_by('-applied_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['recruitments'] = Recruitment.objects.filter(tenant=tenant, status='OPEN')
        context['status_choices'] = JobApplication.APPLICATION_STATUS_CHOICES
        
        context['filters'] = {
            'status': self.request.GET.get('status', ''),
            'recruitment': self.request.GET.get('recruitment', ''),
            'start_date': self.request.GET.get('start_date', ''),
            'end_date': self.request.GET.get('end_date', ''),
            'search': self.request.GET.get('search', ''),
        }
        
        return context


class JobApplicationCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = JobApplication
    form_class = PublicJobApplicationForm
    template_name = 'public/recruitment/application_form.html'
    permission_required = 'hr.add_jobapplication'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:jobapplication_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.applied_date = timezone.now()
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Job application from {form.instance.applicant_name} submitted successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_JOB_APPLICATION',
            resource_type='Job Application',
            resource_id=str(self.object.id),
            details={'applicant': self.object.applicant_name, 'position': self.object.recruitment.position_title},
            severity='INFO'
        )
        
        return response


class JobApplicationDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    model = JobApplication
    template_name = 'hr/recruitment/application_detail.html'
    context_object_name = 'job_application'
    permission_required = 'hr.view_jobapplication'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        job_application = self.get_object()
        context['page_title'] = job_application.applicant_name
        context['parent_text'] = _("Recruitment")
        context['parent_url'] = reverse_lazy('hr:jobapplication_list')
        context['current_text'] = job_application.applicant_name
        return context


class JobApplicationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = JobApplication
    form_class = JobApplicationForm
    template_name = 'hr/recruitment/application_form.html'
    permission_required = 'hr.change_jobapplication'
    roles_required = ['admin', 'hr_manager']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        job_application = self.get_object()
        context['page_title'] = _("Edit Application")
        context['parent_text'] = _("Recruitment")
        context['parent_url'] = reverse_lazy('hr:jobapplication_list')
        context['current_text'] = job_application.applicant_name
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:jobapplication_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Job application from {form.instance.applicant_name} updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_JOB_APPLICATION',
            resource_type='Job Application',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class JobApplicationDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = JobApplication
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_jobapplication'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:jobapplication_list')
    
    def delete(self, request, *args, **kwargs):
        application = self.get_object()
        
        # Perform soft delete
        application.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Job application from {application.applicant_name} has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_JOB_APPLICATION',
            resource_type='Job Application',
            resource_id=str(application.id),
            details={'applicant': application.applicant_name},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class JobApplicationShortlistView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Shortlist job application"""
    permission_required = 'hr.manage_recruitment'
    roles_required = ['admin', 'hr_manager']
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        application = get_object_or_404(
            JobApplication.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        if application.status != 'APPLIED':
            messages.error(request, f"Cannot shortlist application with status: {application.get_status_display()}")
            return redirect('hr:jobapplication_detail', pk=application.pk)
        
        # Shortlist the application
        application.status = 'SHORTLISTED'
        application.save()
        
        audit_log(
            user=request.user,
            action='SHORTLIST_APPLICATION',
            resource_type='Job Application',
            resource_id=str(application.id),
            details={
                'applicant': application.applicant_name,
                'position': application.recruitment.position_title
            },
            severity='INFO'
        )
        
        messages.success(request, "Application shortlisted successfully.")
        return redirect('hr:jobapplication_detail', pk=application.pk)


class JobApplicationInterviewView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Schedule interview for job application"""
    permission_required = 'hr.manage_recruitment'
    roles_required = ['admin', 'hr_manager']
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        application = get_object_or_404(
            JobApplication.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        if application.status != 'SHORTLISTED':
            messages.error(request, f"Cannot schedule interview for application with status: {application.get_status_display()}")
            return redirect('hr:jobapplication_detail', pk=application.pk)
        
        # Update status to interview
        application.status = 'INTERVIEW'
        application.save()
        
        audit_log(
            user=request.user,
            action='SCHEDULE_INTERVIEW',
            resource_type='Job Application',
            resource_id=str(application.id),
            details={
                'applicant': application.applicant_name,
                'position': application.recruitment.position_title
            },
            severity='INFO'
        )
        
        messages.success(request, "Interview scheduled successfully.")
        return redirect('hr:jobapplication_detail', pk=application.pk)


class JobApplicationRejectView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Reject job application"""
    permission_required = 'hr.manage_recruitment'
    roles_required = ['admin', 'hr_manager']
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        application = get_object_or_404(
            JobApplication.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        # Reject the application
        application.status = 'REJECTED'
        application.save()
        
        audit_log(
            user=request.user,
            action='REJECT_APPLICATION',
            resource_type='Job Application',
            resource_id=str(application.id),
            details={
                'applicant': application.applicant_name,
                'position': application.recruitment.position_title
            },
            severity='INFO'
        )
        
        messages.success(request, "Application rejected.")
        return redirect('hr:jobapplication_detail', pk=application.pk)


class JobApplicationOfferView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Make job offer to applicant"""
    permission_required = 'hr.manage_recruitment'
    roles_required = ['admin', 'hr_manager']
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        application = get_object_or_404(
            JobApplication.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        if application.status != 'INTERVIEW':
            return JsonResponse({
                'success': False,
                'error': f'Cannot make offer to application with status: {application.status}'
            })
        
        # Make offer
        application.status = 'OFFERED'
        application.save()
        
        audit_log(
            user=request.user,
            action='MAKE_JOB_OFFER',
            resource_type='Job Application',
            resource_id=str(application.id),
            details={
                'applicant': application.applicant_name,
                'position': application.recruitment.position_title
            },
            severity='INFO'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Job offer made successfully.'
        })


# ==================== TRAINING & DEVELOPMENT VIEWS ====================

class TrainingProgramListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = TrainingProgram
    template_name = 'hr/training/program_list.html'
    context_object_name = 'training_programs'
    permission_required = 'hr.view_trainingprogram'
    paginate_by = 20
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(tenant=tenant)
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Filter by training type
        training_type = self.request.GET.get('training_type')
        if training_type and training_type != 'all':
            queryset = queryset.filter(training_type=training_type)
        
        # Filter by mandatory
        is_mandatory = self.request.GET.get('is_mandatory')
        if is_mandatory and is_mandatory != 'all':
            queryset = queryset.filter(is_mandatory=(is_mandatory == 'true'))
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date and end_date:
            try:
                start = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(
                    Q(start_date__lte=end) & Q(end_date__gte=start)
                )
            except ValueError:
                pass
        
        # Search
        search = self.request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(organizer__icontains=search) |
                Q(venue__icontains=search)
            )
        
        return queryset.order_by('-start_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context['training_types'] = TrainingProgram.TRAINING_TYPE_CHOICES
        context['status_choices'] = [
            ('UPCOMING', 'Upcoming'),
            ('ONGOING', 'Ongoing'),
            ('COMPLETED', 'Completed'),
            ('CANCELLED', 'Cancelled')
        ]
        
        context['filters'] = {
            'status': self.request.GET.get('status', ''),
            'training_type': self.request.GET.get('training_type', ''),
            'is_mandatory': self.request.GET.get('is_mandatory', ''),
            'start_date': self.request.GET.get('start_date', ''),
            'end_date': self.request.GET.get('end_date', ''),
            'search': self.request.GET.get('search', ''),
        }
        
        return context


class TrainingProgramCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = TrainingProgram
    form_class = TrainingProgramForm
    template_name = 'hr/training/program_form.html'
    permission_required = 'hr.add_trainingprogram'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:trainingprogram_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Training program '{form.instance.title}' created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_TRAINING_PROGRAM',
            resource_type='Training Program',
            resource_id=str(self.object.id),
            details={'title': self.object.title, 'type': self.object.training_type},
            severity='INFO'
        )
        
        return response


class TrainingProgramDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    model = TrainingProgram
    template_name = 'hr/training/program_detail.html'
    context_object_name = 'training_program'
    permission_required = 'hr.view_trainingprogram'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        training = self.object
        
        # Get participants
        context['participants'] = training.participants.select_related(
            'staff', 'staff__user'
        ).all()
        
        # Statistics
        participants = context['participants']
        context['total_participants'] = participants.count()
        context['completed'] = participants.filter(status='COMPLETED').count()
        context['attended'] = participants.filter(status='ATTENDED').count()
        context['registered'] = participants.filter(status='REGISTERED').count()
        context['available_slots'] = training.available_slots
        
        # Check if user can register
        context['can_register'] = self.request.user.has_perm('hr.register_training')
        
        return context


class TrainingProgramUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = TrainingProgram
    form_class = TrainingProgramForm
    template_name = 'hr/training/program_form.html'
    permission_required = 'hr.change_trainingprogram'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:trainingprogram_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Training program '{form.instance.title}' updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_TRAINING_PROGRAM',
            resource_type='Training Program',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class TrainingProgramDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = TrainingProgram
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_trainingprogram'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:trainingprogram_list')
    
    def delete(self, request, *args, **kwargs):
        training = self.get_object()
        
        # Check if training has participants
        if training.participants.exists():
            messages.error(
                request,
                f"Cannot delete training program '{training.title}' because it has participants."
            )
            return JsonResponse({'success': False, 'error': 'Training program has participants'})
        
        # Perform soft delete
        training.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Training program '{training.title}' has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_TRAINING_PROGRAM',
            resource_type='Training Program',
            resource_id=str(training.id),
            details={'title': training.title},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class TrainingRegisterView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Register for training program"""
    permission_required = 'hr.register_training'
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        training = get_object_or_404(
            TrainingProgram.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        # Check if training is open for registration
        if training.status != 'UPCOMING':
            return JsonResponse({
                'success': False,
                'error': f'Cannot register for training with status: {training.status}'
            })
        
        # Check if there are available slots
        if training.available_slots <= 0:
            return JsonResponse({
                'success': False,
                'error': 'No available slots for this training program.'
            })
        
        # Get staff member for the current user
        try:
            staff = Staff.objects.get(
                tenant=get_current_tenant(),
                user=request.user
            )
        except Staff.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Staff profile not found.'
            })
        
        # Check if already registered
        if TrainingParticipation.objects.filter(
            training=training,
            staff=staff
        ).exists():
            return JsonResponse({
                'success': False,
                'error': 'You are already registered for this training.'
            })
        
        # Register for training
        TrainingParticipation.objects.create(
            tenant=get_current_tenant(),
            training=training,
            staff=staff,
            status='REGISTERED',
            created_by=request.user,
            updated_by=request.user
        )
        
        audit_log(
            user=request.user,
            action='REGISTER_TRAINING',
            resource_type='Training Program',
            resource_id=str(training.id),
            details={
                'training': training.title,
                'staff': staff.full_name
            },
            severity='INFO'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Successfully registered for the training program.'
        })


class TrainingParticipantsView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """View training participants"""
    model = TrainingProgram
    template_name = 'hr/training/participants.html'
    context_object_name = 'training_program'
    permission_required = 'hr.view_trainingprogram'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        training = self.object
        
        # Get all participants
        participants = training.participants.select_related(
            'staff', 'staff__user', 'staff__department'
        ).all()
        
        # Group by status
        status_groups = {}
        for participant in participants:
            status = participant.status
            if status not in status_groups:
                status_groups[status] = []
            status_groups[status].append(participant)
        
        context['status_groups'] = status_groups
        context['total_participants'] = participants.count()
        
        return context


# ==================== TRAINING PARTICIPATION VIEWS ====================

class TrainingParticipationListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = TrainingParticipation
    template_name = 'hr/training/participation_list.html'
    context_object_name = 'training_participations'
    permission_required = 'hr.view_trainingparticipation'
    paginate_by = 50
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('training', 'staff', 'staff__user')
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Filter by training
        training_id = self.request.GET.get('training')
        if training_id:
            queryset = queryset.filter(training_id=training_id)
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by certificate issued
        certificate_issued = self.request.GET.get('certificate_issued')
        if certificate_issued and certificate_issued != 'all':
            queryset = queryset.filter(certificate_issued=(certificate_issued == 'true'))
        
        return queryset.order_by('-training__start_date')


class TrainingParticipationCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = TrainingParticipation
    form_class = TrainingParticipationForm
    template_name = 'hr/training/participation_form.html'
    permission_required = 'hr.add_trainingparticipation'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:trainingparticipation_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Training participation record created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_TRAINING_PARTICIPATION',
            resource_type='Training Participation',
            resource_id=str(self.object.id),
            details={
                'training': self.object.training.title,
                'staff': self.object.staff.full_name
            },
            severity='INFO'
        )
        
        return response


class TrainingParticipationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = TrainingParticipation
    form_class = TrainingParticipationForm
    template_name = 'hr/training/participation_form.html'
    permission_required = 'hr.change_trainingparticipation'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:trainingparticipation_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Training participation record updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_TRAINING_PARTICIPATION',
            resource_type='Training Participation',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class TrainingParticipationDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = TrainingParticipation
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_trainingparticipation'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:trainingparticipation_list')
    
    def delete(self, request, *args, **kwargs):
        participation = self.get_object()
        
        # Perform soft delete
        participation.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Training participation record has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_TRAINING_PARTICIPATION',
            resource_type='Training Participation',
            resource_id=str(participation.id),
            details={
                'training': participation.training.title,
                'staff': participation.staff.full_name
            },
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class TrainingCertificateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """View training certificate"""
    model = TrainingParticipation
    template_name = 'hr/training/certificate.html'
    context_object_name = 'participation'
    permission_required = 'hr.view_trainingparticipation'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        participation = self.object
        
        # Check if certificate is issued
        if not participation.certificate_issued:
            raise Http404("Certificate not issued for this participation")
        
        return context


# ==================== PERFORMANCE MANAGEMENT VIEWS ====================

class PerformanceReviewListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = PerformanceReview
    template_name = 'hr/performance/review_list.html'
    context_object_name = 'performance_reviews'
    permission_required = 'hr.view_performancereview'
    paginate_by = 20
    
    def get_queryset(self):
        tenant = get_current_tenant()
        user = self.request.user
        
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('staff', 'reviewed_by', 'staff__user', 'staff__department')
        
        # Filter by review type
        review_type = self.request.GET.get('review_type')
        if review_type and review_type != 'all':
            queryset = queryset.filter(review_type=review_type)
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by department
        dept_id = self.request.GET.get('department')
        if dept_id:
            queryset = queryset.filter(staff__department_id=dept_id)
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date and end_date:
            try:
                start = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(review_date__range=[start, end])
            except ValueError:
                pass
        
        # Filter by rating
        min_rating = self.request.GET.get('min_rating')
        if min_rating:
            try:
                min_rating = float(min_rating)
                queryset = queryset.filter(overall_rating__gte=min_rating)
            except ValueError:
                pass
        
        # HR/Admin can see all, others see only their own or reviews they conducted
        if not (user.has_perm('hr.view_all_performance') or user.role in ['admin', 'hr_manager']):
            queryset = queryset.filter(
                Q(staff__user=user) | Q(reviewed_by=user)
            )
        
        return queryset.order_by('-review_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['departments'] = Department.objects.filter(tenant=tenant)
        context['staff_list'] = Staff.objects.filter(tenant=tenant, is_active=True)
        context['review_types'] = PerformanceReview.REVIEW_TYPE_CHOICES
        
        context['filters'] = {
            'review_type': self.request.GET.get('review_type', ''),
            'staff': self.request.GET.get('staff', ''),
            'department': self.request.GET.get('department', ''),
            'start_date': self.request.GET.get('start_date', ''),
            'end_date': self.request.GET.get('end_date', ''),
            'min_rating': self.request.GET.get('min_rating', ''),
        }
        
        return context


class PerformanceReviewCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = PerformanceReview
    form_class = PerformanceReviewForm
    template_name = 'hr/performance/review_form.html'
    permission_required = 'hr.add_performancereview'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:performancereview_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.reviewed_by = self.request.user
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Performance review for {form.instance.staff.full_name} created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_PERFORMANCE_REVIEW',
            resource_type='Performance Review',
            resource_id=str(self.object.id),
            details={'staff': self.object.staff.full_name, 'rating': self.object.overall_rating},
            severity='INFO'
        )
        
        return response


class PerformanceReviewDetailView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    model = PerformanceReview
    template_name = 'hr/performance/review_detail.html'
    context_object_name = 'performance_review'
    permission_required = 'hr.view_performancereview'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        review = self.object
        
        # Check if user can acknowledge
        context['can_acknowledge'] = (
            self.request.user == review.staff.user and 
            not review.employee_acknowledged
        )
        
        # Check if user can approve
        context['can_approve'] = (
            self.request.user.has_perm('hr.approve_performancereview') and
            review.approved_by is None
        )
        
        context['page_title'] = _("Performance Review - ") + review.staff.full_name
        context['parent_text'] = _("Performance Reviews")
        context['parent_url'] = reverse_lazy('hr:performancereview_list')
        context['current_text'] = review.staff.full_name
        
        return context


class PerformanceReviewUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = PerformanceReview
    form_class = PerformanceReviewForm
    template_name = 'hr/performance/review_form.html'
    permission_required = 'hr.change_performancereview'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:performancereview_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Performance review for {form.instance.staff.full_name} updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_PERFORMANCE_REVIEW',
            resource_type='Performance Review',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class PerformanceReviewDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = PerformanceReview
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_performancereview'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:performancereview_list')
    
    def delete(self, request, *args, **kwargs):
        review = self.get_object()
        
        # Perform soft delete
        review.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Performance review for {review.staff.full_name} has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_PERFORMANCE_REVIEW',
            resource_type='Performance Review',
            resource_id=str(review.id),
            details={'staff': review.staff.full_name},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class PerformanceAcknowledgeView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Acknowledge performance review"""
    permission_required = 'hr.acknowledge_performancereview'
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        review = get_object_or_404(
            PerformanceReview.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        # Check if user is the staff member
        if review.staff.user != request.user:
            return JsonResponse({
                'success': False,
                'error': 'You can only acknowledge your own performance review.'
            })
        
        # Check if already acknowledged
        if review.employee_acknowledged:
            return JsonResponse({
                'success': False,
                'error': 'Performance review already acknowledged.'
            })
        
        # Acknowledge the review
        review.employee_acknowledged = True
        review.acknowledgement_date = timezone.now()
        review.save()
        
        audit_log(
            user=request.user,
            action='ACKNOWLEDGE_PERFORMANCE_REVIEW',
            resource_type='Performance Review',
            resource_id=str(review.id),
            details={
                'staff': review.staff.full_name,
                'rating': review.overall_rating
            },
            severity='INFO'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Performance review acknowledged successfully.'
        })


# ==================== EMPLOYMENT HISTORY VIEWS ====================

class EmploymentHistoryListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = EmploymentHistory
    template_name = 'hr/history/list.html'
    context_object_name = 'employment_history'
    permission_required = 'hr.view_employmenthistory'
    paginate_by = 50
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('staff', 'staff__user')
        
        # Filter by action
        action = self.request.GET.get('action')
        if action and action != 'all':
            queryset = queryset.filter(action=action)
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date and end_date:
            try:
                start = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(effective_date__range=[start, end])
            except ValueError:
                pass
        
        return queryset.order_by('-effective_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['staff_list'] = Staff.objects.filter(tenant=tenant, is_active=True)
        context['action_choices'] = EmploymentHistory.ACTION_CHOICES
        
        context['filters'] = {
            'action': self.request.GET.get('action', ''),
            'staff': self.request.GET.get('staff', ''),
            'start_date': self.request.GET.get('start_date', ''),
            'end_date': self.request.GET.get('end_date', ''),
        }
        
        return context


class EmploymentHistoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = EmploymentHistory
    form_class = EmploymentHistoryForm
    template_name = 'hr/history/form.html'
    permission_required = 'hr.add_employmenthistory'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:employmenthistory_list')
    
    def form_valid(self, form):
        form.instance.initiated_by = self.request.user
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Employment history record created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_EMPLOYMENT_HISTORY',
            resource_type='Employment History',
            resource_id=str(self.object.id),
            details={'staff': self.object.staff.full_name, 'action': self.object.action},
            severity='INFO'
        )
        
        return response


# ==================== PROMOTION VIEWS ====================

class PromotionListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = Promotion
    template_name = 'hr/history/promotion_list.html'
    context_object_name = 'promotions'
    permission_required = 'hr.view_promotion'
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('staff', 'previous_designation', 'new_designation', 'staff__user')
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date and end_date:
            try:
                start = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(effective_date__range=[start, end])
            except ValueError:
                pass
        
        return queryset.order_by('-effective_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['staff_list'] = Staff.objects.filter(tenant=tenant, is_active=True)
        
        context['filters'] = {
            'staff': self.request.GET.get('staff', ''),
            'start_date': self.request.GET.get('start_date', ''),
            'end_date': self.request.GET.get('end_date', ''),
        }
        
        return context


class PromotionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Promotion
    form_class = PromotionForm
    template_name = 'hr/history/promotion_form.html'
    permission_required = 'hr.add_promotion'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        
        # Pre-fill staff and previous designation if staff_id is provided
        staff_id = self.request.GET.get('staff_id')
        if staff_id:
            try:
                staff = Staff.objects.get(id=staff_id, tenant=get_current_tenant())
                kwargs['initial'] = {
                    'staff': staff,
                    'previous_designation': staff.designation,
                    'salary_before': staff.basic_salary
                }
            except Staff.DoesNotExist:
                pass
        
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:promotion_list')
    
    def form_valid(self, form):
        form.instance.approved_by = self.request.user
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        
        response = super().form_valid(form)
        
        # Update staff designation and salary
        staff = form.instance.staff
        staff.designation = form.instance.new_designation
        staff.basic_salary = form.instance.salary_after
        staff.save()
        
        messages.success(
            self.request,
            f"Promotion for {form.instance.staff.full_name} recorded successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_PROMOTION',
            resource_type='Promotion',
            resource_id=str(self.object.id),
            details={
                'staff': self.object.staff.full_name,
                'from': self.object.previous_designation.title,
                'to': self.object.new_designation.title
            },
            severity='INFO'
        )
        
        return response


class PromotionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = Promotion
    form_class = PromotionForm
    template_name = 'hr/history/promotion_form.html'
    permission_required = 'hr.change_promotion'
    roles_required = ['admin', 'hr_manager']
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:promotion_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Promotion record for {form.instance.staff.full_name} updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_PROMOTION',
            resource_type='Promotion',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class PromotionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = Promotion
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_promotion'
    roles_required = ['admin', 'hr_manager']
    
    def get_success_url(self):
        return reverse_lazy('hr:promotion_list')
    
    def delete(self, request, *args, **kwargs):
        promotion = self.get_object()
        
        # Perform soft delete
        promotion.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Promotion record for {promotion.staff.full_name} has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_PROMOTION',
            resource_type='Promotion',
            resource_id=str(promotion.id),
            details={'staff': promotion.staff.full_name},
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


# ==================== TRANSFER VIEWS ====================

class TransferListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    """List staff transfers (special type of employment history)"""
    template_name = 'hr/history/transfer_list.html'
    context_object_name = 'transfers'
    permission_required = 'hr.view_employmenthistory'
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = EmploymentHistory.objects.filter(
            tenant=tenant,
            action='TRANSFER'
        ).select_related('staff', 'staff__user', 'staff__department')
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date and end_date:
            try:
                start = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(effective_date__range=[start, end])
            except ValueError:
                pass
        
        return queryset.order_by('-effective_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['staff_list'] = Staff.objects.filter(tenant=tenant, is_active=True)
        
        context['filters'] = {
            'staff': self.request.GET.get('staff', ''),
            'start_date': self.request.GET.get('start_date', ''),
            'end_date': self.request.GET.get('end_date', ''),
        }
        
        return context


# ==================== DOCUMENT VIEWS ====================

class StaffDocumentListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = StaffDocument
    template_name = 'hr/documents/list.html'
    context_object_name = 'documents'
    permission_required = 'hr.view_staffdocument'
    paginate_by = 50
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('staff', 'staff__user', 'verified_by')
        
        # Filter by document type
        doc_type = self.request.GET.get('document_type')
        if doc_type and doc_type != 'all':
            queryset = queryset.filter(document_type=doc_type)
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by verification status
        is_verified = self.request.GET.get('is_verified')
        if is_verified and is_verified != 'all':
            queryset = queryset.filter(is_verified=(is_verified == 'true'))
        
        # Filter by expiry
        expiry_filter = self.request.GET.get('expiry_filter')
        if expiry_filter == 'expired':
            queryset = queryset.filter(expiry_date__lt=timezone.now().date())
        elif expiry_filter == 'expiring_soon':
            thirty_days_later = timezone.now().date() + timezone.timedelta(days=30)
            queryset = queryset.filter(
                expiry_date__gte=timezone.now().date(),
                expiry_date__lte=thirty_days_later
            )
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['staff_list'] = Staff.objects.filter(tenant=tenant, is_active=True)
        context['document_types'] = StaffDocument.DOCUMENT_TYPE_CHOICES
        
        context['filters'] = {
            'document_type': self.request.GET.get('document_type', ''),
            'staff': self.request.GET.get('staff', ''),
            'is_verified': self.request.GET.get('is_verified', ''),
            'expiry_filter': self.request.GET.get('expiry_filter', ''),
        }
        
        return context


class StaffDocumentUploadView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = StaffDocument
    form_class = StaffDocumentForm
    template_name = 'hr/documents/upload.html'
    permission_required = 'hr.add_staffdocument'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        
        # Pre-select staff if provided
        staff_id = self.request.GET.get('staff_id')
        if staff_id:
            try:
                staff = Staff.objects.get(id=staff_id, tenant=get_current_tenant())
                kwargs['initial'] = {'staff': staff}
            except Staff.DoesNotExist:
                pass
        
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:staffdocument_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        
        # Set file name if not provided
        if not form.instance.file_name and form.instance.file:
            form.instance.file_name = form.instance.file.name
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Document uploaded successfully for {form.instance.staff.full_name}."
        )
        
        audit_log(
            user=self.request.user,
            action='UPLOAD_STAFF_DOCUMENT',
            resource_type='Staff Document',
            resource_id=str(self.object.id),
            details={
                'staff': self.object.staff.full_name,
                'document_type': self.object.get_document_type_display(),
                'file_name': self.object.file_name
            },
            severity='INFO'
        )
        
        return response


class StaffDocumentViewView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DetailView):
    """View staff document"""
    model = StaffDocument
    template_name = 'hr/documents/view.html'
    context_object_name = 'document'
    permission_required = 'hr.view_staffdocument'


class StaffDocumentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = StaffDocument
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_staffdocument'
    
    def get_success_url(self):
        return reverse_lazy('hr:staffdocument_list')
    
    def delete(self, request, *args, **kwargs):
        document = self.get_object()
        
        # Perform soft delete
        document.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Document '{document.file_name}' has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_STAFF_DOCUMENT',
            resource_type='Staff Document',
            resource_id=str(document.id),
            details={
                'staff': document.staff.full_name,
                'file_name': document.file_name
            },
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


class StaffDocumentVerifyView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, View):
    """Verify staff document"""
    permission_required = 'hr.verify_staffdocument'
    roles_required = ['admin', 'hr_manager']
    http_method_names = ['post']
    
    def post(self, request, *args, **kwargs):
        document = get_object_or_404(
            StaffDocument.objects.filter(tenant=get_current_tenant()),
            pk=kwargs['pk']
        )
        
        if document.is_verified:
            return JsonResponse({
                'success': False,
                'error': 'Document is already verified.'
            })
        
        # Verify the document
        document.is_verified = True
        document.verified_by = request.user
        document.verified_at = timezone.now()
        document.save()
        
        audit_log(
            user=request.user,
            action='VERIFY_STAFF_DOCUMENT',
            resource_type='Staff Document',
            resource_id=str(document.id),
            details={
                'staff': document.staff.full_name,
                'document_type': document.get_document_type_display(),
                'file_name': document.file_name
            },
            severity='INFO'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Document verified successfully.'
        })


# ==================== ADDRESS VIEWS ====================

class StaffAddressListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = StaffAddress
    template_name = 'hr/addresses/list.html'
    context_object_name = 'addresses'
    permission_required = 'hr.view_staffaddress'
    
    def get_queryset(self):
        tenant = get_current_tenant()
        queryset = super().get_queryset().filter(
            tenant=tenant
        ).select_related('staff', 'staff__user')
        
        # Filter by staff
        staff_id = self.request.GET.get('staff')
        if staff_id:
            queryset = queryset.filter(staff_id=staff_id)
        
        # Filter by address type
        address_type = self.request.GET.get('address_type')
        if address_type and address_type != 'all':
            queryset = queryset.filter(address_type=address_type)
        
        # Filter by current address
        is_current = self.request.GET.get('is_current')
        if is_current and is_current != 'all':
            queryset = queryset.filter(is_current=(is_current == 'true'))
        
        return queryset.order_by('staff__user__first_name', 'address_type')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        context['staff_list'] = Staff.objects.filter(tenant=tenant, is_active=True)
        context['address_types'] = StaffAddress.ADDRESS_TYPE_CHOICES
        
        context['filters'] = {
            'staff': self.request.GET.get('staff', ''),
            'address_type': self.request.GET.get('address_type', ''),
            'is_current': self.request.GET.get('is_current', ''),
        }
        
        return context


class StaffAddressCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = StaffAddress
    form_class = StaffAddressForm
    template_name = 'hr/addresses/form.html'
    permission_required = 'hr.add_staffaddress'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        
        # Pre-select staff if provided
        staff_id = self.request.GET.get('staff_id')
        if staff_id:
            try:
                staff = Staff.objects.get(id=staff_id, tenant=get_current_tenant())
                kwargs['initial'] = {'staff': staff}
            except Staff.DoesNotExist:
                pass
        
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:staffaddress_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        
        # If this is set as current address, update other addresses
        if form.instance.is_current:
            StaffAddress.objects.filter(
                staff=form.instance.staff,
                address_type=form.instance.address_type
            ).update(is_current=False)
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Address for {form.instance.staff.full_name} created successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='CREATE_STAFF_ADDRESS',
            resource_type='Staff Address',
            resource_id=str(self.object.id),
            details={
                'staff': self.object.staff.full_name,
                'address_type': self.object.get_address_type_display()
            },
            severity='INFO'
        )
        
        return response


class StaffAddressUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, UpdateView):
    model = StaffAddress
    form_class = StaffAddressForm
    template_name = 'hr/addresses/form.html'
    permission_required = 'hr.change_staffaddress'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_success_url(self):
        return reverse_lazy('hr:staffaddress_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        
        # If this is set as current address, update other addresses
        if form.instance.is_current:
            StaffAddress.objects.filter(
                staff=form.instance.staff,
                address_type=form.instance.address_type
            ).exclude(id=form.instance.id).update(is_current=False)
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f"Address for {form.instance.staff.full_name} updated successfully."
        )
        
        audit_log(
            user=self.request.user,
            action='UPDATE_STAFF_ADDRESS',
            resource_type='Staff Address',
            resource_id=str(self.object.id),
            details={'changes': form.changed_data},
            severity='INFO'
        )
        
        return response


class StaffAddressDeleteView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, DeleteView):
    model = StaffAddress
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_staffaddress'
    
    def get_success_url(self):
        return reverse_lazy('hr:staffaddress_list')
    
    def delete(self, request, *args, **kwargs):
        address = self.get_object()
        
        # Check if this is the only address of this type for the staff
        other_addresses = StaffAddress.objects.filter(
            staff=address.staff,
            address_type=address.address_type
        ).exclude(id=address.id)
        
        if not other_addresses.exists() and address.is_current:
            return JsonResponse({
                'success': False,
                'error': 'Cannot delete the only current address of this type.'
            })
        
        # Perform soft delete
        address.delete(
            user=request.user,
            reason=request.POST.get('deletion_reason', 'No reason provided'),
            category=request.POST.get('deletion_category', 'ADMIN_ACTION')
        )
        
        messages.success(
            request,
            f"Address for {address.staff.full_name} has been deleted."
        )
        
        audit_log(
            user=request.user,
            action='DELETE_STAFF_ADDRESS',
            resource_type='Staff Address',
            resource_id=str(address.id),
            details={
                'staff': address.staff.full_name,
                'address_type': address.get_address_type_display()
            },
            severity='WARNING'
        )
        
        return JsonResponse({
            'success': True,
            'redirect': self.get_success_url()
        })


# ==================== REPORT VIEWS ====================

class StaffReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Staff report view"""
    template_name = 'hr/reports/staff.html'
    permission_required = 'hr.view_staff_report'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        # Get staff statistics
        staff_queryset = Staff.objects.filter(tenant=tenant, is_active=True)
        
        # Department distribution
        departments = Department.objects.filter(tenant=tenant)
        dept_stats = []
        for dept in departments:
            count = staff_queryset.filter(department=dept).count()
            if count > 0:
                dept_stats.append({
                    'department': dept.name,
                    'count': count,
                    'percentage': (count / staff_queryset.count()) * 100
                })
        
        # Designation distribution
        designations = Designation.objects.filter(tenant=tenant)
        desig_stats = []
        for desig in designations:
            count = staff_queryset.filter(designation=desig).count()
            if count > 0:
                desig_stats.append({
                    'designation': desig.title,
                    'count': count,
                    'percentage': (count / staff_queryset.count()) * 100
                })
        
        # Employment type distribution
        emp_types = Staff.EMPLOYMENT_TYPE_CHOICES
        emp_type_stats = []
        for emp_type_code, emp_type_name in emp_types:
            count = staff_queryset.filter(employment_type=emp_type_code).count()
            if count > 0:
                emp_type_stats.append({
                    'type': emp_type_name,
                    'count': count,
                    'percentage': (count / staff_queryset.count()) * 100
                })
        
        # Gender distribution
        genders = Staff.GENDER_CHOICES
        gender_stats = []
        for gender_code, gender_name in genders:
            count = staff_queryset.filter(gender=gender_code).count()
            if count > 0:
                gender_stats.append({
                    'gender': gender_name,
                    'count': count,
                    'percentage': (count / staff_queryset.count()) * 100
                })
        
        context['dept_stats'] = dept_stats
        context['desig_stats'] = desig_stats
        context['emp_type_stats'] = emp_type_stats
        context['gender_stats'] = gender_stats
        context['total_staff'] = staff_queryset.count()
        
        # Age distribution
        age_groups = {
            'Under 30': staff_queryset.filter(date_of_birth__gte=timezone.now() - timezone.timedelta(days=30*365)).count(),
            '30-40': staff_queryset.filter(
                date_of_birth__lt=timezone.now() - timezone.timedelta(days=30*365),
                date_of_birth__gte=timezone.now() - timezone.timedelta(days=40*365)
            ).count(),
            '40-50': staff_queryset.filter(
                date_of_birth__lt=timezone.now() - timezone.timedelta(days=40*365),
                date_of_birth__gte=timezone.now() - timezone.timedelta(days=50*365)
            ).count(),
            '50-60': staff_queryset.filter(
                date_of_birth__lt=timezone.now() - timezone.timedelta(days=50*365),
                date_of_birth__gte=timezone.now() - timezone.timedelta(days=60*365)
            ).count(),
            'Over 60': staff_queryset.filter(date_of_birth__lt=timezone.now() - timezone.timedelta(days=60*365)).count(),
        }
        
        context['age_groups'] = age_groups
        
        # Service years distribution
        service_years = {}
        for staff in staff_queryset:
            years = staff.service_years
            if years < 1:
                group = 'Less than 1 year'
            elif years < 3:
                group = '1-3 years'
            elif years < 5:
                group = '3-5 years'
            elif years < 10:
                group = '5-10 years'
            else:
                group = '10+ years'
            
            service_years[group] = service_years.get(group, 0) + 1
        
        context['service_years'] = service_years
        
        return context


from django.template.loader import get_template
from xhtml2pdf import pisa

class StaffReportPDFView(StaffReportView):
    """Staff report PDF view"""
    template_name = 'hr/reports/staff_pdf.html'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        template = get_template(self.template_name)
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="staff_report.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response


class AttendanceReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """StaffAttendance report view"""
    template_name = 'hr/reports/StaffAttendance.html'
    permission_required = 'hr.view_attendance_report'
    roles_required = ['admin', 'hr_manager', 'attendance_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        # Filter parameters
        month_str = self.request.GET.get('month')
        dept_id = self.request.GET.get('department')
        
        # Calculate date range
        today = timezone.now().date()
        if month_str:
            try:
                year, month = map(int, month_str.split('-'))
                start_date = timezone.datetime(year, month, 1).date()
                import calendar
                _, last_day = calendar.monthrange(year, month)
                end_date = timezone.datetime(year, month, last_day).date()
            except ValueError:
                start_date = today.replace(day=1)
                end_date = today
                month_str = start_date.strftime('%Y-%m')
        else:
            start_date = today.replace(day=1)
            end_date = today
            month_str = start_date.strftime('%Y-%m')

        # Get departments for filter
        context['departments'] = Department.objects.filter(tenant=tenant)
        context['filters'] = {'month': month_str, 'department': dept_id}
        
        # Staff Query
        staff_qs = Staff.objects.filter(tenant=tenant, is_active=True).select_related('user', 'department').order_by('user__first_name')
        if dept_id:
            staff_qs = staff_qs.filter(department_id=dept_id)
            
        # Attendance Records
        attendance_records = StaffAttendance.objects.filter(
            tenant=tenant,
            date__range=[start_date, end_date]
        ).values('staff_id', 'date', 'status')
        
        # Create lookup map: (staff_id, date) -> status
        attendance_map = {(r['staff_id'], r['date']): r['status'] for r in attendance_records}
        
        # Generate Days List
        report_days = []
        current_date = start_date
        while current_date <= end_date:
            report_days.append({
                'date': current_date,
                'day': current_date.day,
                'weekday': current_date.strftime('%a')
            })
            current_date += timezone.timedelta(days=1)
            
        context['report_days'] = report_days
        
        # Build Matrix Data
        report_data = []
        for staff in staff_qs:
            daily_status = []
            present_count = 0
            absent_count = 0
            
            for day_info in report_days:
                date = day_info['date']
                status = attendance_map.get((staff.id, date), '-')
                
                # Normalize status code to short form for template badge logic if needed
                # Template expects 'P', 'A', 'L', 'H' or full code?
                # Template badges: P, A, L, H.
                # Model choices: PRESENT, ABSENT, LATE, HALF_DAY, LEAVE
                short_status = '-'
                if status == 'PRESENT':
                    short_status = 'P'
                    present_count += 1
                elif status == 'LATE':
                    short_status = 'L'
                    present_count += 1
                elif status == 'ABSENT':
                    short_status = 'A'
                    absent_count += 1
                elif status == 'HALF_DAY':
                    short_status = 'H'
                    present_count += 0.5
                elif status == 'LEAVE':
                    short_status = 'L' # Or use specific leave code
                
                daily_status.append(short_status)
            
            report_data.append({
                'staff': staff,
                'daily_status': daily_status,
                'present_count': present_count,
                'absent_count': absent_count
            })
            
        context['report_data'] = report_data
        
        return context
        
        while current_date <= end_date:
            day_attendance = attendance_records.filter(date=current_date)
            if day_attendance.exists():
                present = day_attendance.filter(status__in=['PRESENT', 'LATE']).count()
                absent = day_attendance.filter(status='ABSENT').count()
                leave = day_attendance.filter(status='LEAVE').count()
                
                daily_data.append({
                    'date': current_date,
                    'present': present,
                    'absent': absent,
                    'leave': leave,
                    'attendance_rate': (present / total_staff) * 100 if total_staff > 0 else 0
                })
            
            current_date += timezone.timedelta(days=1)
        
        context['daily_data'] = daily_data
        
        return context


class LeaveReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Leave report view"""
    template_name = 'hr/reports/leave.html'
    permission_required = 'hr.view_leave_report'
    roles_required = ['admin', 'hr_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        # Get year
        year = int(self.request.GET.get('year', timezone.now().year))
        
        # Get leave applications for the year
        leave_apps = LeaveApplication.objects.filter(
            tenant=tenant,
            start_date__year=year,
            status='APPROVED'
        ).select_related('staff', 'leave_type', 'staff__department')
        
        # Get leave balances for the year
        leave_balances = LeaveBalance.objects.filter(
            tenant=tenant,
            year=year
        ).select_related('staff', 'leave_type')
        
        # Leave type statistics
        leave_types = LeaveType.objects.filter(tenant=tenant, is_active=True)
        leave_type_stats = []
        
        for leave_type in leave_types:
            type_apps = leave_apps.filter(leave_type=leave_type)
            type_balances = leave_balances.filter(leave_type=leave_type)
            
            if type_apps.exists() or type_balances.exists():
                total_entitled = sum(b.total_entitled for b in type_balances)
                total_used = sum(b.used_days for b in type_balances)
                total_applications = type_apps.count()
                total_days = sum(app.total_days for app in type_apps)
                
                leave_type_stats.append({
                    'leave_type': leave_type.name,
                    'total_entitled': total_entitled,
                    'total_used': total_used,
                    'available': total_entitled - total_used,
                    'utilization_rate': (total_used / total_entitled * 100) if total_entitled > 0 else 0,
                    'applications': total_applications,
                    'days': total_days
                })
        
        context['leave_type_stats'] = leave_type_stats
        context['selected_year'] = year
        context['years'] = range(timezone.now().year - 5, timezone.now().year + 2)
        
        # Monthly leave trend
        monthly_data = []
        for month in range(1, 13):
            month_apps = leave_apps.filter(start_date__month=month)
            if month_apps.exists():
                monthly_data.append({
                    'month': month,
                    'applications': month_apps.count(),
                    'days': sum(app.total_days for app in month_apps)
                })
        
        context['monthly_data'] = monthly_data
        
        # Department-wise leave statistics
        departments = Department.objects.filter(tenant=tenant)
        dept_stats = []
        
        for dept in departments:
            dept_staff = Staff.objects.filter(department=dept, is_active=True)
            if dept_staff.exists():
                dept_apps = leave_apps.filter(staff__department=dept)
                dept_balances = leave_balances.filter(staff__department=dept)
                
                if dept_apps.exists() or dept_balances.exists():
                    total_entitled = sum(b.total_entitled for b in dept_balances)
                    total_used = sum(b.used_days for b in dept_balances)
                    
                    dept_stats.append({
                        'department': dept.name,
                        'staff_count': dept_staff.count(),
                        'applications': dept_apps.count(),
                        'days': sum(app.total_days for app in dept_apps),
                        'total_entitled': total_entitled,
                        'total_used': total_used,
                        'utilization_rate': (total_used / total_entitled * 100) if total_entitled > 0 else 0
                    })
        
        context['dept_stats'] = dept_stats
        
        return context


class TurnoverReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Staff turnover report view"""
    template_name = 'hr/reports/turnover.html'
    permission_required = 'hr.view_turnover_report'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        # Get year range
        start_year = int(self.request.GET.get('start_year', timezone.now().year - 5))
        end_year = int(self.request.GET.get('end_year', timezone.now().year))
        
        # Staff turnover data by year
        turnover_data = []
        
        for year in range(start_year, end_year + 1):
            # Staff joined in the year
            joined = Staff.objects.filter(
                tenant=tenant,
                joining_date__year=year,
                is_active=True
            ).count()
            
            # Staff left in the year
            left = Staff.objects.filter(
                tenant=tenant,
                employment_status__in=['TERMINATED', 'RESIGNED', 'RETIRED'],
                updated_at__year=year
            ).count()
            
            # Average staff count (simplified)
            avg_staff = Staff.objects.filter(
                tenant=tenant,
                is_active=True,
                created_at__lt=timezone.datetime(year+1, 1, 1)
            ).count()
            
            turnover_rate = (left / avg_staff * 100) if avg_staff > 0 else 0
            
            turnover_data.append({
                'year': year,
                'joined': joined,
                'left': left,
                'avg_staff': avg_staff,
                'turnover_rate': turnover_rate
            })
        
        context['turnover_data'] = turnover_data
        context['start_year'] = start_year
        context['end_year'] = end_year
        context['years'] = range(timezone.now().year - 10, timezone.now().year + 1)
        
        # Reason for leaving
        employment_history = EmploymentHistory.objects.filter(
            tenant=tenant,
            action__in=['TERMINATION', 'RESIGNATION', 'RETIREMENT'],
            effective_date__year__gte=start_year
        )
        
        reason_stats = {}
        for history in employment_history:
            action = history.action
            reason_stats[action] = reason_stats.get(action, 0) + 1
        
        context['reason_stats'] = reason_stats
        
        # Department-wise turnover
        departments = Department.objects.filter(tenant=tenant)
        dept_turnover = []
        
        for dept in departments:
            dept_staff = Staff.objects.filter(department=dept)
            if dept_staff.exists():
                joined = dept_staff.filter(
                    joining_date__year=end_year,
                    is_active=True
                ).count()
                
                left = dept_staff.filter(
                    employment_status__in=['TERMINATED', 'RESIGNED', 'RETIRED'],
                    updated_at__year=end_year
                ).count()
                
                avg_staff = dept_staff.filter(is_active=True).count()
                turnover_rate = (left / avg_staff * 100) if avg_staff > 0 else 0
                
                dept_turnover.append({
                    'department': dept.name,
                    'joined': joined,
                    'left': left,
                    'avg_staff': avg_staff,
                    'turnover_rate': turnover_rate
                })
        
        context['dept_turnover'] = dept_turnover
        
        return context


class DemographicReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Demographic report view"""
    template_name = 'hr/reports/demographic.html'
    permission_required = 'hr.view_demographic_report'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        # Get all active staff
        staff_queryset = Staff.objects.filter(tenant=tenant, is_active=True)
        
        # Gender distribution
        gender_stats = {}
        for gender_code, gender_name in Staff.GENDER_CHOICES:
            count = staff_queryset.filter(gender=gender_code).count()
            if count > 0:
                gender_stats[gender_name] = count
        
        context['gender_stats'] = gender_stats
        
        # Age distribution
        age_groups = {
            'Under 25': staff_queryset.filter(date_of_birth__gte=timezone.now() - timezone.timedelta(days=25*365)).count(),
            '25-35': staff_queryset.filter(
                date_of_birth__lt=timezone.now() - timezone.timedelta(days=25*365),
                date_of_birth__gte=timezone.now() - timezone.timedelta(days=35*365)
            ).count(),
            '35-45': staff_queryset.filter(
                date_of_birth__lt=timezone.now() - timezone.timedelta(days=35*365),
                date_of_birth__gte=timezone.now() - timezone.timedelta(days=45*365)
            ).count(),
            '45-55': staff_queryset.filter(
                date_of_birth__lt=timezone.now() - timezone.timedelta(days=45*365),
                date_of_birth__gte=timezone.now() - timezone.timedelta(days=55*365)
            ).count(),
            '55-65': staff_queryset.filter(
                date_of_birth__lt=timezone.now() - timezone.timedelta(days=55*365),
                date_of_birth__gte=timezone.now() - timezone.timedelta(days=65*365)
            ).count(),
            'Over 65': staff_queryset.filter(date_of_birth__lt=timezone.now() - timezone.timedelta(days=65*365)).count(),
        }
        
        context['age_groups'] = age_groups
        
        # Marital status distribution
        marital_stats = {}
        for status_code, status_name in Staff.MARITAL_STATUS_CHOICES:
            count = staff_queryset.filter(marital_status=status_code).count()
            if count > 0:
                marital_stats[status_name] = count
        
        context['marital_stats'] = marital_stats
        
        # Nationality distribution
        nationalities = staff_queryset.values_list('nationality', flat=True).distinct()
        nationality_stats = {}
        
        for nationality in nationalities:
            if nationality:
                count = staff_queryset.filter(nationality=nationality).count()
                if count > 0:
                    nationality_stats[nationality] = count
        
        context['nationality_stats'] = nationality_stats
        
        # Blood group distribution
        blood_groups = staff_queryset.values_list('blood_group', flat=True).distinct()
        blood_group_stats = {}
        
        for blood_group in blood_groups:
            if blood_group:
                count = staff_queryset.filter(blood_group=blood_group).count()
                if count > 0:
                    blood_group_stats[blood_group] = count
        
        context['blood_group_stats'] = blood_group_stats
        
        # Service years distribution
        service_years = {}
        for staff in staff_queryset:
            years = staff.service_years
            if years < 1:
                group = 'Less than 1 year'
            elif years < 3:
                group = '1-3 years'
            elif years < 5:
                group = '3-5 years'
            elif years < 10:
                group = '5-10 years'
            elif years < 20:
                group = '10-20 years'
            else:
                group = '20+ years'
            
            service_years[group] = service_years.get(group, 0) + 1
        
        context['service_years'] = service_years
        
        # Educational qualification distribution
        qualification_stats = {}
        for staff in staff_queryset:
            if staff.qualifications:
                for qual in staff.qualifications:
                    degree = qual.get('degree', 'Unknown')
                    qualification_stats[degree] = qualification_stats.get(degree, 0) + 1
        
        context['qualification_stats'] = qualification_stats
        
        return context


class SalaryAnalysisView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Salary analysis report view"""
    template_name = 'hr/reports/salary.html'
    permission_required = 'hr.view_salary_report'
    roles_required = ['admin', 'hr_manager', 'accountant', 'principal']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        # Get all active staff with salaries
        staff_queryset = Staff.objects.filter(
            tenant=tenant,
            is_active=True,
            employment_status='ACTIVE'
        ).select_related('department', 'designation')
        
        # Overall salary statistics
        salaries = [s.basic_salary for s in staff_queryset]
        
        if salaries:
            context['total_salary'] = sum(salaries)
            context['avg_salary'] = sum(salaries) / len(salaries)
            context['min_salary'] = min(salaries)
            context['max_salary'] = max(salaries)
            context['median_salary'] = sorted(salaries)[len(salaries) // 2]
        else:
            context['total_salary'] = 0
            context['avg_salary'] = 0
            context['min_salary'] = 0
            context['max_salary'] = 0
            context['median_salary'] = 0
        
        # Department-wise salary analysis
        departments = Department.objects.filter(tenant=tenant)
        dept_salary_stats = []
        
        for dept in departments:
            dept_staff = staff_queryset.filter(department=dept)
            if dept_staff.exists():
                dept_salaries = [s.basic_salary for s in dept_staff]
                
                dept_salary_stats.append({
                    'department': dept.name,
                    'staff_count': dept_staff.count(),
                    'total_salary': sum(dept_salaries),
                    'avg_salary': sum(dept_salaries) / len(dept_salaries) if dept_salaries else 0,
                    'min_salary': min(dept_salaries) if dept_salaries else 0,
                    'max_salary': max(dept_salaries) if dept_salaries else 0
                })
        
        context['dept_salary_stats'] = dept_salary_stats
        
        # Designation-wise salary analysis
        designations = Designation.objects.filter(tenant=tenant)
        desig_salary_stats = []
        
        for desig in designations:
            desig_staff = staff_queryset.filter(designation=desig)
            if desig_staff.exists():
                desig_salaries = [s.basic_salary for s in desig_staff]
                
                desig_salary_stats.append({
                    'designation': desig.title,
                    'category': desig.get_category_display(),
                    'staff_count': desig_staff.count(),
                    'total_salary': sum(desig_salaries),
                    'avg_salary': sum(desig_salaries) / len(desig_salaries) if desig_salaries else 0,
                    'min_salary': min(desig_salaries) if desig_salaries else 0,
                    'max_salary': max(desig_salaries) if desig_salaries else 0,
                    'designation_min': desig.min_salary,
                    'designation_max': desig.max_salary
                })
        
        context['desig_salary_stats'] = desig_salary_stats
        
        # Employment type salary analysis
        emp_types = Staff.EMPLOYMENT_TYPE_CHOICES
        emp_type_stats = []
        
        for emp_type_code, emp_type_name in emp_types:
            emp_staff = staff_queryset.filter(employment_type=emp_type_code)
            if emp_staff.exists():
                emp_salaries = [s.basic_salary for s in emp_staff]
                
                emp_type_stats.append({
                    'type': emp_type_name,
                    'staff_count': emp_staff.count(),
                    'total_salary': sum(emp_salaries),
                    'avg_salary': sum(emp_salaries) / len(emp_salaries) if emp_salaries else 0,
                    'min_salary': min(emp_salaries) if emp_salaries else 0,
                    'max_salary': max(emp_salaries) if emp_salaries else 0
                })
        
        context['emp_type_stats'] = emp_type_stats
        
        # Gender salary analysis
        gender_stats = []
        for gender_code, gender_name in Staff.GENDER_CHOICES:
            gender_staff = staff_queryset.filter(gender=gender_code)
            if gender_staff.exists():
                gender_salaries = [s.basic_salary for s in gender_staff]
                
                gender_stats.append({
                    'gender': gender_name,
                    'staff_count': gender_staff.count(),
                    'total_salary': sum(gender_salaries),
                    'avg_salary': sum(gender_salaries) / len(gender_salaries) if gender_salaries else 0,
                    'min_salary': min(gender_salaries) if gender_salaries else 0,
                    'max_salary': max(gender_salaries) if gender_salaries else 0
                })
        
        context['gender_stats'] = gender_stats
        
        # Experience vs Salary analysis
        experience_groups = {
            '0-2 years': [],
            '2-5 years': [],
            '5-10 years': [],
            '10-20 years': [],
            '20+ years': []
        }
        
        for staff in staff_queryset:
            experience = staff.total_experience
            
            if experience < 2:
                group = '0-2 years'
            elif experience < 5:
                group = '2-5 years'
            elif experience < 10:
                group = '5-10 years'
            elif experience < 20:
                group = '10-20 years'
            else:
                group = '20+ years'
            
            experience_groups[group].append(float(staff.basic_salary))
        
        experience_stats = []
        for group, salaries in experience_groups.items():
            if salaries:
                experience_stats.append({
                    'experience': group,
                    'staff_count': len(salaries),
                    'avg_salary': sum(salaries) / len(salaries),
                    'min_salary': min(salaries),
                    'max_salary': max(salaries)
                })
        
        context['experience_stats'] = experience_stats
        
        # Salary distribution by ranges
        salary_ranges = {
            'Under ₹20,000': 0,
            '₹20,000 - ₹40,000': 0,
            '₹40,000 - ₹60,000': 0,
            '₹60,000 - ₹80,000': 0,
            '₹80,000 - ₹1,00,000': 0,
            'Over ₹1,00,000': 0
        }
        
        for salary in salaries:
            salary_float = float(salary)
            if salary_float < 20000:
                salary_ranges['Under ₹20,000'] += 1
            elif salary_float < 40000:
                salary_ranges['₹20,000 - ₹40,000'] += 1
            elif salary_float < 60000:
                salary_ranges['₹40,000 - ₹60,000'] += 1
            elif salary_float < 80000:
                salary_ranges['₹60,000 - ₹80,000'] += 1
            elif salary_float < 100000:
                salary_ranges['₹80,000 - ₹1,00,000'] += 1
            else:
                salary_ranges['Over ₹1,00,000'] += 1
        
        context['salary_ranges'] = salary_ranges
        
        return context


# ==================== SETTINGS VIEWS ====================

class HRSettingsView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """HR Settings view"""
    template_name = 'hr/settings/index.html'
    permission_required = 'hr.view_hr_settings'
    roles_required = ['admin', 'hr_manager']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        # Get configuration statistics
        context['department_count'] = Department.objects.filter(tenant=tenant).count()
        context['designation_count'] = Designation.objects.filter(tenant=tenant).count()
        context['leave_type_count'] = LeaveType.objects.filter(tenant=tenant, is_active=True).count()
        context['training_program_count'] = TrainingProgram.objects.filter(tenant=tenant).count()
        
        return context


# Note: Additional view classes like HolidayListView, HolidayCreateView, WorkScheduleView, 
# TaxConfigView, PFESIConfigView would require additional models. 
# These are placeholders that would need to be implemented based on your specific requirements.

# ==================== REPORT GENERATION API ====================

class ReportGenerateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Generate report API"""
    permission_required = 'hr.generate_reports'
    roles_required = ['admin', 'hr_manager']
    
    def post(self, request, *args, **kwargs):
        report_type = request.POST.get('report_type')
        format_type = request.POST.get('format', 'pdf')
        parameters = json.loads(request.POST.get('parameters', '{}'))
        
        # Validate report type
        valid_reports = ['staff', 'StaffAttendance', 'leave', 'turnover', 'demographic', 'salary']
        if report_type not in valid_reports:
            return JsonResponse({
                'success': False,
                'error': f'Invalid report type: {report_type}'
            })
        
        # Generate report based on type
        report_data = self.generate_report_data(report_type, parameters)
        
        # Generate file based on format
        if format_type == 'pdf':
            file_content = self.generate_pdf_report(report_type, report_data)
            file_name = f"{report_type}_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            content_type = 'application/pdf'
        elif format_type == 'excel':
            file_content = self.generate_excel_report(report_type, report_data)
            file_name = f"{report_type}_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif format_type == 'csv':
            file_content = self.generate_csv_report(report_type, report_data)
            file_name = f"{report_type}_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
            content_type = 'text/csv'
        else:
            return JsonResponse({
                'success': False,
                'error': f'Invalid format type: {format_type}'
            })
        
        # Log the report generation
        audit_log(
            user=request.user,
            action='GENERATE_REPORT',
            resource_type='Report',
            details={
                'report_type': report_type,
                'format': format_type,
                'parameters': parameters
            },
            severity='INFO'
        )
        
        # Create HTTP response with file
        response = HttpResponse(file_content, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        
        return response
    
    def generate_report_data(self, report_type, parameters):
        """Generate report data based on type"""
        tenant = get_current_tenant()
        
        if report_type == 'staff':
            # Staff report data
            staff_queryset = Staff.objects.filter(tenant=tenant, is_active=True)
            return {
                'total_staff': staff_queryset.count(),
                'departments': Department.objects.filter(tenant=tenant),
                'staff_by_dept': staff_queryset.values('department__name').annotate(count=Count('id')),
                'staff_by_designation': staff_queryset.values('designation__title').annotate(count=Count('id')),
            }
        
        elif report_type == 'StaffAttendance':
            # StaffAttendance report data
            start_date = parameters.get('start_date', timezone.now().date() - timezone.timedelta(days=30))
            end_date = parameters.get('end_date', timezone.now().date())
            
            staff_attendance_qs = StaffAttendance.objects.filter(
                tenant=tenant,
                date__range=[start_date, end_date]
            ).select_related('staff', 'staff__user', 'staff__department').order_by('date', 'staff__user__first_name')
            
            records = staff_attendance_qs.values(
                'date', 
                'staff__employee_id', 
                'staff__user__first_name', 
                'staff__user__last_name', 
                'staff__department__name', 
                'status',
                'check_in',
                'check_out'
            )

            return {
                'start_date': start_date,
                'end_date': end_date,
                'total_records': staff_attendance_qs.count(),
                'attendance_by_status': staff_attendance_qs.values('status').annotate(count=Count('id')),
                'attendance_by_dept': staff_attendance_qs.values('staff__department__name').annotate(count=Count('id')),
                'records': list(records) # Convert queryset to list for passing to export functions
            }
        
        elif report_type == 'leave':
            # Leave Analysis Data
            year = int(parameters.get('year', timezone.now().year))
            
            leave_apps = LeaveApplication.objects.filter(
                tenant=tenant,
                start_date__year=year,
                status='APPROVED'
            ).select_related('staff', 'leave_type', 'staff__department')
            
            leave_balances = LeaveBalance.objects.filter(
                tenant=tenant,
                year=year
            ).select_related('staff', 'leave_type')
            
            # Leave Type Stats
            leave_types = LeaveType.objects.filter(tenant=tenant, is_active=True)
            leave_type_stats = []
            for lt in leave_types:
                type_apps = leave_apps.filter(leave_type=lt)
                type_balances = leave_balances.filter(leave_type=lt)
                
                if type_apps.exists() or type_balances.exists():
                    total_entitled = sum(b.total_entitled for b in type_balances)
                    total_used = sum(b.used_days for b in type_balances)
                    leave_type_stats.append({
                        'leave_type': lt.name,
                        'total_entitled': total_entitled,
                        'total_used': total_used,
                        'available': total_entitled - total_used,
                        'utilization_rate': (total_used/total_entitled*100) if total_entitled else 0,
                        'applications': type_apps.count()
                    })
            
            # Monthly Trend
            monthly_data = []
            for m in range(1, 13):
                m_apps = leave_apps.filter(start_date__month=m)
                if m_apps.exists():
                    monthly_data.append({
                        'month': m, # Formatting to name in template is preferred
                        'applications': m_apps.count(),
                        'days': sum(a.total_days for a in m_apps)
                    })
            
            # Dept Stats
            departments = Department.objects.filter(tenant=tenant)
            dept_stats = []
            for d in departments:
                d_staff = Staff.objects.filter(department=d, is_active=True)
                if d_staff.exists():
                    d_apps = leave_apps.filter(staff__department=d)
                    d_balances = leave_balances.filter(staff__department=d)
                    
                    if d_apps.exists() or d_balances.exists():
                        total_entitled = sum(b.total_entitled for b in d_balances)
                        total_used = sum(b.used_days for b in d_balances)
                        dept_stats.append({
                            'department': d.name,
                            'staff_count': d_staff.count(),
                            'applications': d_apps.count(),
                            'days': sum(a.total_days for a in d_apps),
                            'utilization_rate': (total_used/total_entitled*100) if total_entitled else 0
                        })
            
            return {
                'selected_year': year,
                'leave_type_stats': leave_type_stats,
                'monthly_data': monthly_data,
                'dept_stats': dept_stats,
                'generated_at': timezone.now()
            }

        elif report_type == 'salary':
            # Salary Analysis Data
            staff_queryset = Staff.objects.filter(
                tenant=tenant,
                is_active=True,
                employment_status='ACTIVE'
            ).select_related('department', 'designation')

            # Overall Stats
            salaries = [s.basic_salary for s in staff_queryset if s.basic_salary]
            stats = {
                'total_salary': sum(salaries) if salaries else 0,
                'avg_salary': sum(salaries) / len(salaries) if salaries else 0,
                'min_salary': min(salaries) if salaries else 0,
                'max_salary': max(salaries) if salaries else 0,
                'median_salary': sorted(salaries)[len(salaries) // 2] if salaries else 0,
                'staff_count': len(salaries)
            }

            # Dept Stats
            departments = Department.objects.filter(tenant=tenant)
            dept_salary_stats = []
            for d in departments:
                d_staff = [s for s in staff_queryset if s.department_id == d.id and s.basic_salary]
                if d_staff:
                    d_salaries = [s.basic_salary for s in d_staff]
                    dept_salary_stats.append({
                        'department': d.name,
                        'staff_count': len(d_staff),
                        'total_salary': sum(d_salaries),
                        'avg_salary': sum(d_salaries) / len(d_salaries),
                        'min_salary': min(d_salaries),
                        'max_salary': max(d_salaries)
                    })

            # Designation Stats
            designations = Designation.objects.filter(tenant=tenant)
            desig_salary_stats = []
            for desig in designations:
                desig_staff = [s for s in staff_queryset if s.designation_id == desig.id and s.basic_salary]
                if desig_staff:
                    desig_salaries = [s.basic_salary for s in desig_staff]
                    desig_salary_stats.append({
                        'designation': desig.title,
                        'category': desig.get_category_display(),
                        'staff_count': len(desig_staff),
                        'avg_salary': sum(desig_salaries) / len(desig_salaries),
                        'min_salary': min(desig_salaries),
                        'max_salary': max(desig_salaries),
                        'designation_min': desig.min_salary,
                        'designation_max': desig.max_salary
                    })
            
            return {
                'stats': stats,
                'dept_salary_stats': dept_salary_stats,
                'desig_salary_stats': desig_salary_stats,
                'generated_at': timezone.now()
            }
    
    
    def generate_pdf_report(self, report_type, data):
        """Generate PDF report using xhtml2pdf"""
        from xhtml2pdf import pisa
        from django.template.loader import get_template
        
        template_name = f"hr/reports/{report_type}_pdf.html"
        # Fallback to StaffAttendance if type is capitalized differently or generic
        if report_type == 'StaffAttendance':
             template_name = "hr/reports/attendance_pdf.html"

        try:
            template = get_template(template_name)
            html = template.render(data)
            result = io.BytesIO()
            pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=result)
            if pisa_status.err:
                return b"Error generating PDF"
            return result.getvalue()
        except:
             # Fallback if specific template missing
             return b"PDF Generation Error: Template not found."

    def generate_excel_report(self, report_type, data):
        """Generate Excel report using openpyxl"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{report_type} Report"
        
        # Style
        header_font = Font(bold=True)
        
        if report_type == 'StaffAttendance':
            # Headers
            headers = ['Date', 'Staff ID', 'Staff Name', 'Department', 'Status', 'Check In', 'Check Out']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
            
            # Data
            # Note: parameters need to be passed correctly in generate_report_data to get detailed list
            row_num = 2
            for record in data.get('records', []):
                ws.cell(row=row_num, column=1, value=str(record['date']))
                ws.cell(row=row_num, column=2, value=record['staff__employee_id'])
                ws.cell(row=row_num, column=3, value=f"{record['staff__user__first_name']} {record['staff__user__last_name']}")
                ws.cell(row=row_num, column=4, value=record['staff__department__name'])
                ws.cell(row=row_num, column=5, value=record['status'])
                ws.cell(row=row_num, column=6, value=str(record['check_in']) if record['check_in'] else '')
                ws.cell(row=row_num, column=7, value=str(record['check_out']) if record['check_out'] else '')
                row_num += 1
                
        elif report_type == 'staff':
             headers = ['Employee ID', 'Name', 'Department', 'Designation', 'Joining Date', 'Status']
             for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
             
             row_num = 2
             for staff in data.get('staff_list', []):
                 ws.cell(row=row_num, column=1, value=staff['employee_id'])
                 ws.cell(row=row_num, column=2, value=f"{staff['user__first_name']} {staff['user__last_name']}")
                 ws.cell(row=row_num, column=3, value=staff['department__name'])
                 ws.cell(row=row_num, column=4, value=staff['designation__title'])
                 ws.cell(row=row_num, column=5, value=str(staff['joining_date']))
                 ws.cell(row=row_num, column=6, value=staff['status'])
                 row_num += 1
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_csv_report(self, report_type, data):
        """Generate CSV report"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        if report_type == 'StaffAttendance':
            writer.writerow(['Date', 'Staff ID', 'Staff Name', 'Department', 'Status', 'Check In', 'Check Out'])
            for record in data.get('records', []):
                writer.writerow([
                    record['date'],
                    record['staff__employee_id'],
                    f"{record['staff__user__first_name']} {record['staff__user__last_name']}",
                    record['staff__department__name'],
                    record['status'],
                    record['check_in'],
                    record['check_out']
                ])
                
        elif report_type == 'staff':
            writer.writerow(['Employee ID', 'Name', 'Department', 'Designation', 'Joining Date', 'Status'])
            for staff in data.get('staff_list', []):
                 writer.writerow([
                     staff['employee_id'],
                     f"{staff['user__first_name']} {staff['user__last_name']}",
                     staff['department__name'],
                     staff['designation__title'],
                     staff['joining_date'],
                     staff['status']
                 ])

        return output.getvalue().encode('utf-8')


# ==================== PERFORMANCE TEMPLATE VIEWS ====================

class PerformanceTemplateListView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Performance template list view"""
    template_name = 'hr/performance/template_list.html'
    permission_required = 'hr.view_performancetemplate'
    roles_required = ['admin', 'hr_manager']


class PerformanceGoalListView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Performance goal list view"""
    template_name = 'hr/performance/goal_list.html'
    permission_required = 'hr.view_performancegoal'
    roles_required = ['admin', 'hr_manager']


class PerformanceReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Performance report view"""
    template_name = 'hr/performance/report.html'
    permission_required = 'hr.view_performance_report'
    roles_required = ['admin', 'hr_manager', 'principal']

# ==================== SETTINGS VIEWS ====================

class HRSettingsView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'hr/settings/index.html'
    permission_required = 'hr.view_settings'
    roles_required = ['admin', 'hr_manager']

class HolidayListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Holiday
    template_name = 'hr/settings/holiday_list.html'
    context_object_name = 'holidays'
    permission_required = 'hr.view_holiday'

class HolidayCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Holiday
    fields = ['name', 'date', 'is_recurring', 'description']
    template_name = 'hr/settings/holiday_form.html'
    success_url = reverse_lazy('hr:holiday_list')
    permission_required = 'hr.add_holiday'

    def form_valid(self, form):
        form.instance.tenant = get_current_tenant()
        audit_log(self.request.user, "Created Holiday", f"Created holiday {form.instance.name}")
        return super().form_valid(form)

class WorkScheduleView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = WorkSchedule
    template_name = 'hr/settings/work_schedule.html'
    context_object_name = 'schedules'
    permission_required = 'hr.view_workschedule'

class TaxConfigView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = TaxConfig
    template_name = 'hr/settings/tax_config.html'
    context_object_name = 'tax_configs'
    permission_required = 'hr.view_taxconfig'

class PFESIConfigView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = PFESIConfig
    template_name = 'hr/settings/pf_esi_config.html'
    context_object_name = 'pf_esi_configs'
    permission_required = 'hr.view_pfesiconfig'


# ==================== Qualification VIEWS ====================

class QualificationListView(LoginRequiredMixin, PermissionRequiredMixin, TenantAccessMixin, ListView):
    model = Qualification
    template_name = 'hr/qualification/list.html'
    context_object_name = 'qualifications'
    permission_required = 'hr.view_qualification'

class QualificationCreateView(BaseCreateView):
    model = Qualification
    form_class = QualificationForm
    template_name = 'hr/qualification/form.html'
    permission_required = 'hr.add_qualification'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_success_url(self):
        return reverse_lazy('hr:qualification_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

class QualificationUpdateView(BaseUpdateView):
    model = Qualification
    form_class = QualificationForm
    template_name = 'hr/qualification/form.html'
    permission_required = 'hr.change_qualification'
    roles_required = ['admin', 'hr_manager', 'principal']
    
    def get_success_url(self):
        return reverse_lazy('hr:qualification_list')
    
    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

class QualificationDeleteView(BaseDeleteView):
    model = Qualification
    template_name = 'hr/common/confirm_delete.html'
    permission_required = 'hr.delete_qualification'
    roles_required = ['admin', 'principal']
    
    def get_success_url(self):
        return reverse_lazy('hr:qualification_list')
