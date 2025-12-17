import logging
from django.conf import settings

from django.db import transaction
from django.db.models import Count, Sum, Q
from django.http import JsonResponse, Http404
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import PermissionDenied, ValidationError
from django.contrib import messages

from django.views.generic import (
    View, TemplateView, ListView, DetailView,
    CreateView, UpdateView, DeleteView, FormView
)

# Core
from apps.core.forms import TenantAwareModelForm
from apps.core.middleware.tenant import get_dynamic_tenant
from apps.core.permissions.mixins import (
    PermissionRequiredMixin, RoleRequiredMixin,
    TenantAccessMixin, ObjectPermissionMixin,
    RoleBasedViewMixin,TenantRequiredMixin
)
from apps.core.services.audit_service import AuditService
from apps.core.utils.tenant import get_current_tenant

# Project Models
from apps.students.models import Student
from apps.users.models import User
from apps.academics.models import SchoolClass
from apps.finance.models import Invoice, Payment
from apps.library.models import Book, BookIssue
from apps.hostel.models import Hostel, HostelAllocation
from apps.hr.models import Staff
from apps.inventory.models import Item
from apps.transportation.models import Vehicle, Route
from apps.events.models import Event
from apps.exams.models import Exam
from apps.security.models import SecurityIncident, AuditLog


logger = logging.getLogger(__name__)

# ============================================================================
# BASE VIEW CLASSES
# ============================================================================

class BaseView(TenantRequiredMixin, RoleBasedViewMixin, View):
    """
    Base view with authentication, role checking, and audit logging
    """
    
    # Configuration
    permission_required = None
    roles_required = None
    min_role_level = None
    allow_superuser = True
    raise_exception = True
    
    # Tenant configuration
    tenant_field = 'tenant'
    tenant_required = True
    
    # Audit configuration
    audit_enabled = True
    audit_resource_type = None  # Auto-detected from model
    audit_action = None  # Auto-detected from HTTP method
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.tenant = None
    
    def setup(self, request, *args, **kwargs):
        """Initialize view with request context"""
        super().setup(request, *args, **kwargs)
        
        # Set tenant from request
        self.tenant = getattr(request, 'tenant', None)
        
        # Validate tenant if required
        if self.tenant_required and not self.tenant:
            logger.warning(f"Tenant context missing in {self.__class__.__name__}")
    
    def get_audit_resource_type(self):
        """Get resource type for audit logging"""
        if self.audit_resource_type:
            return self.audit_resource_type
        
        # Try to determine from model
        if hasattr(self, 'model'):
            return self.model.__name__
        
        # Use class name as fallback
        return self.__class__.__name__.replace('View', '')
    
    def get_audit_action(self):
        """Get action for audit logging"""
        if self.audit_action:
            return self.audit_action
        
        # Map HTTP method to audit action
        method = self.request.method.upper()
        action_map = {
            'GET': 'READ',
            'POST': 'CREATE',
            'PUT': 'UPDATE',
            'PATCH': 'UPDATE',
            'DELETE': 'DELETE'
        }
        return action_map.get(method, 'READ')
    
    def dispatch(self, request, *args, **kwargs):
        """Dispatch request with enhanced logging and audit"""
        try:
            # Log request
            if settings.DEBUG:
                logger.debug(f"Processing {request.method} request for {self.__class__.__name__}")
            
            # Check permissions
            if not self.has_permission():
                if self.raise_exception:
                    raise PermissionDenied(
                        f"You don't have permission to access {self.__class__.__name__}"
                    )
                return self.handle_no_permission()
            
            # Check role permissions
            if not self.has_role_permission():
                if self.raise_exception:
                    raise PermissionDenied(
                        f"Your role doesn't have access to {self.__class__.__name__}"
                    )
                return self.handle_no_permission()
            
            # Process request
            response = super().dispatch(request, *args, **kwargs)
            
            # Audit the request
            if self.audit_enabled and request.user.is_authenticated:
                try:
                    AuditService.create_audit_entry(
                        action=self.get_audit_action(),
                        resource_type=self.get_audit_resource_type(),
                        user=request.user,
                        request=request,
                        severity='INFO',
                        extra_data={
                            'view_class': self.__class__.__name__,
                            'http_method': request.method,
                            'status_code': response.status_code,
                            'path': request.path,
                        }
                    )
                except Exception as e:
                    logger.error(f"Failed to create audit entry: {e}")
            
            return response
            
        except PermissionDenied as e:
            logger.warning(f"Permission denied for user {request.user}: {e}")
            return self.handle_no_permission()
            
        except Exception as e:
            logger.error(f"Error in {self.__class__.__name__}: {e}", exc_info=True)
            raise
    
    def get_context_data(self, **kwargs):
        """Add common context to all views"""
        context = super().get_context_data(**kwargs) if hasattr(super(), 'get_context_data') else {}
        
        # Add tenant to context
        if self.tenant:
            context['tenant'] = self.tenant
        
        # Add user to context
        if self.request.user.is_authenticated:
            context['current_user'] = self.request.user
        
        # Add current path for navigation
        context['current_path'] = self.request.path
        
        # Add settings for templates
        from django.conf import settings
        context['DEBUG'] = settings.DEBUG
        context['APP_VERSION'] = getattr(settings, 'APP_VERSION', '1.0.0')
        
        return context
    
    def get_success_url(self):
        """Default success URL - override in subclasses"""
        return reverse_lazy('home')
    
    def handle_no_permission(self):
        """Handle permission denied"""
        messages.error(
            self.request,
            "You don't have permission to access this page."
        )
        return redirect('login')


class BaseTemplateView(BaseView, TemplateView):
    """Base template view with tenant and security integration"""
    
    def get_template_names(self):
        """Get template names with tenant-specific overrides"""
        template_names = super().get_template_names()
        
        # Add tenant-specific template if tenant exists
        if self.tenant and hasattr(self.tenant, 'schema_name'):
            tenant_template = f"{self.template_name.split('.')[0]}_{self.tenant.schema_name}.html"
            template_names.insert(0, tenant_template)
        
        return template_names


class BaseListView(BaseView, ListView):
    """
    Base list view with tenant isolation, pagination, and filtering
    """
    
    # Configuration
    # paginate_by = 25
    # paginate_orphans = 5
    ordering = ['-created_at']
    context_object_name = 'object_list'
    
    # Search and filter configuration
    search_fields = []
    filter_form_class = None
    
    def get_queryset(self):
        """Get tenant-isolated queryset"""
        queryset = super().get_queryset() if hasattr(super(), 'get_queryset') else self.model.objects.all()
        
        # Apply tenant filtering
        if self.tenant and hasattr(self.model, 'tenant'):
            queryset = queryset.filter(tenant=self.tenant)
        
        # Apply active filter for soft-delete models
        try:
            self.model._meta.get_field('is_active')
            queryset = queryset.filter(is_active=True)
        except Exception:
            # Field doesn't exist or is not a database field
            pass
        
        # Apply search
        search_query = self.request.GET.get('q', '')
        if search_query and self.search_fields:
            from django.db.models import Q
            search_q = Q()
            for field in self.search_fields:
                search_q |= Q(**{f"{field}__icontains": search_query})
            queryset = queryset.filter(search_q)
        
        # Apply ordering
        if self.ordering:
            queryset = queryset.order_by(*self.ordering)
        
        # Apply additional filtering
        queryset = self.apply_additional_filters(queryset)
        
        return queryset
    
    def apply_additional_filters(self, queryset):
        """Override to add custom filtering logic"""
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add pagination and filtering context"""
        context = super().get_context_data(**kwargs)
        
        # Add filter form if specified
        if self.filter_form_class:
            filter_form = self.filter_form_class(
                data=self.request.GET,
                tenant=self.tenant
            )
            if filter_form.is_valid():
                context['filter_form'] = filter_form
                # Apply filters to queryset
                context['object_list'] = filter_form.get_filtered_queryset(
                    context['object_list']
                )
            else:
                context['filter_form'] = filter_form
        
        # Add search query
        context['search_query'] = self.request.GET.get('q', '')
        
        # Add pagination information
        page_obj = context.get('page_obj')
        if page_obj:
            context['page_range'] = self.get_page_range(page_obj)
            context['total_items'] = page_obj.paginator.count
            context['items_per_page'] = page_obj.paginator.per_page
        
        return context
    
    def get_page_range(self, page_obj, delta=2):
        """Generate page range for pagination"""
        current = page_obj.number
        total = page_obj.paginator.num_pages
        
        start = max(1, current - delta)
        end = min(total, current + delta)
        
        return range(start, end + 1)
    
    def render_to_response(self, context, **response_kwargs):
        """Handle AJAX requests for pagination"""
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Return JSON for AJAX requests
            data = {
                'object_list': [
                    obj.to_dict() if hasattr(obj, 'to_dict') 
                    else {'id': str(obj.id), 'name': str(obj)}
                    for obj in context['object_list']
                ],
                'page_obj': {
                    'number': context['page_obj'].number,
                    'has_previous': context['page_obj'].has_previous(),
                    'has_next': context['page_obj'].has_next(),
                    'num_pages': context['page_obj'].paginator.num_pages,
                } if 'page_obj' in context else None,
                'paginator': {
                    'count': context['paginator'].count,
                    'num_pages': context['paginator'].num_pages,
                } if 'paginator' in context else None,
            }
            return JsonResponse(data, safe=False)
        
        return super().render_to_response(context, **response_kwargs)


class BaseDetailView(BaseView, DetailView):
    """
    Base detail view with object permission checking
    """
    
    object_permission_required = None
    
    def get_queryset(self):
        """Get tenant-isolated queryset"""
        queryset = super().get_queryset() if hasattr(super(), 'get_queryset') else self.model.objects.all()
        
        # Apply tenant filtering
        if self.tenant and hasattr(self.model, 'tenant'):
            queryset = queryset.filter(tenant=self.tenant)
        
        # Apply active filter for soft-delete models
        if hasattr(self.model, 'is_active'):
            queryset = queryset.filter(is_active=True)
        
        return queryset
    
    def get_object(self, queryset=None):
        """Get object with permission checking"""
        obj = super().get_object(queryset)
        
        # Check object-level permissions
        if not self.has_object_permission(obj):
            raise PermissionDenied(
                f"You don't have permission to access this {self.model.__name__}"
            )
        
        # Log object access
        if self.audit_enabled:
            AuditService.create_audit_entry(
                action='READ',
                resource_type=self.model.__name__,
                user=self.request.user,
                request=self.request,
                instance=obj,
                extra_data={'detail_view': True}
            )
        
        return obj
    
    def has_object_permission(self, obj):
        """Check object-level permissions"""
        user = self.request.user
        
        if user.is_superuser:
            return True
        
        if self.object_permission_required:
            return user.has_perm(self.object_permission_required, obj)
        
        # Default: check if user is in the same tenant
        if hasattr(obj, 'tenant') and hasattr(user, 'tenant'):
            return obj.tenant == user.tenant
        
        # Check if user created the object
        if hasattr(obj, 'created_by'):
            return obj.created_by == user
        
        return True


class BaseCreateView(BaseView, CreateView):
    """
    Base create view with tenant integration and audit logging
    """
    
    # Form configuration
    form_class = None  # Should be a TenantAwareModelForm subclass
    
    def get_form_class(self):
        """Get form class with tenant integration"""
        if self.form_class:
            return self.form_class
        
        # Auto-generate form from model
        from django.forms import modelform_factory
        from apps.core.forms import TenantAwareModelForm
        
        class AutoForm(TenantAwareModelForm):
            class Meta:
                model = self.model
                fields = '__all__'
                exclude = ['tenant', 'created_by', 'updated_by']
        
        return AutoForm
    
    def get_form_kwargs(self):
        """Add tenant and user to form kwargs"""
        kwargs = super().get_form_kwargs()
        
        # Add tenant to form
        if self.tenant:
            kwargs['tenant'] = self.tenant
        
        # Add user to form
        if self.request.user.is_authenticated:
            kwargs['user'] = self.request.user
        
        return kwargs
    
    def form_valid(self, form):
        """Handle successful form submission"""
        response = super().form_valid(form)
        
        # Add success message
        model_name = self.model.__name__ if hasattr(self, 'model') else 'Object'
        messages.success(
            self.request,
            f"{model_name} created successfully!"
        )
        
        # Audit the creation
        if self.audit_enabled:
            AuditService.log_creation(
                user=self.request.user,
                instance=form.instance,
                request=self.request,
                extra_data={'created_via': 'web_form'}
            )
        
        return response
    
    def form_invalid(self, form):
        """Handle invalid form submission"""
        # Log form errors
        logger.warning(f"Form validation failed: {form.errors}")
        
        # Add error message
        messages.error(
            self.request,
            "Please correct the errors below."
        )
        
        # Add non-field errors as specific messages
        if '__all__' in form.errors:
            for error in form.errors['__all__']:
                messages.error(self.request, error)
        
        return super().form_invalid(form)
    
    def get_success_url(self):
        """Get success URL - override in subclasses"""
        if self.success_url:
            return super().get_success_url()

        if hasattr(self, 'model'):
            return reverse_lazy(f'{self.model._meta.app_label}:{self.model._meta.model_name}_list')
        return super().get_success_url()


class BaseUpdateView(BaseView, UpdateView):
    """
    Base update view with tenant isolation and change tracking
    """
    
    # Form configuration
    form_class = None  # Should be a TenantAwareModelForm subclass
    audit_changes = True
    
    def get_form_class(self):
        """Get form class with tenant integration"""
        if self.form_class:
            return self.form_class
        
        # Auto-generate form from model
        from django.forms import modelform_factory
        from apps.core.forms import TenantAwareModelForm
        
        class AutoForm(TenantAwareModelForm):
            class Meta:
                model = self.model
                fields = '__all__'
                exclude = ['tenant', 'created_by', 'updated_by']
        
        return AutoForm
    
    def get_form_kwargs(self):
        """Add tenant to form kwargs"""
        kwargs = super().get_form_kwargs()
        
        if self.tenant:
            kwargs['tenant'] = self.tenant
        
        return kwargs
    
    def get_queryset(self):
        """Get tenant-isolated queryset"""
        queryset = super().get_queryset() if hasattr(super(), 'get_queryset') else self.model.objects.all()
        
        # Apply tenant filtering
        if self.tenant and hasattr(self.model, 'tenant'):
            queryset = queryset.filter(tenant=self.tenant)
        
        return queryset
    
    def form_valid(self, form):
        """Handle successful form update with change tracking"""
        # Store old instance for change tracking
        old_instance = None
        if self.audit_changes:
            old_instance = self.model.objects.get(pk=form.instance.pk)
        
        response = super().form_valid(form)
        
        # Add success message
        model_name = self.model.__name__ if hasattr(self, 'model') else 'Object'
        messages.success(
            self.request,
            f"{model_name} updated successfully!"
        )
        
        # Audit the update with changes
        if self.audit_enabled and old_instance:
            AuditService.log_update(
                user=self.request.user,
                instance=form.instance,
                old_instance=old_instance,
                request=self.request,
                extra_data={'updated_via': 'web_form'}
            )
        
        return response
    
    def get_success_url(self):
        """Get success URL"""
        if self.success_url:
            return super().get_success_url()

        if hasattr(self.object, 'get_absolute_url'):
            return self.object.get_absolute_url()
        
        if hasattr(self, 'model'):
            return reverse_lazy(
                f'{self.model._meta.app_label}:{self.model._meta.model_name}_detail',
                kwargs={'pk': self.object.pk}
            )
        
        return super().get_success_url()

    def form_invalid(self, form):
        """Handle invalid form submission"""
        # Log form errors
        logger.warning(f"Form validation failed: {form.errors}")
        
        # Add error message
        messages.error(
            self.request,
            "Please correct the errors below."
        )
        
        # Add non-field errors as specific messages
        if '__all__' in form.errors:
            for error in form.errors['__all__']:
                messages.error(self.request, error)
        
        return super().form_invalid(form)

# In apps/core/views.py, update the BaseDeleteView class:

class BaseDeleteView(BaseView, DeleteView):
    """
    Base delete view with soft delete support and audit logging
    """
    
    # Delete configuration
    soft_delete = True  # Use soft delete if model supports it
    success_message = "Object deleted successfully."
    require_delete_reason = True  # Set to True for models with SoftDeleteModel
    
    def get_queryset(self):
        """Get tenant-isolated queryset"""
        queryset = super().get_queryset() if hasattr(super(), 'get_queryset') else self.model.objects.all()
        
        # Apply tenant filtering
        if self.tenant and hasattr(self.model, 'tenant'):
            queryset = queryset.filter(tenant=self.tenant)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add delete confirmation context"""
        context = super().get_context_data(**kwargs)
        
        # Add delete reason fields if required
        if self.require_delete_reason:
            context['require_reason'] = True
        
        # Add deletion categories for forms
        if hasattr(self.model, 'DELETION_CATEGORIES'):
            context['deletion_categories'] = self.model.DELETION_CATEGORIES
        else:
            context['deletion_categories'] = [
                ('USER_REQUEST', 'User Request'),
                ('ADMIN_ACTION', 'Administrative Action'),
                ('SYSTEM_CLEANUP', 'System Cleanup'),
                ('COMPLIANCE', 'Compliance Requirement'),
                ('OTHER', 'Other')
            ]
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Handle deletion with soft delete support"""
        self.object = self.get_object()
        
        try:
            if self.soft_delete and hasattr(self.object, 'is_active'):
                # Use soft delete with reason
                delete_reason = request.POST.get('deletion_reason', '')
                delete_category = request.POST.get('deletion_category', '')
                
                if self.require_delete_reason and (not delete_reason or not delete_category):
                    messages.error(
                        request,
                        "Deletion reason and category are required."
                    )
                    return self.render_to_response(self.get_context_data())
                
                # Call the model's delete method with parameters
                self.object.delete(
                    user=request.user,
                    reason=delete_reason,
                    category=delete_category
                )
                
                # Audit soft deletion
                if self.audit_enabled:
                    AuditService.log_deletion(
                        user=request.user,
                        instance=self.object,
                        request=request,
                        hard_delete=False,
                        extra_data={
                            'deletion_reason': delete_reason,
                            'deletion_category': delete_category
                        }
                    )
                
            else:
                # Hard delete (for models without soft delete)
                success_url = self.get_success_url()
                self.object.delete()
                
                # Audit hard deletion
                if self.audit_enabled:
                    AuditService.log_deletion(
                        user=request.user,
                        instance=self.object,
                        request=request,
                        hard_delete=True
                    )
            
            # Add success message
            messages.success(request, self.success_message)
            
            return redirect(self.get_success_url())
            
        except ValidationError as e:
            messages.error(request, f"Deletion failed: {e}")
            return self.render_to_response(self.get_context_data())
        
        except Exception as e:
            logger.error(f"Delete error: {e}", exc_info=True)
            messages.error(request, "An error occurred during deletion.")
            return self.render_to_response(self.get_context_data())


class BaseFormView(BaseView, FormView):
    """
    Base form view for non-model forms
    """
    
    def form_valid(self, form):
        """Handle successful form submission"""
        response = super().form_valid(form)
        
        # Add success message
        messages.success(
            self.request,
            "Action completed successfully!"
        )
        
        # Audit the action
        if self.audit_enabled:
            AuditService.create_audit_entry(
                action='FORM_SUBMIT',
                resource_type=self.get_audit_resource_type(),
                user=self.request.user,
                request=self.request,
                severity='INFO',
                extra_data={'form_class': form.__class__.__name__}
            )
        
        return response
    
    def form_invalid(self, form):
        """Handle invalid form submission"""
        messages.error(
            self.request,
            "Please correct the errors below."
        )
        return super().form_invalid(form)


# ============================================================================
# AJAX VIEW CLASSES
# ============================================================================

class BaseAjaxView(BaseView):
    """
    Base view for AJAX requests
    """
    
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        """Ensure AJAX requests only"""
        if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'error': 'This endpoint only accepts AJAX requests'
            }, status=400)
        
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        """Handle AJAX GET requests"""
        try:
            data = self.get_ajax_data(request, *args, **kwargs)
            return JsonResponse({
                'success': True,
                'data': data
            })
        except Exception as e:
            logger.error(f"AJAX GET error: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    def post(self, request, *args, **kwargs):
        """Handle AJAX POST requests"""
        try:
            data = self.post_ajax_data(request, *args, **kwargs)
            return JsonResponse({
                'success': True,
                'data': data
            })
        except Exception as e:
            logger.error(f"AJAX POST error: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    def get_ajax_data(self, request, *args, **kwargs):
        """Override to provide data for GET requests"""
        return {}
    
    def post_ajax_data(self, request, *args, **kwargs):
        """Override to handle POST data"""
        return {}


class BaseModelAjaxView(BaseAjaxView):
    """
    Base AJAX view for model operations
    """
    
    model = None
    
    def get_queryset(self):
        """Get tenant-isolated queryset"""
        if not self.model:
            raise NotImplementedError("Model must be specified")
        
        queryset = self.model.objects.all()
        
        # Apply tenant filtering
        if self.tenant and hasattr(self.model, 'tenant'):
            queryset = queryset.filter(tenant=self.tenant)
        
        # Apply active filter for soft-delete models
        if hasattr(self.model, 'is_active'):
            queryset = queryset.filter(is_active=True)
        
        return queryset


# ============================================================================
# API VIEW CLASSES
# ============================================================================

class BaseAPIView(BaseView):
    """
    Base view for API endpoints
    """
    
    # API configuration
    authentication_classes = []  # Override for API authentication
    permission_classes = []  # Override for API permissions
    throttle_classes = []  # Override for rate limiting
    
    def dispatch(self, request, *args, **kwargs):
        """Handle API requests"""
        # Set content type for API responses
        request.accepted_renderer = 'json'
        return super().dispatch(request, *args, **kwargs)
    
    def handle_exception(self, exc):
        """Handle exceptions in API requests"""
        logger.error(f"API error: {exc}", exc_info=True)
        
        if isinstance(exc, PermissionDenied):
            return JsonResponse({
                'error': 'Permission denied',
                'detail': str(exc)
            }, status=403)
        
        if isinstance(exc, Http404):
            return JsonResponse({
                'error': 'Not found',
                'detail': str(exc)
            }, status=404)
        
        if isinstance(exc, ValidationError):
            return JsonResponse({
                'error': 'Validation error',
                'detail': exc.message_dict if hasattr(exc, 'message_dict') else str(exc)
            }, status=400)
        
        # Generic error
        return JsonResponse({
            'error': 'Internal server error',
            'detail': str(exc) if settings.DEBUG else 'An error occurred'
        }, status=500)
    
    def get_paginated_response(self, data):
        """Return paginated response"""
        from django.core.paginator import Paginator
        
        page = self.request.GET.get('page', 1)
        per_page = self.request.GET.get('per_page', 25)
        
        paginator = Paginator(data, per_page)
        page_obj = paginator.get_page(page)
        
        return JsonResponse({
            'count': paginator.count,
            'next': page_obj.next_page_number() if page_obj.has_next() else None,
            'previous': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'results': list(page_obj.object_list),
            'page': page_obj.number,
            'pages': paginator.num_pages,
        })


# ============================================================================
# EXPORT MIXINS
# ============================================================================

class ExportMixin:
    """
    Mixin for adding export functionality to views
    """
    
    export_formats = ['csv', 'excel', 'pdf']
    export_filename = 'export'
    
    def get_export_queryset(self):
        """Get queryset for export"""
        return self.get_queryset()
    
    def export_csv(self, request, *args, **kwargs):
        """Export data as CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        headers = self.get_export_headers()
        writer.writerow(headers)
        
        # Write data
        for obj in self.get_export_queryset():
            writer.writerow(self.get_export_row(obj))
        
        return response
    
    def export_excel(self, request, *args, **kwargs):
        """Export data as Excel"""
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Write headers
        headers = self.get_export_headers()
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data
        for row_idx, obj in enumerate(self.get_export_queryset(), 2):
            row_data = self.get_export_row(obj)
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.xlsx"'
        wb.save(response)
        
        return response
    
    def get_export_headers(self):
        """Get headers for export"""
        raise NotImplementedError("Subclasses must implement get_export_headers")
    
    def get_export_row(self, obj):
        """Get row data for export"""
        raise NotImplementedError("Subclasses must implement get_export_row")
    
    def export(self, request, *args, **kwargs):
        """Handle export request"""
        export_format = request.GET.get('format', 'csv')
        
        if export_format == 'csv':
            return self.export_csv(request, *args, **kwargs)
        elif export_format == 'excel':
            return self.export_excel(request, *args, **kwargs)
        elif export_format == 'pdf':
            # Implement PDF export if needed
            pass
        
        return JsonResponse({'error': 'Invalid export format'}, status=400)


# ============================================================================
# BULK OPERATION MIXINS
# ============================================================================

class BulkOperationMixin:
    """
    Mixin for bulk operations on model instances
    """
    
    def post(self, request, *args, **kwargs):
        """Handle bulk operations"""
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_ids[]')
        
        if not action or not selected_ids:
            return JsonResponse({
                'success': False,
                'error': 'Action and selected items required'
            })
        
        try:
            with transaction.atomic():
                result = self.perform_bulk_action(action, selected_ids)
            
            # Audit bulk operation
            if self.audit_enabled:
                AuditService.create_audit_entry(
                    action='BULK_OPERATION',
                    resource_type=self.get_audit_resource_type(),
                    user=request.user,
                    request=request,
                    severity='INFO',
                    extra_data={
                        'action': action,
                        'count': len(selected_ids),
                        'selected_ids': selected_ids
                    }
                )
            
            return JsonResponse({
                'success': True,
                'message': f'Bulk action completed: {result}'
            })
            
        except Exception as e:
            logger.error(f"Bulk operation error: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    def perform_bulk_action(self, action, selected_ids):
        """Perform bulk action - override in subclasses"""
        raise NotImplementedError("Subclasses must implement perform_bulk_action")


class MasterDashboardView(TenantRequiredMixin, TemplateView):
    template_name = 'dashboard/master_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_current_tenant()
        
        # Students Statistics
        context['total_students'] = Student.objects.filter(tenant=tenant).count()
        context['active_students'] = Student.objects.filter(tenant=tenant, status='ACTIVE').count()
        context['alumni'] = Student.objects.filter(tenant=tenant, status='ALUMNI').count()
        
        # Users Statistics
        context['total_users'] = User.objects.filter(tenant=tenant).count()
        context['active_users'] = User.objects.filter(tenant=tenant, is_active=True).count()
        context['staff_count'] = Staff.objects.filter(tenant=tenant, is_active=True).count()
        
        # Academics Statistics
        context['total_courses'] = Course.objects.filter(tenant=tenant).count()
        context['total_classes'] = Class.objects.filter(tenant=tenant).count()
        
        # Finance Statistics
        context['total_invoices'] = Invoice.objects.filter(tenant=tenant).count()
        context['total_revenue'] = Payment.objects.filter(tenant=tenant).aggregate(
            total=Sum('amount')
        )['total'] or 0
        context['pending_invoices'] = Invoice.objects.filter(
            tenant=tenant, status='PENDING'
        ).count()
        
        # Library Statistics
        context['total_books'] = Book.objects.filter(tenant=tenant).count()
        context['books_issued'] = BookIssue.objects.filter(
            tenant=tenant, return_date__isnull=True
        ).count()
        
        # Hostel Statistics
        context['total_hostels'] = Hostel.objects.filter(tenant=tenant).count()
        context['hostel_allocations'] = HostelAllocation.objects.filter(
            tenant=tenant, is_active=True
        ).count()
        
        # Inventory Statistics
        context['total_items'] = Item.objects.filter(tenant=tenant, is_active=True).count()
        context['low_stock_items'] = Item.objects.filter(
            tenant=tenant, is_active=True, current_stock__lte=10
        ).count()
        
        # Transportation Statistics
        context['total_vehicles'] = Vehicle.objects.filter(tenant=tenant, is_active=True).count()
        context['total_routes'] = Route.objects.filter(tenant=tenant, is_active=True).count()
        
        # Events Statistics
        context['upcoming_events'] = Event.objects.filter(
            tenant=tenant, start_date__gte=timezone.now()
        ).count()
        
        # Exams Statistics
        context['upcoming_exams'] = Exam.objects.filter(
            tenant=tenant, start_date__gte=timezone.now()
        ).count()
        
        # Security Statistics
        context['open_incidents'] = SecurityIncident.objects.filter(
            tenant=tenant, status='OPEN'
        ).count()
        context['recent_audits'] = AuditLog.objects.filter(tenant=tenant).count()
        
        # Chart Data - Student Status Distribution
        student_status = Student.objects.filter(tenant=tenant).values('status').annotate(
            count=Count('id')
        )
        context['student_status_labels'] = [item['status'] for item in student_status]
        context['student_status_data'] = [item['count'] for item in student_status]
        
        # Chart Data - User Roles Distribution
        user_roles = User.objects.filter(tenant=tenant).values('role').annotate(
            count=Count('id')
        )
        context['user_roles_labels'] = [item['role'] for item in user_roles]
        context['user_roles_data'] = [item['count'] for item in user_roles]
        
        # Recent Activities (last 10 audit logs)
        context['recent_activities'] = AuditLog.objects.filter(
            tenant=tenant
        ).order_by('-created_at')[:10]
        
        return context

# ============================================
# CUSTOM ERROR VIEWS
# ============================================



def custom_page_not_found_view(request, exception):
    """Custom 404 error handler"""
    return render(request, 'errors/404.html', status=404)

def custom_error_view(request):
    """Custom 500 error handler"""
    return render(request, 'errors/500.html', status=500)

def custom_permission_denied_view(request, exception):
    """Custom 403 error handler"""
    return render(request, 'errors/403.html', status=403)

def custom_bad_request_view(request, exception):
    """Custom 400 error handler"""
    return render(request, 'errors/400.html', status=400)