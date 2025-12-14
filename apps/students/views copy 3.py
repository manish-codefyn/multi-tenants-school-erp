from datetime import datetime
import csv
import openpyxl
import io
import logging
import re
from django.utils import timezone
from django.conf import Settings
logger = logging.getLogger(__name__)

from django.core.exceptions import ValidationError
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse, FileResponse, Http404, JsonResponse
from django.urls import reverse_lazy

# Core imports
from apps.core.views import BaseView, BaseListView, BaseCreateView, BaseUpdateView, BaseDeleteView,BaseDetailView
from apps.core.services.audit_service import AuditService
from apps.core.middleware.tenant import get_dynamic_tenant
from apps.core.permissions.mixins import ( PermissionRequiredMixin, RoleRequiredMixin, 
TenantAccessMixin, ObjectPermissionMixin,
RoleBasedViewMixin,TenantRequiredMixin )

# Model imports
from apps.academics.models import AcademicYear, SchoolClass, Section, Stream
from .models import Student, Guardian, StudentAddress, StudentDocument
from .forms import StudentForm, GuardianForm, StudentDocumentForm, StudentFilterForm
from .idcard import StudentIDCardGenerator


# ============================================================================
# DASHBOARD VIEWS
# ============================================================================

class StudentDashboardView(BaseView):
    """Comprehensive student dashboard with analytics"""
    template_name = 'students/dashboard.html'
    permission_required = 'students.view_student_dashboard'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = self.request.tenant
        
        # Get statistics with audit logging
        students = Student.get_secure_queryset(self.request.user)
        
        context['total_students'] = students.count()
        context['active_students'] = students.filter(status='ACTIVE').count()
        context['inactive_students'] = students.filter(status='INACTIVE').count()
        context['graduated_students'] = students.filter(status='GRADUATED').count()
        
        # Class distribution with audit
        class_stats = []
        for school_class in SchoolClass.objects.filter(tenant=tenant):
            count = students.filter(current_class=school_class).count()
            class_stats.append({
                'name': school_class.name,
                'count': count,
                'percentage': (count / context['total_students'] * 100) if context['total_students'] > 0 else 0
            })
        context['class_stats'] = class_stats
        
        # Gender distribution
        gender_stats = students.values('gender').annotate(count=models.Count('id'))
        context['gender_stats'] = list(gender_stats)
        
        # Recent activities (from audit logs)
        context['recent_activities'] = self.get_recent_activities()
        
        # Audit the dashboard access
        AuditService.create_audit_entry(
            action='READ',
            resource_type='StudentDashboard',
            user=self.request.user,
            request=self.request,
            severity='INFO',
            extra_data={'dashboard_type': 'student_overview'}
        )
        
        return context
    
    def get_recent_activities(self):
        """Get recent student-related activities from audit logs"""
        try:
            from apps.core.models import AuditLog
            return AuditLog.objects.filter(
                resource_type='Student',
                tenant_id=self.request.tenant.id
            ).order_by('-timestamp')[:10]
        except Exception:
            return []


# ============================================================================
# CRUD VIEWS WITH TENANT & AUDIT INTEGRATION
# ============================================================================

class StudentListView(BaseListView):
    """Secure student list with tenant isolation and audit logging"""
    model = Student
    template_name = 'students/student_list.html'
    context_object_name = 'students'
    permission_required = 'students.view_student'
    
    def get_queryset(self):
        # Use the secure queryset method from BaseModel
        queryset = Student.get_secure_queryset(self.request.user)
        
        # Apply filters
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(admission_number__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(personal_email__icontains=search) |
                Q(mobile_primary__icontains=search)
            )
        
        # Apply other filters
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        class_id = self.request.GET.get('class_id')
        if class_id:
            queryset = queryset.filter(current_class_id=class_id)
        
        section_id = self.request.GET.get('section_id')
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        
        return queryset.select_related(
            'current_class', 'section', 'academic_year', 'stream'
        ).prefetch_related('guardians')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = Student.STATUS_CHOICES
        context['filter_form'] = StudentFilterForm(
            tenant=self.request.tenant,
            initial=self.request.GET.dict()
        )
        
        # Add classes for filter dropdown
        from apps.academics.models import SchoolClass
        context['classes'] = SchoolClass.objects.filter(tenant=self.request.tenant)
        
        # Log the list view access
        AuditService.create_audit_entry(
            action='READ',
            resource_type='StudentList',
            user=self.request.user,
            request=self.request,
            severity='INFO',
            extra_data={
                'filter_count': self.get_queryset().count(),
                'filters_applied': dict(self.request.GET)
            }
        )
        
        return context


class StudentDetailView(BaseDetailView):
    """Detailed student view with all related information"""
    template_name = 'students/student_detail.html'
    permission_required = 'students.view_student'
    
    def get_object(self):
        """Get student with tenant isolation"""
        student_id = self.kwargs.get('pk')
        return get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=student_id
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        # Add all related data
        context['student'] = student
        context['guardians'] = student.guardians.all()
        context['addresses'] = student.addresses.all()
        context['documents'] = student.documents.all()
        
        # Academic information
        context['academic_history'] = student.academic_history.all()
        
        # Performance metrics
        context['attendance_rate'] = self.calculate_attendance_rate(student)
        context['academic_performance'] = self.get_academic_performance(student)
        
        # Audit the detail view access
        AuditService.log_update(
            user=self.request.user,
            instance=student,
            request=self.request,
            extra_data={'action': 'viewed_details'}
        )
        
        return context
    
    def calculate_attendance_rate(self, student):
        """Calculate student attendance rate"""
        # This would integrate with attendance module
        try:
            from apps.attendance.models import Attendance
            total_days = Attendance.objects.filter(
                student=student,
                academic_year=student.academic_year
            ).count()
            present_days = Attendance.objects.filter(
                student=student,
                academic_year=student.academic_year,
                status='PRESENT'
            ).count()
            
            return (present_days / total_days * 100) if total_days > 0 else 0
        except Exception:
            return 0
    
    def get_academic_performance(self, student):
        """Get student academic performance"""
        # This would integrate with exams module
        try:
            from apps.exams.models import ExamResult
            results = ExamResult.objects.filter(
                student=student,
                is_published=True
            ).order_by('-exam__start_date')[:5]
            return results
        except Exception:
            return []


class StudentCreateView(BaseCreateView):
    """Secure student creation with tenant and audit integration"""
    form_class = StudentForm
    template_name = 'students/student_form.html'
    permission_required = 'students.add_student'
    success_url = reverse_lazy('students:student_list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.tenant
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Add New Student')
        context['is_create'] = True
        return context
    
    def form_valid(self, form):
        # Save but don't commit yet
        student = form.save(commit=False)

        # IMPORTANT: Assign tenant BEFORE clean() or save() runs
        student.tenant = self.request.tenant

        # Save to DB
        student.save()

        # Now save M2M fields if any
        form.save_m2m()

        # Log creation
        AuditService.log_creation(
            user=self.request.user,
            instance=student,
            request=self.request,
            extra_data={'created_via': 'web_form'}
        )

        messages.success(
            self.request,
            f"Student {student.full_name} created successfully!"
        )

        return redirect(self.get_success_url())


class StudentUpdateView(BaseUpdateView):
    """Secure student update with tenant isolation and audit logging"""
    form_class = StudentForm
    template_name = 'students/student_form.html'
    permission_required = 'students.change_student'
    
    def get_object(self):
        return get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs.get('pk')
        )
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.tenant
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Edit Student')
        context['is_create'] = False
        return context
    
    def form_valid(self, form):
        """Capture changes and audit the update"""
        old_instance = Student.objects.get(id=form.instance.id)
        
        response = super().form_valid(form)
        
        # Log the update with changes
        AuditService.log_update(
            user=self.request.user,
            instance=form.instance,
            old_instance=old_instance,
            request=self.request,
            extra_data={'updated_via': 'web_form'}
        )
        
        messages.success(
            self.request,
            f"Student {form.instance.full_name} updated successfully!"
        )
        
        return response
    
    def get_success_url(self):
        return reverse_lazy('students:student_detail', kwargs={'pk': self.object.id})


class StudentDeleteView(BaseDeleteView):
    """Secure student deletion with soft delete and audit logging"""
    template_name = 'students/student_confirm_delete.html'
    permission_required = 'students.delete_student'
    
    def get_object(self):
        return get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs.get('pk')
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = self.get_object()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Override delete to use soft delete with audit"""
        student = self.get_object()
        
        try:
            # Perform soft delete with reason
            student.delete(
                user=request.user,
                reason=request.POST.get('deletion_reason', 'Administrative action'),
                category=request.POST.get('deletion_category', 'ADMIN_ACTION')
            )
            
            # Log the deletion
            AuditService.log_deletion(
                user=request.user,
                instance=student,
                request=request,
                hard_delete=False,
                extra_data={
                    'deletion_reason': request.POST.get('deletion_reason'),
                    'deletion_category': request.POST.get('deletion_category')
                }
            )
            
            messages.success(
                request,
                f"Student {student.full_name} has been deleted successfully."
            )
            
        except ValidationError as e:
            messages.error(request, f"Deletion failed: {e}")
            return self.render_to_response(self.get_context_data())
        
        return redirect('students:student_list')
    
    def get_success_url(self):
        return reverse_lazy('students:student_list')


# ============================================================================
# GUARDIAN MANAGEMENT VIEWS
# ============================================================================

class GuardianCreateView(BaseCreateView):
    """Create guardian for student with tenant isolation"""
    form_class = GuardianForm
    template_name = 'students/guardian_form.html'
    permission_required = 'students.add_student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs['student_id']
        )
        context['student'] = student
        return context
    
    def form_valid(self, form):
        student = get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs['student_id']
        )
        
        form.instance.student = student
        form.instance.tenant = self.request.tenant
        
        response = super().form_valid(form)
        
        # Log guardian creation
        AuditService.log_creation(
            user=self.request.user,
            instance=form.instance,
            request=self.request,
            extra_data={'student_id': str(student.id)}
        )
        
        messages.success(
            self.request,
            f"Guardian {form.instance.full_name} added successfully!"
        )
        
        return response
    
    def get_success_url(self):
        return reverse_lazy('students:student_detail', kwargs={'pk': self.kwargs['student_id']})


class GuardianUpdateView(BaseUpdateView):
    """Update guardian information"""
    form_class = GuardianForm
    template_name = 'students/guardian_form.html'
    permission_required = 'students.change_student'
    
    def get_object(self):
        return get_object_or_404(
            Guardian.objects.filter(
                tenant=self.request.tenant,
                student_id=self.kwargs['student_id']
            ),
            id=self.kwargs['guardian_id']
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs['student_id']
        )
        context['is_update'] = True
        return context
    
    def get_success_url(self):
        return reverse_lazy('students:student_detail', kwargs={'pk': self.kwargs['student_id']})


# ============================================================================
# DOCUMENT MANAGEMENT VIEWS
# ============================================================================

class DocumentUploadView(BaseCreateView):
    """Secure document upload with validation and audit"""
    form_class = StudentDocumentForm
    template_name = 'students/document_form.html'
    permission_required = 'students.add_student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs['student_id']
        )
        return context
    
    def form_valid(self, form):
        student = get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs['student_id']
        )
        
        form.instance.student = student
        form.instance.tenant = self.request.tenant
        
        response = super().form_valid(form)
        
        # Log document upload
        AuditService.log_creation(
            user=self.request.user,
            instance=form.instance,
            request=self.request,
            extra_data={
                'student_id': str(student.id),
                'document_type': form.instance.doc_type,
                'file_size': form.instance.file.size if form.instance.file else 0
            }
        )
        
        messages.success(
            self.request,
            f"Document {form.instance.file_name} uploaded successfully!"
        )
        
        return response
    
    def get_success_url(self):
        return reverse_lazy('students:student_detail', kwargs={'pk': self.kwargs['student_id']})


class DocumentDownloadView(BaseView):
    """Secure document download with permission checking"""
    permission_required = 'students.view_student'
    
    def get(self, request, *args, **kwargs):
        document = get_object_or_404(
            StudentDocument.objects.filter(tenant=request.tenant),
            id=kwargs.get('document_id')
        )
        
        # Verify student access
        if not Student.get_secure_queryset(request.user).filter(id=document.student_id).exists():
            raise Http404(_("Document not found or access denied"))
        
        if not document.file:
            raise Http404(_("File not found"))
        
        # Log the download
        AuditService.create_audit_entry(
            action='READ',
            resource_type='StudentDocument',
            user=request.user,
            request=request,
            instance=document,
            extra_data={
                'student_id': str(document.student_id),
                'action': 'document_download'
            }
        )
        
        # Serve the file
        response = FileResponse(document.file)
        response['Content-Disposition'] = f'attachment; filename="{document.file_name}"'
        return response


# ============================================================================
# BULK OPERATIONS VIEWS
# ============================================================================

class StudentBulkUploadView(BaseView):
    """Comprehensive bulk upload with validation, error handling, and progress tracking"""
    template_name = 'students/student_bulk_upload.html'
    permission_required = 'students.add_student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = self.request.tenant
        
        # Get academic years
        context['academic_years'] = AcademicYear.objects.filter(
            tenant=tenant
        ).order_by('-is_current', '-start_date')
        
        # Get available classes
        context['classes'] = SchoolClass.objects.filter(
            tenant=tenant
        ).order_by('order', 'name')
        
        # Get available streams
        context['streams'] = Stream.objects.filter(
            tenant=tenant
        ).order_by('name')
        
        # Get recent uploads for context
        context['recent_uploads'] = self.get_recent_uploads()
        
        return context
    
    def get_recent_uploads(self):
        """Get recent bulk upload history from audit logs"""
        try:
            from apps.core.models import AuditLog
            return AuditLog.objects.filter(
                resource_type='Student',
                action='CREATE',
                tenant_id=self.request.tenant.id,
                extra_data__contains={'bulk_upload': True}
            ).order_by('-timestamp')[:5]
        except Exception:
            return []
    
    def post(self, request, *args, **kwargs):
        """Handle file upload and processing"""
        file = request.FILES.get('file')
        academic_year_id = request.POST.get('academic_year')
        update_existing = request.POST.get('update_existing') == 'on'
        skip_errors = request.POST.get('skip_errors') == 'on'
        
        if not file:
            messages.error(request, 'Please select a file to upload.')
            return self.render_to_response(self.get_context_data())
        
        # Validate file size (max 10MB)
        if file.size > 10 * 1024 * 1024:
            messages.error(request, 'File size exceeds 10MB limit.')
            return self.render_to_response(self.get_context_data())
        
        try:
            # Get academic year
            if academic_year_id:
                academic_year = AcademicYear.objects.filter(
                    id=academic_year_id,
                    tenant=request.tenant
                ).first()
                if not academic_year:
                    messages.error(request, 'Selected academic year not found.')
                    return self.render_to_response(self.get_context_data())
            else:
                academic_year = AcademicYear.objects.filter(
                    tenant=request.tenant, 
                    is_current=True
                ).first()
                if not academic_year:
                    messages.error(request, 'No active academic year found. Please select one.')
                    return self.render_to_response(self.get_context_data())
            
            # Process file based on type
            file_extension = file.name.split('.')[-1].lower()
            
            if file_extension == 'csv':
                result = self.process_csv(file, request.tenant, academic_year, update_existing, skip_errors)
            elif file_extension in ['xls', 'xlsx']:
                result = self.process_excel(file, request.tenant, academic_year, update_existing, skip_errors)
            else:
                messages.error(request, 'Invalid file format. Please upload CSV or Excel file.')
                return self.render_to_response(self.get_context_data())
            
            # Handle result
            return self.handle_upload_result(request, result)
            
        except Exception as e:
            logger.error(f"Bulk upload error: {str(e)}", exc_info=True)
            messages.error(request, f'Error processing file: {str(e)}')
            return self.render_to_response(self.get_context_data())
    
    def handle_upload_result(self, request, result):
        """Handle upload result and display appropriate messages"""
        if 'error' in result:
            messages.error(request, result['error'])
            if 'errors' in result and len(result['errors']) <= 10:
                for error in result['errors']:
                    messages.warning(request, error)
        elif 'success' in result:
            msg = result['success']
            if result.get('warnings'):
                msg += f" ({len(result['warnings'])} warnings)"
            messages.success(request, msg)
            
            # Store details in session for summary page
            upload_summary = {
                'created': result.get('created_count', 0),
                'updated': result.get('updated_count', 0),
                'skipped': result.get('skipped_count', 0),
                'total': result.get('total_rows', 0),
                'warnings': result.get('warnings', [])[:20],
                'errors': result.get('errors', [])[:10],
                'timestamp': timezone.now().isoformat(),
            }
            request.session['last_upload_summary'] = upload_summary
            
            # Log bulk upload operation
            AuditService.create_audit_entry(
                action='BULK_OPERATION',
                resource_type='Student',
                user=request.user,
                request=request,
                severity='INFO',
                extra_data={
                    'bulk_upload': True,
                    'summary': upload_summary,
                    'file_name': request.FILES['file'].name
                }
            )
            
            # Redirect based on warnings
            if result.get('warnings') or result.get('errors'):
                return redirect('students:bulk_upload_summary')
            else:
                return redirect('students:student_list')
        
        return self.render_to_response(self.get_context_data())
    
    def process_csv(self, file, tenant, academic_year, update_existing, skip_errors):
        """Process CSV file"""
        try:
            content = file.read()
            
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'windows-1252', 'cp1252']
            decoded_content = None
            
            for encoding in encodings:
                try:
                    decoded_content = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if decoded_content is None:
                return {'error': 'Unable to decode CSV file. Please save as UTF-8.'}
            
            # Clean BOM if present
            if decoded_content.startswith('\ufeff'):
                decoded_content = decoded_content[1:]
            
            csv_file = io.StringIO(decoded_content)
            
            # Detect delimiter
            sample = csv_file.read(2048)
            csv_file.seek(0)
            delimiter = self.detect_delimiter(sample)
            
            # Read CSV
            reader = csv.reader(csv_file, delimiter=delimiter)
            rows = list(reader)
            
            if len(rows) < 2:
                return {'error': 'CSV file is empty or has insufficient data.'}
            
            # Process headers
            headers = self.clean_headers(rows[0])
            
            # Check for required headers
            required_check = self.check_required_headers(headers)
            if not required_check['valid']:
                return {'error': required_check['message']}
            
            # Process data rows
            data = []
            for i in range(1, len(rows)):
                row = rows[i]
                
                # Skip empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                # Convert row to dictionary
                row_dict = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        value = str(cell).strip() if cell is not None else ''
                        row_dict[headers[j]] = value
                
                # Only add rows with at least first_name or last_name
                if row_dict.get('first_name') or row_dict.get('last_name'):
                    data.append(row_dict)
            
            if not data:
                return {'error': 'No valid student data found in CSV file.'}
            
            return self._create_students(data, tenant, academic_year, update_existing, skip_errors)
            
        except Exception as e:
            logger.error(f"CSV processing error: {str(e)}", exc_info=True)
            return {'error': f'Error processing CSV file: {str(e)}'}
    
    def process_excel(self, file, tenant, academic_year, update_existing, skip_errors):
        """Process Excel file"""
        try:
            wb = openpyxl.load_workbook(
                filename=io.BytesIO(file.read()),
                read_only=True,
                data_only=True
            )
            
            sheet = wb.active
            
            # Get headers
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower().replace(' ', '_'))
                else:
                    if len(headers) > 0 and cell.column > 30:
                        break
                    headers.append(f'column_{cell.column}')
            
            # Clean headers
            headers = self.clean_headers(headers)
            
            # Check required headers
            required_check = self.check_required_headers(headers)
            if not required_check['valid']:
                return {'error': required_check['message']}
            
            # Process data rows
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                    continue
                
                row_dict = {}
                for i, cell_value in enumerate(row):
                    if i < len(headers):
                        value = self.convert_excel_cell(cell_value, headers[i])
                        row_dict[headers[i]] = value
                
                # Only add rows with at least first_name or last_name
                if row_dict.get('first_name') or row_dict.get('last_name'):
                    data.append(row_dict)
            
            if not data:
                return {'error': 'No valid student data found in Excel file.'}
            
            return self._create_students(data, tenant, academic_year, update_existing, skip_errors)
            
        except Exception as e:
            logger.error(f"Excel processing error: {str(e)}", exc_info=True)
            return {'error': f'Error processing Excel file: {str(e)}'}
    
    def _create_students(self, data, tenant, academic_year, update_existing, skip_errors):
        """Main method to create/update students"""
        created_count = 0
        updated_count = 0
        skipped_count = 0
        warnings = []
        errors = []
        
        # Pre-fetch data for validation
        all_classes = {c.name.lower(): c for c in SchoolClass.objects.filter(tenant=tenant)}
        all_sections = {}
        for class_obj in all_classes.values():
            sections = {s.name.lower(): s for s in class_obj.sections.all()}
            all_sections[class_obj.name.lower()] = sections
        
        existing_students = {
            s.personal_email.lower(): s for s in Student.objects.filter(tenant=tenant)
        }
        
        try:
            with transaction.atomic():
                for row_index, row in enumerate(data, start=1):
                    try:
                        # Basic validation
                        first_name = str(row.get('first_name', '')).strip()
                        last_name = str(row.get('last_name', '')).strip()
                        
                        if not first_name and not last_name:
                            errors.append(f"Row {row_index}: Missing both first and last name")
                            if skip_errors:
                                skipped_count += 1
                                continue
                            else:
                                return {
                                    'error': f"Row {row_index}: Missing required fields",
                                    'errors': errors[:10]
                                }
                        
                        # Email handling
                        personal_email = str(row.get('personal_email', '')).strip().lower()
                        if not personal_email:
                            # Generate email
                            first_part = re.sub(r'[^a-z0-9]', '', first_name.lower())[:15]
                            last_part = re.sub(r'[^a-z0-9]', '', last_name.lower())[:15]
                            domain = getattr(tenant, 'domain', f"{tenant.schema_name}.edu")
                            personal_email = f"{first_part}.{last_part}{row_index}@{domain}"
                            warnings.append(f"Row {row_index}: Generated email: {personal_email}")
                        
                        # Check if student exists
                        existing_student = existing_students.get(personal_email)
                        
                        if existing_student and not update_existing:
                            errors.append(f"Row {row_index}: Student with email {personal_email} already exists")
                            if skip_errors:
                                skipped_count += 1
                                continue
                            else:
                                return {
                                    'error': f"Duplicate email found at row {row_index}",
                                    'errors': errors[:10]
                                }
                        
                        # Prepare student data
                        student_data = {
                            'tenant': tenant,
                            'first_name': first_name,
                            'last_name': last_name,
                            'personal_email': personal_email,
                            'academic_year': academic_year,
                            'gender': row.get('gender', 'U').upper()[:1],
                            'date_of_birth': self.parse_date(row.get('date_of_birth')),
                            'mobile_primary': str(row.get('mobile_primary', ''))[:15],
                            'current_class': all_classes.get(row.get('class_name', '').lower()),
                            'status': row.get('status', 'ACTIVE').upper(),
                            'category': row.get('category', 'GENERAL').upper(),
                        }
                        
                        # Handle class name if provided
                        class_name = row.get('class_name', '').strip()
                        if class_name:
                            class_key = class_name.lower()
                            if class_key in all_classes:
                                student_data['current_class'] = all_classes[class_key]
                            else:
                                warnings.append(f"Row {row_index}: Class '{class_name}' not found")
                        
                        # Handle section if provided
                        section_name = row.get('section_name', '').strip()
                        if student_data['current_class'] and section_name:
                            class_key = student_data['current_class'].name.lower()
                            section_key = section_name.lower()
                            if class_key in all_sections and section_key in all_sections[class_key]:
                                student_data['section'] = all_sections[class_key][section_key]
                            else:
                                warnings.append(f"Row {row_index}: Section '{section_name}' not found in class")
                        
                        # Create or update student
                        if existing_student and update_existing:
                            # Update existing
                            for key, value in student_data.items():
                                if key != 'tenant':
                                    setattr(existing_student, key, value)
                            existing_student.save()
                            updated_count += 1
                            
                            # Log individual update
                            AuditService.log_update(
                                user=None,  # System action
                                instance=existing_student,
                                request=None,
                                extra_data={'updated_via': 'bulk_upload'}
                            )
                        else:
                            # Create new
                            student = Student.objects.create(**student_data)
                            created_count += 1
                            existing_students[personal_email] = student
                            
                            # Log individual creation
                            AuditService.log_creation(
                                user=None,  # System action
                                instance=student,
                                request=None,
                                extra_data={'created_via': 'bulk_upload'}
                            )
                        
                    except Exception as e:
                        error_msg = f"Row {row_index}: {str(e)}"
                        errors.append(error_msg)
                        if skip_errors:
                            skipped_count += 1
                        else:
                            raise
        
            # Prepare final result
            result = {
                'created_count': created_count,
                'updated_count': updated_count,
                'skipped_count': skipped_count,
                'total_rows': len(data),
                'warnings': warnings[:50],
                'errors': errors[:50] if errors else None,
            }
            
            if errors and not skip_errors:
                result['error'] = f"Failed to process {len(errors)} row(s)"
            else:
                total_processed = created_count + updated_count
                result['success'] = (
                    f"Successfully processed {total_processed} student(s): "
                    f"Created {created_count}, Updated {updated_count}"
                )
            
            return result
            
        except Exception as e:
            logger.error(f"Bulk create error: {str(e)}", exc_info=True)
            return {
                'error': f'Critical error in transaction: {str(e)}',
                'created_count': created_count,
                'updated_count': updated_count,
                'skipped_count': skipped_count,
            }
    
    # Helper methods for bulk upload
    def detect_delimiter(self, sample):
        delimiters = [',', ';', '\t', '|']
        counts = {d: sample.count(d) for d in delimiters}
        return max(counts, key=counts.get) if counts else ','
    
    def clean_headers(self, headers):
        cleaned = []
        mapping = {
            'firstname': 'first_name',
            'lastname': 'last_name',
            'dob': 'date_of_birth',
            'email': 'personal_email',
            'phone': 'mobile_primary',
            'class': 'class_name',
            'section': 'section_name',
        }
        for header in headers:
            if header is None:
                cleaned.append('')
                continue
            header_str = str(header).strip().lower().replace(' ', '_')
            cleaned.append(mapping.get(header_str, header_str))
        return cleaned
    
    def check_required_headers(self, headers):
        required = ['first_name', 'last_name']
        missing = [h for h in required if h not in headers]
        return {
            'valid': len(missing) == 0,
            'message': f'Missing required columns: {", ".join(missing)}' if missing else ''
        }
    
    def convert_excel_cell(self, cell_value, header):
        if cell_value is None:
            return ''
        if header in ['date_of_birth']:
            if isinstance(cell_value, datetime):
                return cell_value.date().isoformat()
        return str(cell_value).strip()
    
    def parse_date(self, date_str):
        if not date_str:
            return None
        formats = ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']
        for fmt in formats:
            try:
                return datetime.strptime(str(date_str).strip(), fmt).date()
            except ValueError:
                continue
        return None


class BulkUploadSummaryView(BaseView):
    """View to show detailed upload summary"""
    template_name = 'students/bulk_upload_summary.html'
    permission_required = 'students.add_student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        summary = self.request.session.get('last_upload_summary', {})
        
        context['summary'] = summary
        context['has_warnings'] = bool(summary.get('warnings'))
        context['has_errors'] = bool(summary.get('errors'))
        context['success_rate'] = self.calculate_success_rate(summary)
        
        return context
    
    def calculate_success_rate(self, summary):
        total = summary.get('total', 0)
        processed = summary.get('created', 0) + summary.get('updated', 0)
        return (processed / total * 100) if total > 0 else 0


class BulkUploadSampleView(BaseView):
    """Download sample templates"""
    permission_required = 'students.add_student'
    
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'csv')
        template_type = request.GET.get('type', 'simple')
        
        if format_type == 'excel':
            return self.generate_excel_sample(template_type)
        else:
            return self.generate_csv_sample(template_type)
    
    def generate_csv_sample(self, template_type):
        response = HttpResponse(content_type='text/csv')
        filename = f'student_upload_{template_type}_template.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        if template_type == 'detailed':
            headers = [
                'first_name', 'last_name', 'date_of_birth', 'gender',
                'personal_email', 'mobile_primary', 'class_name', 'section_name',
                'status', 'category', 'admission_type'
            ]
            writer.writerow(headers)
            
            # Sample data
            sample_data = [
                ['John', 'Doe', '2010-01-15', 'M', 
                 'john.doe@example.com', '9876543210', 'Class 1', 'A',
                 'ACTIVE', 'GENERAL', 'REGULAR'],
                ['Jane', 'Smith', '2010-05-20', 'F',
                 'jane.smith@example.com', '9876543211', 'Class 2', 'B',
                 'ACTIVE', 'OBC', 'REGULAR'],
            ]
            
            for row in sample_data:
                writer.writerow(row)
        else:
            headers = ['first_name', 'last_name', 'personal_email']
            writer.writerow(headers)
            
            sample_data = [
                ['John', 'Doe', 'john.doe@example.com'],
                ['Jane', 'Smith', 'jane.smith@example.com'],
            ]
            
            for row in sample_data:
                writer.writerow(row)
        
        return response
    
    def generate_excel_sample(self, template_type):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if template_type == 'detailed':
            headers = ['First Name', 'Last Name', 'Date of Birth', 'Gender',
                      'Email', 'Phone', 'Class', 'Section', 'Status']
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
        else:
            headers = ['First Name', 'Last Name', 'Email']
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'student_upload_{template_type}_template.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class BulkUploadValidateView(BaseView):
    """AJAX validation endpoint"""
    permission_required = 'students.add_student'
    
    def post(self, request, *args, **kwargs):
        try:
            file = request.FILES.get('file')
            if not file:
                return JsonResponse({'valid': False, 'error': 'No file provided'})
            
            # Quick validation
            if file.size > 10 * 1024 * 1024:
                return JsonResponse({'valid': False, 'error': 'File too large'})
            
            ext = file.name.split('.')[-1].lower()
            if ext not in ['csv', 'xls', 'xlsx']:
                return JsonResponse({'valid': False, 'error': 'Invalid file type'})
            
            # Basic structure check
            if ext == 'csv':
                result = self.validate_csv_structure(file)
            else:
                result = self.validate_excel_structure(file)
            
            return JsonResponse(result)
            
        except Exception as e:
            return JsonResponse({'valid': False, 'error': str(e)})
    
    def validate_csv_structure(self, file):
        content = file.read(10240).decode('utf-8', errors='ignore')
        file.seek(0)
        lines = content.splitlines()
        
        if len(lines) < 2:
            return {'valid': False, 'error': 'File appears empty'}
        
        return {'valid': True, 'row_count': len(lines) - 1}
    
    def validate_excel_structure(self, file):
        wb = openpyxl.load_workbook(filename=file, read_only=True)
        ws = wb.active
        row_count = sum(1 for _ in ws.iter_rows(min_row=2) if any(cell.value for cell in _))
        return {'valid': True, 'row_count': row_count}


# ============================================================================
# EXPORT VIEWS
# ============================================================================

class StudentExportView(BaseView):
    """Export student data with filtering"""
    permission_required = 'students.export_student_data'
    
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'csv')
        queryset = self.get_filtered_queryset(request)
        
        if format_type == 'excel':
            return self.export_excel(queryset)
        else:
            return self.export_csv(queryset)
    
    def get_filtered_queryset(self, request):
        queryset = Student.get_secure_queryset(request.user)
        
        # Apply filters from request
        status = request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        class_id = request.GET.get('class_id')
        if class_id:
            queryset = queryset.filter(current_class_id=class_id)
        
        # Log the export
        AuditService.create_audit_entry(
            action='EXPORT',
            resource_type='Student',
            user=request.user,
            request=request,
            severity='INFO',
            extra_data={
                'export_format': request.GET.get('format', 'csv'),
                'record_count': queryset.count(),
                'filters': dict(request.GET)
            }
        )
        
        return queryset
    
    def export_csv(self, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
            'Email', 'Phone', 'Class', 'Section', 'Status', 'Category'
        ])
        
        for student in queryset:
            writer.writerow([
                student.admission_number,
                student.full_name,
                student.date_of_birth,
                student.get_gender_display(),
                student.personal_email,
                student.mobile_primary,
                student.current_class.name if student.current_class else '',
                student.section.name if student.section else '',
                student.get_status_display(),
                student.get_category_display()
            ])
        
        return response
    
    def export_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Headers
        headers = ['Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                  'Email', 'Phone', 'Class', 'Section', 'Status', 'Category']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Data
        for row_idx, student in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=student.admission_number)
            ws.cell(row=row_idx, column=2, value=student.full_name)
            ws.cell(row=row_idx, column=3, value=str(student.date_of_birth))
            ws.cell(row=row_idx, column=4, value=student.get_gender_display())
            ws.cell(row=row_idx, column=5, value=student.personal_email)
            ws.cell(row=row_idx, column=6, value=student.mobile_primary)
            ws.cell(row=row_idx, column=7, value=student.current_class.name if student.current_class else '')
            ws.cell(row=row_idx, column=8, value=student.section.name if student.section else '')
            ws.cell(row=row_idx, column=9, value=student.get_status_display())
            ws.cell(row=row_idx, column=10, value=student.get_category_display())
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
        wb.save(response)
        return response


# ============================================================================
# SPECIALIZED VIEWS
# ============================================================================

class StudentAcademicHistoryView(BaseView):
    """View student academic history"""
    template_name = 'students/student_academic_history.html'
    permission_required = 'students.view_student'
    
    def get_object(self):
        return get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs.get('pk')
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        context['student'] = student
        
        # Get exam results
        try:
            from apps.exams.models import ExamResult
            context['exam_results'] = ExamResult.objects.filter(
                student=student
            ).order_by('-exam__start_date')
        except ImportError:
            context['exam_results'] = []
        
        # Get attendance summary
        try:
            from apps.attendance.models import Attendance
            attendance = Attendance.objects.filter(student=student)
            context['attendance_summary'] = attendance.aggregate(
                total=models.Count('id'),
                present=models.Count('id', filter=models.Q(status='PRESENT')),
                absent=models.Count('id', filter=models.Q(status='ABSENT'))
            )
        except ImportError:
            context['attendance_summary'] = {}
        
        return context


class StudentPromoteView(BaseView):
    """Promote student to next class"""
    template_name = 'students/student_promote.html'
    permission_required = 'students.change_student'
    
    def get_object(self):
        return get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs.get('pk')
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        context['student'] = student
        context['next_classes'] = SchoolClass.objects.filter(
            tenant=self.request.tenant,
            order__gt=student.current_class.order if student.current_class else 0
        ).order_by('order')
        
        return context
    
    def post(self, request, *args, **kwargs):
        student = self.get_object()
        new_class_id = request.POST.get('new_class_id')
        
        try:
            new_class = SchoolClass.objects.get(
                id=new_class_id,
                tenant=request.tenant
            )
            
            old_class = student.current_class
            
            # Update student
            student.current_class = new_class
            student.save()
            
            # Log the promotion
            AuditService.log_update(
                user=request.user,
                instance=student,
                request=request,
                extra_data={
                    'action': 'promotion',
                    'old_class': str(old_class),
                    'new_class': str(new_class)
                }
            )
            
            messages.success(
                request,
                f"Student {student.full_name} promoted to {new_class.name}"
            )
            
        except Exception as e:
            messages.error(request, f"Promotion failed: {str(e)}")
        
        return redirect('students:student_detail', pk=student.id)


class StudentReportView(BaseView):
    """Generate student report"""
    template_name = 'students/student_report.html'
    permission_required = 'students.view_student'
    
    def get_object(self):
        return get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs.get('pk')
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        context['student'] = student
        
        # Generate report data
        context['report_date'] = timezone.now()
        context['academic_year'] = student.academic_year
        context['current_class'] = student.current_class
        
        # You can add more report data here
        # For example: grades, attendance, comments, etc.
        
        return context


class StudentIdCardView(BaseDetailView):
    """Generate and download student ID card"""
    permission_required = 'students.view_student'
    
    def get_object(self):
        return get_object_or_404(
            Student.get_secure_queryset(self.request.user),
            id=self.kwargs.get('pk')
        )

    def get(self, request, *args, **kwargs):
        """Generate ID card on the fly"""
        student = self.get_object()
        
        # Log the ID card generation
        AuditService.create_audit_entry(
            action='READ',
            resource_type='StudentIDCard',
            user=request.user,
            request=request,
            instance=student,
            extra_data={'action': 'generated_id_card'}
        )
        
        
        generator = StudentIDCardGenerator(student)
        return generator.get_id_card_response()


# ============================================================================
# MULTI-STEP REGISTRATION VIEWS
# ============================================================================

class StudentRegistrationStep1View(StudentCreateView):
    """Step 1: Student Information"""
    template_name = 'students/registration/step1_student.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['step'] = 1
        return context
    
    def get_success_url(self):
        # Redirect to Step 2 with the created student ID
        return reverse_lazy('students:registration_step2', kwargs={'student_id': self.object.id})


class StudentRegistrationStep2View(GuardianCreateView):
    """Step 2: Guardian Information"""
    template_name = 'students/registration/step2_guardian.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['step'] = 2
        return context
        
    def get_success_url(self):
        # Redirect to Step 3
        return reverse_lazy('students:registration_step3', kwargs={'student_id': self.kwargs['student_id']})


class StudentRegistrationStep3View(DocumentUploadView):
    """Step 3: Document Upload"""
    template_name = 'students/registration/step3_documents.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['step'] = 3
        # Pass document form explicitly if needed, but BaseCreateView handles 'form'
        return context
        
    
    def get_success_url(self):
        # Final step - redirect to student detail
        return reverse_lazy('students:student_detail', kwargs={'pk': self.kwargs['student_id']})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Create user account after successful registration
        try:
            student = self.object.student
            student.create_user_account()
            messages.success(self.request, f"User account created for {student.full_name}")
        except Exception as e:
            logger.error(f"Failed to create user account: {e}")
            messages.warning(self.request, "Student registered, but user account creation failed.")
            
        return response
