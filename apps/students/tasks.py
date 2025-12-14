"""
Background tasks for student operations using Celery
"""

import logging
import csv
import io
import tempfile
import os
from typing import Dict, List, Optional, Any
from datetime import datetime
from celery import shared_task
from django.utils import timezone
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Q
from django.contrib.auth import get_user_model

# Import models and services
from .models import Student, Guardian, StudentDocument, StudentAcademicHistory
from .services import StudentService, GuardianService, DocumentService, StudentImportService
from apps.core.services.audit_service import AuditService
from apps.core.services.notification_service import NotificationService
from apps.academics.models import AcademicYear, SchoolClass, Section

logger = logging.getLogger(__name__)
User = get_user_model()


@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def process_bulk_upload_async(self, upload_data: Dict) -> Dict:
    """
    Process bulk student upload asynchronously
    
    Args:
        upload_data: Dictionary containing upload information
        
    Returns:
        Dictionary with processing results
    """
    try:
        # Extract data
        file_content = upload_data.get('file_content')
        file_name = upload_data.get('file_name')
        file_type = upload_data.get('file_type', 'csv')
        tenant_id = upload_data.get('tenant_id')
        academic_year_id = upload_data.get('academic_year_id')
        update_existing = upload_data.get('update_existing', False)
        skip_errors = upload_data.get('skip_errors', False)
        user_id = upload_data.get('user_id')
        send_welcome_email = upload_data.get('send_welcome_email', False)
        
        # Get related objects
        from django.apps import apps
        Tenant = apps.get_model('core', 'Tenant')
        
        tenant = Tenant.objects.get(id=tenant_id)
        academic_year = AcademicYear.objects.get(id=academic_year_id) if academic_year_id else None
        user = User.objects.get(id=user_id) if user_id else None
        
        # Update task status
        self.update_state(
            state='PROGRESS',
            meta={
                'current': 0,
                'total': 100,
                'status': 'Starting import process...'
            }
        )
        
        # Process based on file type
        if file_type.lower() == 'csv':
            result = _process_csv_upload(
                file_content=file_content,
                file_name=file_name,
                tenant=tenant,
                academic_year=academic_year,
                update_existing=update_existing,
                skip_errors=skip_errors,
                user=user,
                send_welcome_email=send_welcome_email
            )
        elif file_type.lower() in ['xls', 'xlsx']:
            result = _process_excel_upload(
                file_content=file_content,
                file_name=file_name,
                tenant=tenant,
                academic_year=academic_year,
                update_existing=update_existing,
                skip_errors=skip_errors,
                user=user,
                send_welcome_email=send_welcome_email
            )
        else:
            result = {
                'success': False,
                'error': f'Unsupported file type: {file_type}'
            }
        
        # Create audit log
        AuditService.create_audit_entry(
            action='BULK_OPERATION',
            resource_type='Student',
            user=user,
            tenant=tenant,
            request=None,
            severity='INFO' if result.get('success', False) else 'ERROR',
            extra_data={
                'task_id': self.request.id,
                'operation': 'bulk_upload',
                'file_name': file_name,
                'result': result
            }
        )
        
        return result
        
    except Exception as e:
        logger.error(f"Error in bulk upload task: {str(e)}", exc_info=True)
        
        # Retry the task
        try:
            raise self.retry(exc=e, countdown=60)
        except self.MaxRetriesExceededError:
            return {
                'success': False,
                'error': f'Max retries exceeded: {str(e)}',
                'task_id': self.request.id
            }


def _process_csv_upload(file_content, file_name, tenant, academic_year, update_existing, skip_errors, user, send_welcome_email):
    """Process CSV file upload"""
    results = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
        'warnings': []
    }
    
    try:
        # Decode file content
        if isinstance(file_content, bytes):
            file_content = file_content.decode('utf-8')
        
        # Create CSV reader
        csv_file = io.StringIO(file_content)
        
        # Detect delimiter
        sample = csv_file.read(1024)
        csv_file.seek(0)
        
        # Try common delimiters
        delimiters = [',', ';', '\t', '|']
        delimiter_counts = {d: sample.count(d) for d in delimiters}
        delimiter = max(delimiter_counts, key=delimiter_counts.get) if any(delimiter_counts.values()) else ','
        
        # Read CSV
        reader = csv.reader(csv_file, delimiter=delimiter)
        rows = list(reader)
        
        if len(rows) < 2:
            results['errors'].append('CSV file has insufficient data')
            results['success'] = False
            return results
        
        # Process headers
        headers = [h.strip().lower().replace(' ', '_') for h in rows[0]]
        
        # Map headers to field names
        header_mapping = {
            'firstname': 'first_name',
            'lastname': 'last_name',
            'dob': 'date_of_birth',
            'email': 'personal_email',
            'phone': 'mobile_primary',
            'class': 'class_name',
            'section': 'section_name',
            'gender': 'gender',
            'status': 'status',
            'category': 'category'
        }
        
        headers = [header_mapping.get(h, h) for h in headers]
        
        # Process data rows
        for i, row in enumerate(rows[1:], start=2):
            try:
                if len(row) < len(headers):
                    row.extend([''] * (len(headers) - len(row)))
                
                # Create row dictionary
                row_dict = {headers[j]: row[j].strip() for j in range(len(headers))}
                
                # Skip empty rows
                if all(value == '' for value in row_dict.values()):
                    results['skipped'] += 1
                    continue
                
                # Process student
                result = _process_student_row(
                    row_dict=row_dict,
                    row_number=i,
                    tenant=tenant,
                    academic_year=academic_year,
                    update_existing=update_existing,
                    user=user,
                    send_welcome_email=send_welcome_email
                )
                
                if result['success']:
                    if result['action'] == 'created':
                        results['created'] += 1
                    elif result['action'] == 'updated':
                        results['updated'] += 1
                    else:
                        results['skipped'] += 1
                else:
                    if skip_errors:
                        results['skipped'] += 1
                        results['warnings'].append(f"Row {i}: {result['error']}")
                    else:
                        results['errors'].append(f"Row {i}: {result['error']}")
                        if not skip_errors:
                            break
                        
            except Exception as e:
                error_msg = f"Row {i}: {str(e)}"
                if skip_errors:
                    results['skipped'] += 1
                    results['warnings'].append(error_msg)
                else:
                    results['errors'].append(error_msg)
                    if not skip_errors:
                        break
        
        results['success'] = len(results['errors']) == 0 or skip_errors
        results['total_rows'] = len(rows) - 1
        results['processed'] = results['created'] + results['updated'] + results['skipped']
        
        return results
        
    except Exception as e:
        logger.error(f"Error processing CSV: {str(e)}", exc_info=True)
        results['success'] = False
        results['errors'].append(str(e))
        return results


def _process_excel_upload(file_content, file_name, tenant, academic_year, update_existing, skip_errors, user, send_welcome_email):
    """Process Excel file upload"""
    results = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
        'warnings': []
    }
    
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(file_content)
            tmp_file_path = tmp_file.name
        
        try:
            # Load workbook
            wb = load_workbook(filename=tmp_file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    header = str(cell.value).strip().lower().replace(' ', '_')
                    headers.append(header)
                else:
                    headers.append(f'column_{cell.column}')
            
            # Map headers to field names
            header_mapping = {
                'firstname': 'first_name',
                'lastname': 'last_name',
                'dob': 'date_of_birth',
                'email': 'personal_email',
                'phone': 'mobile_primary',
                'class': 'class_name',
                'section': 'section_name',
                'gender': 'gender',
                'status': 'status',
                'category': 'category'
            }
            
            headers = [header_mapping.get(h, h) for h in headers]
            
            # Process data rows
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
                        results['skipped'] += 1
                        continue
                    
                    # Create row dictionary
                    row_dict = {}
                    for j, cell_value in enumerate(row):
                        if j < len(headers):
                            value = _convert_excel_value(cell_value)
                            row_dict[headers[j]] = value
                    
                    # Process student
                    result = _process_student_row(
                        row_dict=row_dict,
                        row_number=i,
                        tenant=tenant,
                        academic_year=academic_year,
                        update_existing=update_existing,
                        user=user,
                        send_welcome_email=send_welcome_email
                    )
                    
                    if result['success']:
                        if result['action'] == 'created':
                            results['created'] += 1
                        elif result['action'] == 'updated':
                            results['updated'] += 1
                        else:
                            results['skipped'] += 1
                    else:
                        if skip_errors:
                            results['skipped'] += 1
                            results['warnings'].append(f"Row {i}: {result['error']}")
                        else:
                            results['errors'].append(f"Row {i}: {result['error']}")
                            if not skip_errors:
                                break
                            
                except Exception as e:
                    error_msg = f"Row {i}: {str(e)}"
                    if skip_errors:
                        results['skipped'] += 1
                        results['warnings'].append(error_msg)
                    else:
                        results['errors'].append(error_msg)
                        if not skip_errors:
                            break
            
            results['success'] = len(results['errors']) == 0 or skip_errors
            results['total_rows'] = ws.max_row - 1
            results['processed'] = results['created'] + results['updated'] + results['skipped']
            
        finally:
            # Clean up temporary file
            os.unlink(tmp_file_path)
        
        return results
        
    except Exception as e:
        logger.error(f"Error processing Excel: {str(e)}", exc_info=True)
        results['success'] = False
        results['errors'].append(str(e))
        return results


def _convert_excel_value(value):
    """Convert Excel cell value to string"""
    if value is None:
        return ''
    elif isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    else:
        return str(value).strip()


def _process_student_row(row_dict, row_number, tenant, academic_year, update_existing, user, send_welcome_email):
    """Process a single student row"""
    result = {
        'success': False,
        'action': 'skipped',
        'error': None,
        'student_id': None
    }
    
    try:
        with transaction.atomic():
            # Extract basic information
            first_name = row_dict.get('first_name', '').strip()
            last_name = row_dict.get('last_name', '').strip()
            personal_email = row_dict.get('personal_email', '').strip().lower()
            
            # Validate required fields
            if not first_name and not last_name:
                result['error'] = 'Missing both first and last name'
                return result
            
            # Get or create email
            if not personal_email:
                # Generate email
                import re
                first_part = re.sub(r'[^a-z0-9]', '', first_name.lower())[:10]
                last_part = re.sub(r'[^a-z0-9]', '', last_name.lower())[:10]
                domain = getattr(tenant, 'domain', f"{tenant.schema_name}.edu")
                personal_email = f"{first_part}.{last_part}{row_number}@{domain}"
            
            # Check if student exists
            existing_student = Student.objects.filter(
                tenant=tenant,
                personal_email=personal_email
            ).first()
            
            if existing_student and not update_existing:
                result['error'] = f'Student with email {personal_email} already exists'
                return result
            
            # Prepare student data
            student_data = {
                'tenant': tenant,
                'first_name': first_name,
                'last_name': last_name,
                'personal_email': personal_email,
                'gender': row_dict.get('gender', 'U').upper()[:1],
                'date_of_birth': _parse_date(row_dict.get('date_of_birth')),
                'mobile_primary': row_dict.get('mobile_primary', '')[:15],
                'status': row_dict.get('status', 'ACTIVE').upper(),
                'category': row_dict.get('category', 'GENERAL').upper(),
            }
            
            # Set academic year
            if academic_year:
                student_data['academic_year'] = academic_year
            
            # Handle class
            class_name = row_dict.get('class_name', '').strip()
            if class_name:
                school_class = SchoolClass.objects.filter(
                    tenant=tenant,
                    name__iexact=class_name
                ).first()
                
                if school_class:
                    student_data['current_class'] = school_class
                    
                    # Handle section
                    section_name = row_dict.get('section_name', '').strip()
                    if section_name:
                        section = Section.objects.filter(
                            class_name=school_class,
                            name__iexact=section_name
                        ).first()
                        if section:
                            student_data['section'] = section
            
            # Create or update student
            if existing_student and update_existing:
                # Update existing student
                for key, value in student_data.items():
                    if key != 'tenant':
                        setattr(existing_student, key, value)
                existing_student.save()
                student = existing_student
                result['action'] = 'updated'
            else:
                # Create new student
                student = Student.objects.create(**student_data)
                result['action'] = 'created'
            
            # Generate admission number if not present
            if not student.admission_number:
                student.admission_number = StudentService.generate_admission_number(tenant, academic_year)
                student.save()
            
            # Handle guardian information
            guardian_data = _extract_guardian_data(row_dict)
            if guardian_data:
                _process_guardian_info(student, guardian_data, tenant)
            
            # Send welcome email if requested
            if send_welcome_email and result['action'] == 'created':
                try:
                    send_student_welcome_email.delay(student.id)
                except Exception as e:
                    logger.warning(f"Failed to queue welcome email: {str(e)}")
            
            # Create audit log for individual student
            AuditService.create_audit_entry(
                action='CREATE' if result['action'] == 'created' else 'UPDATE',
                resource_type='Student',
                user=user,
                tenant=tenant,
                request=None,
                severity='INFO',
                instance=student,
                extra_data={
                    'action': result['action'],
                    'source': 'bulk_upload',
                    'row_number': row_number
                }
            )
            
            result['success'] = True
            result['student_id'] = student.id
            return result
            
    except Exception as e:
        logger.error(f"Error processing student row {row_number}: {str(e)}", exc_info=True)
        result['error'] = str(e)
        return result


def _parse_date(date_str):
    """Parse date string to date object"""
    if not date_str:
        return None
    
    date_formats = [
        '%Y-%m-%d',
        '%d-%m-%Y',
        '%d/%m/%Y',
        '%m/%d/%Y',
        '%Y/%m/%d'
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(str(date_str).strip(), fmt).date()
        except ValueError:
            continue
    
    return None


def _extract_guardian_data(row_dict):
    """Extract guardian information from row data"""
    guardian_data = {}
    
    # Father information
    if row_dict.get('father_name'):
        guardian_data['father'] = {
            'first_name': row_dict.get('father_name', '').split()[0] if row_dict.get('father_name') else '',
            'last_name': ' '.join(row_dict.get('father_name', '').split()[1:]) if row_dict.get('father_name') else '',
            'email': row_dict.get('father_email', ''),
            'phone': row_dict.get('father_phone', ''),
            'occupation': row_dict.get('father_occupation', ''),
            'relationship': 'FATHER'
        }
    
    # Mother information
    if row_dict.get('mother_name'):
        guardian_data['mother'] = {
            'first_name': row_dict.get('mother_name', '').split()[0] if row_dict.get('mother_name') else '',
            'last_name': ' '.join(row_dict.get('mother_name', '').split()[1:]) if row_dict.get('mother_name') else '',
            'email': row_dict.get('mother_email', ''),
            'phone': row_dict.get('mother_phone', ''),
            'occupation': row_dict.get('mother_occupation', ''),
            'relationship': 'MOTHER'
        }
    
    # Guardian information
    if row_dict.get('guardian_name'):
        guardian_data['guardian'] = {
            'first_name': row_dict.get('guardian_name', '').split()[0] if row_dict.get('guardian_name') else '',
            'last_name': ' '.join(row_dict.get('guardian_name', '').split()[1:]) if row_dict.get('guardian_name') else '',
            'email': row_dict.get('guardian_email', ''),
            'phone': row_dict.get('guardian_phone', ''),
            'occupation': row_dict.get('guardian_occupation', ''),
            'relationship': row_dict.get('guardian_relationship', 'GUARDIAN')
        }
    
    return guardian_data


def _process_guardian_info(student, guardian_data, tenant):
    """Process guardian information"""
    for role, data in guardian_data.items():
        try:
            if data.get('first_name') or data.get('last_name'):
                # Check if guardian already exists
                existing_guardian = Guardian.objects.filter(
                    tenant=tenant,
                    email=data.get('email')
                ).first() if data.get('email') else None
                
                if not existing_guardian:
                    # Create new guardian
                    guardian = Guardian.objects.create(
                        tenant=tenant,
                        first_name=data.get('first_name', ''),
                        last_name=data.get('last_name', ''),
                        email=data.get('email', ''),
                        phone=data.get('phone', ''),
                        occupation=data.get('occupation', ''),
                        relationship=data.get('relationship', 'OTHER'),
                        is_primary=(role == 'father')  # Father as primary by default
                    )
                else:
                    guardian = existing_guardian
                
                # Add guardian to student
                if guardian not in student.guardians.all():
                    student.guardians.add(guardian)
                    
        except Exception as e:
            logger.warning(f"Error processing guardian {role} for student {student.id}: {str(e)}")


@shared_task
def generate_id_cards_batch(student_ids: List[int], user_id: int, tenant_id: int) -> Dict:
    """
    Generate ID cards for multiple students in batch
    
    Args:
        student_ids: List of student IDs
        user_id: User ID who requested the batch
        tenant_id: Tenant ID
        
    Returns:
        Dictionary with results
    """
    try:
        from django.apps import apps
        
        Tenant = apps.get_model('core', 'Tenant')
        tenant = Tenant.objects.get(id=tenant_id)
        user = User.objects.get(id=user_id) if user_id else None
        
        # Get students
        students = Student.objects.filter(
            id__in=student_ids,
            tenant=tenant,
            is_active=True
        ).select_related('current_class', 'section', 'academic_year')
        
        results = {
            'total': len(students),
            'generated': 0,
            'failed': 0,
            'errors': []
        }
        
        # This would normally generate PDFs or images
        # For now, simulate the process
        for student in students:
            try:
                # Simulate ID card generation
                # In production, this would create actual ID card files
                logger.info(f"Generating ID card for student: {student.full_name}")
                
                # Create audit log for each ID card
                AuditService.create_audit_entry(
                    action='CREATE',
                    resource_type='StudentIDCard',
                    user=user,
                    tenant=tenant,
                    request=None,
                    severity='INFO',
                    instance=student,
                    extra_data={
                        'action': 'batch_id_card_generation',
                        'student_name': student.full_name,
                        'admission_number': student.admission_number
                    }
                )
                
                results['generated'] += 1
                
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Student {student.admission_number}: {str(e)}")
                logger.error(f"Error generating ID card for student {student.id}: {str(e)}")
        
        # Create overall audit log
        AuditService.create_audit_entry(
            action='BULK_OPERATION',
            resource_type='StudentIDCard',
            user=user,
            tenant=tenant,
            request=None,
            severity='INFO',
            extra_data={
                'operation': 'batch_id_card_generation',
                'results': results,
                'student_count': len(students)
            }
        )
        
        return {
            'success': results['failed'] == 0,
            'results': results,
            'message': f"Generated {results['generated']} ID cards, failed: {results['failed']}"
        }
        
    except Exception as e:
        logger.error(f"Error in batch ID card generation: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': str(e)
        }


@shared_task
def send_student_welcome_email(student_id: int) -> Dict:
    """
    Send welcome email to student
    
    Args:
        student_id: Student ID
        
    Returns:
        Dictionary with results
    """
    try:
        student = Student.objects.get(id=student_id, is_active=True)
        
        # Check if student has email
        if not student.personal_email:
            return {
                'success': False,
                'error': 'Student has no email address'
            }
        
        # Prepare email content
        subject = f"Welcome to {student.tenant.name if hasattr(student.tenant, 'name') else 'Our School'}!"
        
        message = f"""
Dear {student.first_name},

Welcome to {student.tenant.name if hasattr(student.tenant, 'name') else 'Our School'}!

Your registration has been successfully completed.

Registration Details:
- Admission Number: {student.admission_number}
- Full Name: {student.full_name}
- Class: {student.current_class.name if student.current_class else 'Not assigned'}
- Section: {student.section.name if student.section else 'Not assigned'}

You can now access the student portal using your credentials.

Best regards,
{student.tenant.name if hasattr(student.tenant, 'name') else 'School Administration'}
"""
        
        # Send email
        from django.core.mail import send_mail
        
        sent_count = send_mail(
            subject=subject,
            message=message,
            from_email=None,  # Use DEFAULT_FROM_EMAIL
            recipient_list=[student.personal_email],
            fail_silently=False
        )
        
        # Create audit log
        AuditService.create_audit_entry(
            action='CREATE',
            resource_type='StudentWelcomeEmail',
            user=None,  # System action
            tenant=student.tenant,
            request=None,
            severity='INFO',
            instance=student,
            extra_data={
                'recipient': student.personal_email,
                'sent_count': sent_count,
                'admission_number': student.admission_number
            }
        )
        
        return {
            'success': sent_count > 0,
            'sent_count': sent_count,
            'recipient': student.personal_email
        }
        
    except Student.DoesNotExist:
        return {
            'success': False,
            'error': 'Student not found'
        }
    except Exception as e:
        logger.error(f"Error sending welcome email: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': str(e)
        }


@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def export_student_data_async(self, export_params: Dict) -> Dict:
    """
    Export student data asynchronously
    
    Args:
        export_params: Dictionary containing export parameters
        
    Returns:
        Dictionary with export results
    """
    try:
        # Extract parameters
        export_format = export_params.get('format', 'csv')
        export_type = export_params.get('type', 'basic')
        filters = export_params.get('filters', {})
        tenant_id = export_params.get('tenant_id')
        user_id = export_params.get('user_id')
        
        # Get related objects
        from django.apps import apps
        Tenant = apps.get_model('core', 'Tenant')
        
        tenant = Tenant.objects.get(id=tenant_id)
        user = User.objects.get(id=user_id) if user_id else None
        
        # Update task status
        self.update_state(
            state='PROGRESS',
            meta={
                'current': 0,
                'total': 100,
                'status': 'Preparing export...'
            }
        )
        
        # Get filtered students
        students = StudentService.get_secure_queryset(user, tenant)
        
        # Apply filters
        if filters:
            students = _apply_export_filters(students, filters)
        
        # Update task status
        self.update_state(
            state='PROGRESS',
            meta={
                'current': 50,
                'total': 100,
                'status': f'Processing {students.count()} students...'
            }
        )
        
        # Generate export file
        if export_format == 'csv':
            result = _generate_csv_export(students, export_type)
        elif export_format == 'excel':
            result = _generate_excel_export(students, export_type)
        else:
            result = {
                'success': False,
                'error': f'Unsupported export format: {export_format}'
            }
        
        # Create audit log
        AuditService.create_audit_entry(
            action='EXPORT',
            resource_type='Student',
            user=user,
            tenant=tenant,
            request=None,
            severity='INFO',
            extra_data={
                'task_id': self.request.id,
                'operation': 'data_export',
                'format': export_format,
                'type': export_type,
                'student_count': students.count(),
                'filters': filters
            }
        )
        
        return result
        
    except Exception as e:
        logger.error(f"Error in export task: {str(e)}", exc_info=True)
        
        # Retry the task
        try:
            raise self.retry(exc=e, countdown=60)
        except self.MaxRetriesExceededError:
            return {
                'success': False,
                'error': f'Max retries exceeded: {str(e)}',
                'task_id': self.request.id
            }


def _apply_export_filters(queryset, filters):
    """Apply filters to student queryset"""
    q_objects = Q()
    
    # Status filter
    if filters.get('status'):
        q_objects &= Q(status=filters['status'])
    
    # Class filter
    if filters.get('class_id'):
        q_objects &= Q(current_class_id=filters['class_id'])
    
    # Section filter
    if filters.get('section_id'):
        q_objects &= Q(section_id=filters['section_id'])
    
    # Academic year filter
    if filters.get('academic_year_id'):
        q_objects &= Q(academic_year_id=filters['academic_year_id'])
    
    # Category filter
    if filters.get('category'):
        q_objects &= Q(category=filters['category'])
    
    # Gender filter
    if filters.get('gender'):
        q_objects &= Q(gender=filters['gender'])
    
    # Date range filters
    if filters.get('created_from'):
        try:
            date_from = datetime.strptime(filters['created_from'], '%Y-%m-%d').date()
            q_objects &= Q(created_at__date__gte=date_from)
        except ValueError:
            pass
    
    if filters.get('created_to'):
        try:
            date_to = datetime.strptime(filters['created_to'], '%Y-%m-%d').date()
            q_objects &= Q(created_at__date__lte=date_to)
        except ValueError:
            pass
    
    return queryset.filter(q_objects).distinct()


def _generate_csv_export(students, export_type):
    """Generate CSV export"""
    try:
        import csv
        from django.http import HttpResponse
        
        # Create response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="students_export_{timestamp}.csv"'
        
        writer = csv.writer(response)
        
        # Write headers based on export type
        if export_type == 'detailed':
            headers = [
                'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                'Email', 'Phone', 'Class', 'Section', 'Stream', 'Academic Year',
                'Status', 'Category', 'Blood Group', 'Nationality', 'Religion',
                'Caste', 'Admission Date', 'Created At', 'Updated At'
            ]
        else:
            headers = [
                'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                'Email', 'Phone', 'Class', 'Section', 'Status', 'Category'
            ]
        
        writer.writerow(headers)
        
        # Write data
        for student in students.select_related('current_class', 'section', 'stream', 'academic_year'):
            if export_type == 'detailed':
                row = [
                    student.admission_number,
                    student.full_name,
                    student.date_of_birth,
                    student.get_gender_display(),
                    student.personal_email,
                    student.mobile_primary,
                    student.current_class.name if student.current_class else '',
                    student.section.name if student.section else '',
                    student.stream.name if student.stream else '',
                    student.academic_year.name if student.academic_year else '',
                    student.get_status_display(),
                    student.get_category_display(),
                    student.blood_group or '',
                    student.nationality or '',
                    student.religion or '',
                    student.caste or '',
                    student.admission_date or '',
                    student.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    student.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                ]
            else:
                row = [
                    student.admission_number,
                    student.full_name,
                    student.date_of_birth,
                    student.get_gender_display(),
                    student.personal_email,
                    student.mobile_primary,
                    student.current_class.name if student.current_class else '',
                    student.section.name if student.section else '',
                    student.get_status_display(),
                    student.get_category_display()
                ]
            
            writer.writerow(row)
        
        return {
            'success': True,
            'content': response.content,
            'content_type': 'text/csv',
            'filename': f"students_export_{timestamp}.csv"
        }
        
    except Exception as e:
        logger.error(f"Error generating CSV export: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': str(e)
        }


def _generate_excel_export(students, export_type):
    """Generate Excel export"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Write headers based on export type
        if export_type == 'detailed':
            headers = [
                'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                'Email', 'Phone', 'Class', 'Section', 'Stream', 'Academic Year',
                'Status', 'Category', 'Blood Group', 'Nationality', 'Religion',
                'Caste', 'Admission Date', 'Created At', 'Updated At'
            ]
        else:
            headers = [
                'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                'Email', 'Phone', 'Class', 'Section', 'Status', 'Category'
            ]
        
        # Write headers with formatting
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="CCCCCC",
                end_color="CCCCCC",
                fill_type="solid"
            )
        
        # Write data
        for row_idx, student in enumerate(
            students.select_related('current_class', 'section', 'stream', 'academic_year'),
            start=2
        ):
            if export_type == 'detailed':
                ws.cell(row=row_idx, column=1, value=student.admission_number)
                ws.cell(row=row_idx, column=2, value=student.full_name)
                ws.cell(row=row_idx, column=3, value=str(student.date_of_birth))
                ws.cell(row=row_idx, column=4, value=student.get_gender_display())
                ws.cell(row=row_idx, column=5, value=student.personal_email)
                ws.cell(row=row_idx, column=6, value=student.mobile_primary)
                ws.cell(row=row_idx, column=7, value=student.current_class.name if student.current_class else '')
                ws.cell(row=row_idx, column=8, value=student.section.name if student.section else '')
                ws.cell(row=row_idx, column=9, value=student.stream.name if student.stream else '')
                ws.cell(row=row_idx, column=10, value=student.academic_year.name if student.academic_year else '')
                ws.cell(row=row_idx, column=11, value=student.get_status_display())
                ws.cell(row=row_idx, column=12, value=student.get_category_display())
                ws.cell(row=row_idx, column=13, value=student.blood_group or '')
                ws.cell(row=row_idx, column=14, value=student.nationality or '')
                ws.cell(row=row_idx, column=15, value=student.religion or '')
                ws.cell(row=row_idx, column=16, value=student.caste or '')
                ws.cell(row=row_idx, column=17, value=str(student.admission_date) if student.admission_date else '')
                ws.cell(row=row_idx, column=18, value=student.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_idx, column=19, value=student.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_idx, column=1, value=student.admission_number)
                ws.cell(row=row_idx, column=2, value=student.full_name)
                ws.cell(row=row_idx, column=3, value=str(student.date_of_birth))
                ws.cell(row=row_idx, column=4, value=student.get_gender_display())
                ws.cell(row=row_idx, column=5, value=student.personal_email)
                ws.cell(row=row_idx, column=6, value=student.mobile_primary)
                ws.cell(row=row_idx, column=7, value=student.current_class.name if student.current_class else '')
                ws.cell(row=row_idx, column=8, value=student.section.name if student.section else '')
                ws.cell(row=row_idx, column=9, value=student.get_status_display())
                ws.cell(row=row_idx, column=10, value=student.get_category_display())
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        return {
            'success': True,
            'content': output.getvalue(),
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'filename': f"students_export_{timestamp}.xlsx"
        }
        
    except Exception as e:
        logger.error(f"Error generating Excel export: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': str(e)
        }


@shared_task
def sync_student_user_accounts(tenant_id: int) -> Dict:
    """
    Sync student user accounts - create missing user accounts for students
    
    Args:
        tenant_id: Tenant ID
        
    Returns:
        Dictionary with sync results
    """
    try:
        from django.apps import apps
        
        Tenant = apps.get_model('core', 'Tenant')
        tenant = Tenant.objects.get(id=tenant_id)
        
        # Get students without user accounts
        students = Student.objects.filter(
            tenant=tenant,
            is_active=True,
            user__isnull=True
        )
        
        results = {
            'total': students.count(),
            'created': 0,
            'failed': 0,
            'errors': []
        }
        
        for student in students:
            try:
                # Check if user already exists with this email
                existing_user = User.objects.filter(
                    email=student.personal_email,
                    tenant=tenant
                ).first()
                
                if existing_user:
                    # Link existing user to student
                    student.user = existing_user
                    student.save()
                    results['created'] += 1
                else:
                    # Create new user account
                    username = f"student_{student.admission_number.lower().replace('/', '_')}"
                    password = User.objects.make_random_password()
                    
                    user = User.objects.create(
                        username=username,
                        email=student.personal_email,
                        first_name=student.first_name,
                        last_name=student.last_name,
                        tenant=tenant,
                        is_active=True
                    )
                    user.set_password(password)
                    user.save()
                    
                    # Add to student group
                    from django.contrib.auth.models import Group
                    student_group, _ = Group.objects.get_or_create(name='Student')
                    user.groups.add(student_group)
                    
                    # Link user to student
                    student.user = user
                    student.save()
                    
                    # Send account creation email
                    try:
                        _send_account_creation_email(student, user, password)
                    except Exception as e:
                        logger.warning(f"Failed to send account creation email: {str(e)}")
                    
                    results['created'] += 1
                    
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Student {student.admission_number}: {str(e)}")
                logger.error(f"Error creating user account for student {student.id}: {str(e)}")
        
        # Create audit log
        AuditService.create_audit_entry(
            action='SYNC',
            resource_type='StudentUserAccount',
            user=None,  # System action
            tenant=tenant,
            request=None,
            severity='INFO',
            extra_data={
                'operation': 'sync_user_accounts',
                'results': results
            }
        )
        
        return {
            'success': results['failed'] == 0,
            'results': results,
            'message': f"Created {results['created']} user accounts, failed: {results['failed']}"
        }
        
    except Exception as e:
        logger.error(f"Error syncing student user accounts: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': str(e)
        }


def _send_account_creation_email(student, user, password):
    """Send account creation email to student"""
    subject = f"Your Student Account - {student.tenant.name if hasattr(student.tenant, 'name') else 'Our School'}"
    
    message = f"""
Dear {student.first_name},

Your student account has been created.

Account Details:
- Username: {user.username}
- Email: {user.email}
- Password: {password}
- Admission Number: {student.admission_number}

Please log in and change your password immediately for security.

Login URL: {getattr(student.tenant, 'login_url', '/login/')}

Best regards,
{student.tenant.name if hasattr(student.tenant, 'name') else 'School Administration'}
"""
    
    # Send email
    from django.core.mail import send_mail
    send_mail(
        subject=subject,
        message=message,
        from_email=None,  # Use DEFAULT_FROM_EMAIL
        recipient_list=[student.personal_email],
        fail_silently=False
    )


@shared_task
def cleanup_student_data(tenant_id: int, days_old: int = 365) -> Dict:
    """
    Clean up old student data (soft delete inactive students)
    
    Args:
        tenant_id: Tenant ID
        days_old: Clean up students older than this many days
        
    Returns:
        Dictionary with cleanup results
    """
    try:
        from django.apps import apps
        
        Tenant = apps.get_model('core', 'Tenant')
        tenant = Tenant.objects.get(id=tenant_id)
        
        cutoff_date = timezone.now() - timezone.timedelta(days=days_old)
        
        # Find inactive students older than cutoff date
        students = Student.objects.filter(
            tenant=tenant,
            is_active=True,
            status='INACTIVE',
            updated_at__lt=cutoff_date
        )
        
        results = {
            'total': students.count(),
            'cleaned': 0,
            'failed': 0,
            'errors': []
        }
        
        for student in students:
            try:
                # Soft delete student
                student.soft_delete(
                    deleted_by=None,  # System action
                    reason=f'Auto-cleanup: Inactive for more than {days_old} days',
                    category='AUTO_CLEANUP'
                )
                results['cleaned'] += 1
                
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Student {student.admission_number}: {str(e)}")
                logger.error(f"Error cleaning up student {student.id}: {str(e)}")
        
        # Create audit log
        AuditService.create_audit_entry(
            action='CLEANUP',
            resource_type='Student',
            user=None,  # System action
            tenant=tenant,
            request=None,
            severity='INFO',
            extra_data={
                'operation': 'data_cleanup',
                'cutoff_date': cutoff_date.isoformat(),
                'days_old': days_old,
                'results': results
            }
        )
        
        return {
            'success': True,
            'results': results,
            'message': f"Cleaned up {results['cleaned']} inactive students, failed: {results['failed']}"
        }
        
    except Exception as e:
        logger.error(f"Error in student data cleanup: {str(e)}", exc_info=True)
        return {
            'success': False,
            'error': str(e)
        }