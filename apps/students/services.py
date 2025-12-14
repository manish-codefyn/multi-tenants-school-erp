

import logging
import re
import uuid
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Union
from django.db import models
from django.db.models import Q, Count, Sum, Avg, Max, Min
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.core.exceptions import ValidationError, PermissionDenied
from django.contrib.auth import get_user_model
from django.core.cache import cache

# Import models
from .models import Student, Guardian, StudentAddress, StudentDocument, StudentAcademicHistory
from apps.academics.models import AcademicYear, SchoolClass, Section, Stream
from apps.core.services.audit_service import AuditService
from apps.core.services.notification_service import NotificationService

logger = logging.getLogger(__name__)
User = get_user_model()


class StudentService:
    """
    Service for student-related operations
    """
    
    @staticmethod
    def get_secure_queryset(user: User, tenant=None):
        """
        Get student queryset with tenant and permission filtering
        
        Args:
            user: Current user
            tenant: Tenant object
            
        Returns:
            Filtered Student queryset
        """
        try:
            # Start with all students for tenant
            queryset = Student.objects.filter(tenant=tenant, is_active=True)
            
            # Superusers and admins can see all students
            if user.is_superuser or user.groups.filter(name__in=['Admin', 'Administrator']).exists():
                return queryset
            
            # Teachers can see students in their classes
            if user.groups.filter(name='Teacher').exists():
                # Get classes taught by teacher
                from apps.academics.models import TeacherAssignment
                teacher_classes = TeacherAssignment.objects.filter(
                    teacher=user,
                    academic_year__is_current=True
                ).values_list('class_name', flat=True)
                
                return queryset.filter(current_class_id__in=teacher_classes)
            
            # Guardians can see only their own students
            if user.groups.filter(name='Guardian').exists():
                guardian = Guardian.objects.filter(user=user).first()
                if guardian:
                    return queryset.filter(guardians=guardian)
                return Student.objects.none()
            
            # Students can see only themselves
            if user.groups.filter(name='Student').exists():
                return queryset.filter(user=user)
            
            # Default: no access
            return Student.objects.none()
            
        except Exception as e:
            logger.error(f"Error getting secure queryset: {str(e)}", exc_info=True)
            return Student.objects.none()
    
    @staticmethod
    def has_student_access(user: User, student: Student) -> bool:
        """
        Check if user has access to specific student
        
        Args:
            user: Current user
            student: Student object
            
        Returns:
            Boolean indicating access permission
        """
        try:
            # Superusers and admins have access
            if user.is_superuser or user.groups.filter(name__in=['Admin', 'Administrator']).exists():
                return True
            
            # Check if user is the student
            if hasattr(student, 'user') and student.user == user:
                return True
            
            # Check if user is a guardian of the student
            if Guardian.objects.filter(student=student, user=user).exists():
                return True
            
            # Check if user is a teacher of the student's class
            if user.groups.filter(name='Teacher').exists():
                from apps.academics.models import TeacherAssignment
                if student.current_class:
                    return TeacherAssignment.objects.filter(
                        teacher=user,
                        class_name=student.current_class,
                        academic_year__is_current=True
                    ).exists()
            
            return False
            
        except Exception as e:
            logger.error(f"Error checking student access: {str(e)}", exc_info=True)
            return False
    
    @staticmethod
    def generate_admission_number(tenant, academic_year=None) -> str:
        """
        Generate unique admission number
        
        Args:
            tenant: Tenant object
            academic_year: Academic year object
            
        Returns:
            Unique admission number string
        """
        try:
            # Get current academic year if not provided
            if not academic_year:
                academic_year = AcademicYear.objects.filter(
                    tenant=tenant,
                    is_current=True
                ).first()
            
            # Generate prefix
            prefix = "ADM"
            if academic_year:
                # Use last 2 digits of academic year
                year_code = str(academic_year.start_date.year)[-2:]
                prefix = f"ADM{year_code}"
            
            # Get tenant code (first 3 letters)
            tenant_code = tenant.schema_name[:3].upper() if hasattr(tenant, 'schema_name') else "SCH"
            
            # Get last admission number for this tenant and academic year
            last_student = Student.objects.filter(
                tenant=tenant,
                admission_number__isnull=False
            ).exclude(admission_number='').order_by('-admission_number').first()
            
            if last_student and last_student.admission_number:
                # Extract number from last admission number
                match = re.search(r'\d+', last_student.admission_number)
                if match:
                    last_number = int(match.group())
                    new_number = last_number + 1
                else:
                    new_number = 1
            else:
                new_number = 1
            
            # Format number with leading zeros
            formatted_number = f"{new_number:05d}"
            
            # Construct admission number
            admission_number = f"{tenant_code}/{prefix}/{formatted_number}"
            
            # Ensure uniqueness
            counter = 1
            original_number = admission_number
            while Student.objects.filter(
                tenant=tenant,
                admission_number=admission_number
            ).exists():
                admission_number = f"{original_number}-{counter}"
                counter += 1
            
            return admission_number
            
        except Exception as e:
            logger.error(f"Error generating admission number: {str(e)}", exc_info=True)
            # Fallback to UUID-based number
            return f"ADM-{uuid.uuid4().hex[:8].upper()}"
    
    @staticmethod
    def create_student(data: Dict, user: User, tenant=None) -> Tuple[Optional[Student], List[str]]:
        """
        Create new student with validation
        
        Args:
            data: Student data dictionary
            user: User creating the student
            tenant: Tenant object
            
        Returns:
            Tuple of (student object, list of errors)
        """
        errors = []
        
        try:
            from .forms import StudentForm
            
            # Prepare form data
            form_data = data.copy()
            form_data['tenant'] = tenant
            
            # Create form instance
            form = StudentForm(data=form_data, tenant=tenant, user=user)
            
            if form.is_valid():
                # Save student
                student = form.save(commit=False)
                student.tenant = tenant
                
                # Generate admission number if not provided
                if not student.admission_number:
                    student.admission_number = StudentService.generate_admission_number(
                        tenant, student.academic_year
                    )
                
                student.save()
                form.save_m2m()
                
                # Create audit log
                AuditService.log_creation(
                    user=user,
                    instance=student,
                    request=None,
                    extra_data={'created_via': 'service_api'}
                )
                
                return student, errors
            else:
                # Collect form errors
                for field, field_errors in form.errors.items():
                    for error in field_errors:
                        errors.append(f"{field}: {error}")
                
                return None, errors
                
        except Exception as e:
            logger.error(f"Error creating student: {str(e)}", exc_info=True)
            errors.append(str(e))
            return None, errors
    
    @staticmethod
    def update_student(student_id: int, data: Dict, user: User, tenant=None) -> Tuple[Optional[Student], List[str]]:
        """
        Update existing student
        
        Args:
            student_id: Student ID
            data: Update data dictionary
            user: User updating the student
            tenant: Tenant object
            
        Returns:
            Tuple of (updated student object, list of errors)
        """
        errors = []
        
        try:
            # Get student
            student = Student.objects.get(id=student_id, tenant=tenant, is_active=True)
            
            # Check permission
            if not StudentService.has_student_access(user, student):
                raise PermissionDenied(_("You don't have permission to update this student."))
            
            # Store old data for audit
            old_data = {
                'first_name': student.first_name,
                'last_name': student.last_name,
                'personal_email': student.personal_email,
                'mobile_primary': student.mobile_primary,
                'current_class': student.current_class,
                'status': student.status,
            }
            
            # Update fields
            for field, value in data.items():
                if hasattr(student, field) and field not in ['id', 'tenant', 'created_at', 'updated_at']:
                    setattr(student, field, value)
            
            # Validate and save
            student.full_clean()
            student.save()
            
            # Create audit log with changes
            changes = {}
            for field, old_value in old_data.items():
                new_value = getattr(student, field)
                if old_value != new_value:
                    changes[field] = {
                        'old': str(old_value),
                        'new': str(new_value)
                    }
            
            AuditService.log_update(
                user=user,
                instance=student,
                old_instance=old_data,
                request=None,
                extra_data={'updated_via': 'service_api', 'changes': changes}
            )
            
            return student, errors
            
        except Student.DoesNotExist:
            errors.append(_("Student not found."))
            return None, errors
        except ValidationError as e:
            errors.extend(e.messages)
            return None, errors
        except Exception as e:
            logger.error(f"Error updating student: {str(e)}", exc_info=True)
            errors.append(str(e))
            return None, errors
    
    @staticmethod
    def delete_student(student_id: int, user: User, reason: str, tenant=None) -> Tuple[bool, List[str]]:
        """
        Soft delete student
        
        Args:
            student_id: Student ID
            user: User deleting the student
            reason: Reason for deletion
            tenant: Tenant object
            
        Returns:
            Tuple of (success status, list of errors)
        """
        errors = []
        
        try:
            # Get student
            student = Student.objects.get(id=student_id, tenant=tenant, is_active=True)
            
            # Check permission
            if not StudentService.has_student_access(user, student):
                raise PermissionDenied(_("You don't have permission to delete this student."))
            
            # Check if student can be deleted
            can_delete, dependencies = StudentService._check_deletion_dependencies(student)
            if not can_delete:
                error_msg = _("Cannot delete student. Dependencies found: ") + ", ".join(dependencies)
                errors.append(error_msg)
                return False, errors
            
            # Perform soft delete
            student.soft_delete(
                deleted_by=user,
                reason=reason,
                category="MANUAL_DELETION"
            )
            
            # Create audit log
            AuditService.log_deletion(
                user=user,
                instance=student,
                request=None,
                hard_delete=False,
                extra_data={
                    'deletion_reason': reason,
                    'deleted_by_user': str(user)
                }
            )
            
            return True, errors
            
        except Student.DoesNotExist:
            errors.append(_("Student not found."))
            return False, errors
        except Exception as e:
            logger.error(f"Error deleting student: {str(e)}", exc_info=True)
            errors.append(str(e))
            return False, errors
    
    @staticmethod
    def _check_deletion_dependencies(student: Student) -> Tuple[bool, List[str]]:
        """
        Check if student has dependencies preventing deletion
        
        Args:
            student: Student object
            
        Returns:
            Tuple of (can_delete, list of dependencies)
        """
        dependencies = []
        
        # Check fee payments
        try:
            from apps.finance.models import FeePayment
            if FeePayment.objects.filter(student=student).exists():
                dependencies.append(_("Fee payments"))
        except ImportError:
            pass
        
        # Check exam results
        try:
            from apps.exams.models import ExamResult
            if ExamResult.objects.filter(student=student).exists():
                dependencies.append(_("Exam results"))
        except ImportError:
            pass
        
        # Check attendance records
        try:
            from apps.attendance.models import Attendance
            if Attendance.objects.filter(student=student).exists():
                dependencies.append(_("Attendance records"))
        except ImportError:
            pass
        
        # Check library records
        try:
            from apps.library.models import BookIssue
            if BookIssue.objects.filter(student=student).exists():
                dependencies.append(_("Library book issues"))
        except ImportError:
            pass
        
        return len(dependencies) == 0, dependencies
    
    @staticmethod
    def get_student_statistics(tenant, user: User = None) -> Dict:
        """
        Get student statistics for dashboard
        
        Args:
            tenant: Tenant object
            user: User requesting statistics
            
        Returns:
            Dictionary of statistics
        """
        try:
            # Use cache for performance
            cache_key = f"student_stats_{tenant.id}_{user.id if user else 'anon'}"
            cached_stats = cache.get(cache_key)
            
            if cached_stats:
                return cached_stats
            
            # Get student queryset
            if user:
                students = StudentService.get_secure_queryset(user, tenant)
            else:
                students = Student.objects.filter(tenant=tenant, is_active=True)
            
            # Calculate statistics
            stats = {
                'total': students.count(),
                'by_status': dict(students.values('status').annotate(count=Count('id')).values_list('status', 'count')),
                'by_gender': dict(students.values('gender').annotate(count=Count('id')).values_list('gender', 'count')),
                'by_category': dict(students.values('category').annotate(count=Count('id')).values_list('category', 'count')),
                'by_class': {},
                'new_this_month': students.filter(
                    created_at__month=timezone.now().month,
                    created_at__year=timezone.now().year
                ).count(),
                'average_age': None,
            }
            
            # Calculate class distribution
            class_stats = students.values('current_class__name').annotate(
                count=Count('id')
            ).order_by('current_class__order')
            
            stats['by_class'] = {
                item['current_class__name']: item['count'] 
                for item in class_stats 
                if item['current_class__name']
            }
            
            # Calculate average age
            from dateutil.relativedelta import relativedelta
            ages = []
            for student in students.filter(date_of_birth__isnull=False):
                age = relativedelta(timezone.now().date(), student.date_of_birth).years
                ages.append(age)
            
            if ages:
                stats['average_age'] = sum(ages) / len(ages)
            
            # Cache for 5 minutes
            cache.set(cache_key, stats, 300)
            
            return stats
            
        except Exception as e:
            logger.error(f"Error getting student statistics: {str(e)}", exc_info=True)
            return {}
    
    @staticmethod
    def search_students(query: str, user: User, tenant=None, limit: int = 20) -> List[Dict]:
        """
        Search students with auto-complete
        
        Args:
            query: Search query
            user: User performing search
            tenant: Tenant object
            limit: Maximum results
            
        Returns:
            List of student dictionaries
        """
        try:
            # Get secure queryset
            students = StudentService.get_secure_queryset(user, tenant)
            
            if not query or len(query) < 2:
                return []
            
            # Search in multiple fields
            results = students.filter(
                Q(admission_number__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(personal_email__icontains=query) |
                Q(mobile_primary__icontains=query) |
                Q(guardians__first_name__icontains=query) |
                Q(guardians__last_name__icontains=query)
            ).select_related('current_class', 'section').distinct()[:limit]
            
            # Format results
            formatted_results = []
            for student in results:
                formatted_results.append({
                    'id': student.id,
                    'admission_number': student.admission_number,
                    'full_name': student.full_name,
                    'first_name': student.first_name,
                    'last_name': student.last_name,
                    'email': student.personal_email,
                    'phone': student.mobile_primary,
                    'class': student.current_class.name if student.current_class else '',
                    'section': student.section.name if student.section else '',
                    'status': student.status,
                    'status_display': student.get_status_display(),
                    'profile_picture': student.profile_picture.url if student.profile_picture else None,
                    'detail_url': f"/students/{student.id}/"
                })
            
            return formatted_results
            
        except Exception as e:
            logger.error(f"Error searching students: {str(e)}", exc_info=True)
            return []
    
    @staticmethod
    def promote_student(student_id: int, new_class_id: int, user: User, tenant=None, remarks: str = "") -> Tuple[bool, List[str]]:
        """
        Promote student to next class
        
        Args:
            student_id: Student ID
            new_class_id: New class ID
            user: User promoting the student
            tenant: Tenant object
            remarks: Promotion remarks
            
        Returns:
            Tuple of (success status, list of errors)
        """
        errors = []
        
        try:
            # Get student
            student = Student.objects.get(id=student_id, tenant=tenant, is_active=True)
            
            # Get new class
            new_class = SchoolClass.objects.get(id=new_class_id, tenant=tenant)
            
            # Check if promotion is valid
            if student.current_class and student.current_class.order >= new_class.order:
                errors.append(_("Cannot promote to same or lower class."))
                return False, errors
            
            # Store old data
            old_class = student.current_class
            old_academic_year = student.academic_year
            
            # Get next academic year
            new_academic_year = AcademicYear.objects.filter(
                tenant=tenant,
                start_date__gt=old_academic_year.start_date if old_academic_year else timezone.now()
            ).order_by('start_date').first()
            
            if not new_academic_year:
                errors.append(_("No next academic year found."))
                return False, errors
            
            # Update student
            student.current_class = new_class
            student.academic_year = new_academic_year
            student.save()
            
            # Create academic history record
            StudentAcademicHistory.objects.create(
                student=student,
                academic_year=old_academic_year,
                class_name=old_class,
                section=student.section,
                promoted_to=new_class,
                promotion_date=timezone.now().date(),
                remarks=remarks,
                promoted_by=user,
                tenant=tenant
            )
            
            # Create audit log
            AuditService.log_update(
                user=user,
                instance=student,
                request=None,
                extra_data={
                    'action': 'promotion',
                    'old_class': str(old_class),
                    'new_class': str(new_class),
                    'old_academic_year': str(old_academic_year),
                    'new_academic_year': str(new_academic_year),
                    'remarks': remarks
                }
            )
            
            # Send notification
            NotificationService.send_student_promotion_notification(
                student=student,
                old_class=old_class,
                new_class=new_class,
                promoted_by=user,
                tenant=tenant
            )
            
            return True, errors
            
        except Student.DoesNotExist:
            errors.append(_("Student not found."))
            return False, errors
        except SchoolClass.DoesNotExist:
            errors.append(_("Class not found."))
            return False, errors
        except Exception as e:
            logger.error(f"Error promoting student: {str(e)}", exc_info=True)
            errors.append(str(e))
            return False, errors
    
    @staticmethod
    def bulk_update_status(student_ids: List[int], new_status: str, user: User, tenant=None, reason: str = "") -> Dict:
        """
        Bulk update student status
        
        Args:
            student_ids: List of student IDs
            new_status: New status value
            user: User performing update
            tenant: Tenant object
            reason: Reason for status change
            
        Returns:
            Dictionary with results
        """
        results = {
            'total': len(student_ids),
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        try:
            # Validate status
            if new_status not in dict(Student.STATUS_CHOICES):
                results['errors'].append(_("Invalid status value."))
                return results
            
            # Get students
            students = Student.objects.filter(
                id__in=student_ids,
                tenant=tenant,
                is_active=True
            )
            
            for student in students:
                try:
                    # Check permission
                    if not StudentService.has_student_access(user, student):
                        results['failed'] += 1
                        results['errors'].append(f"No permission for student {student.admission_number}")
                        continue
                    
                    # Store old status
                    old_status = student.status
                    
                    # Update status
                    student.status = new_status
                    student.save()
                    
                    # Create audit log
                    AuditService.log_update(
                        user=user,
                        instance=student,
                        request=None,
                        extra_data={
                            'action': 'bulk_status_update',
                            'old_status': old_status,
                            'new_status': new_status,
                            'reason': reason
                        }
                    )
                    
                    results['success'] += 1
                    
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"Student {student.admission_number}: {str(e)}")
                    logger.error(f"Error updating student {student.id}: {str(e)}")
            
            return results
            
        except Exception as e:
            logger.error(f"Error in bulk status update: {str(e)}", exc_info=True)
            results['errors'].append(str(e))
            return results


class GuardianService:
    """
    Service for guardian-related operations
    """
    
    @staticmethod
    def get_student_guardians(student: Student, user: User = None) -> models.QuerySet:
        """
        Get guardians for student with permission check
        
        Args:
            student: Student object
            user: User requesting guardians
            
        Returns:
            Guardian queryset
        """
        try:
            # Check permission
            if user and not StudentService.has_student_access(user, student):
                return Guardian.objects.none()
            
            return Guardian.objects.filter(
                student=student,
                is_active=True
            ).order_by('-is_primary', 'created_at')
            
        except Exception as e:
            logger.error(f"Error getting student guardians: {str(e)}", exc_info=True)
            return Guardian.objects.none()
    
    @staticmethod
    def create_guardian(data: Dict, student: Student, user: User, tenant=None) -> Tuple[Optional[Guardian], List[str]]:
        """
        Create guardian for student
        
        Args:
            data: Guardian data
            student: Student object
            user: User creating guardian
            tenant: Tenant object
            
        Returns:
            Tuple of (guardian object, list of errors)
        """
        errors = []
        
        try:
            from .forms import GuardianForm
            
            # Prepare form data
            form_data = data.copy()
            form_data['student'] = student
            form_data['tenant'] = tenant
            
            # Create form instance
            form = GuardianForm(data=form_data, student=student, tenant=tenant, user=user)
            
            if form.is_valid():
                # Save guardian
                guardian = form.save(commit=False)
                guardian.tenant = tenant
                guardian.save()
                
                # Add to student
                student.guardians.add(guardian)
                
                # Create audit log
                AuditService.log_creation(
                    user=user,
                    instance=guardian,
                    request=None,
                    extra_data={
                        'student_id': str(student.id),
                        'created_via': 'service_api'
                    }
                )
                
                return guardian, errors
            else:
                # Collect form errors
                for field, field_errors in form.errors.items():
                    for error in field_errors:
                        errors.append(f"{field}: {error}")
                
                return None, errors
                
        except Exception as e:
            logger.error(f"Error creating guardian: {str(e)}", exc_info=True)
            errors.append(str(e))
            return None, errors
    
    @staticmethod
    def update_guardian(guardian_id: int, data: Dict, user: User, tenant=None) -> Tuple[Optional[Guardian], List[str]]:
        """
        Update guardian
        
        Args:
            guardian_id: Guardian ID
            data: Update data
            user: User updating guardian
            tenant: Tenant object
            
        Returns:
            Tuple of (updated guardian object, list of errors)
        """
        errors = []
        
        try:
            # Get guardian
            guardian = Guardian.objects.get(id=guardian_id, tenant=tenant, is_active=True)
            
            # Check permission via student
            if not StudentService.has_student_access(user, guardian.student):
                raise PermissionDenied(_("You don't have permission to update this guardian."))
            
            # Store old data for audit
            old_data = {
                'first_name': guardian.first_name,
                'last_name': guardian.last_name,
                'email': guardian.email,
                'phone': guardian.phone,
                'relationship': guardian.relationship,
                'is_primary': guardian.is_primary,
            }
            
            # Update fields
            for field, value in data.items():
                if hasattr(guardian, field) and field not in ['id', 'tenant', 'created_at', 'updated_at']:
                    setattr(guardian, field, value)
            
            # Validate and save
            guardian.full_clean()
            guardian.save()
            
            # Create audit log with changes
            changes = {}
            for field, old_value in old_data.items():
                new_value = getattr(guardian, field)
                if old_value != new_value:
                    changes[field] = {
                        'old': str(old_value),
                        'new': str(new_value)
                    }
            
            AuditService.log_update(
                user=user,
                instance=guardian,
                old_instance=old_data,
                request=None,
                extra_data={'updated_via': 'service_api', 'changes': changes}
            )
            
            return guardian, errors
            
        except Guardian.DoesNotExist:
            errors.append(_("Guardian not found."))
            return None, errors
        except ValidationError as e:
            errors.extend(e.messages)
            return None, errors
        except Exception as e:
            logger.error(f"Error updating guardian: {str(e)}", exc_info=True)
            errors.append(str(e))
            return None, errors
    
    @staticmethod
    def set_primary_guardian(student: Student, guardian_id: int, user: User, tenant=None) -> Tuple[bool, List[str]]:
        """
        Set primary guardian for student
        
        Args:
            student: Student object
            guardian_id: Guardian ID to set as primary
            user: User performing action
            tenant: Tenant object
            
        Returns:
            Tuple of (success status, list of errors)
        """
        errors = []
        
        try:
            # Check permission
            if not StudentService.has_student_access(user, student):
                raise PermissionDenied(_("You don't have permission to modify guardians."))
            
            # Get guardian
            guardian = Guardian.objects.get(id=guardian_id, student=student, tenant=tenant, is_active=True)
            
            # Remove primary status from all other guardians
            Guardian.objects.filter(
                student=student,
                is_primary=True
            ).update(is_primary=False)
            
            # Set new primary guardian
            guardian.is_primary = True
            guardian.save()
            
            # Create audit log
            AuditService.log_update(
                user=user,
                instance=guardian,
                request=None,
                extra_data={
                    'action': 'set_primary_guardian',
                    'student_id': str(student.id)
                }
            )
            
            return True, errors
            
        except Guardian.DoesNotExist:
            errors.append(_("Guardian not found."))
            return False, errors
        except Exception as e:
            logger.error(f"Error setting primary guardian: {str(e)}", exc_info=True)
            errors.append(str(e))
            return False, errors


class DocumentService:
    """
    Service for student document operations
    """
    
    @staticmethod
    def get_student_documents(student: Student, user: User = None) -> models.QuerySet:
        """
        Get documents for student with permission check
        
        Args:
            student: Student object
            user: User requesting documents
            
        Returns:
            StudentDocument queryset
        """
        try:
            # Check permission
            if user and not StudentService.has_student_access(user, student):
                return StudentDocument.objects.none()
            
            return StudentDocument.objects.filter(
                student=student,
                is_active=True
            ).order_by('-uploaded_at', 'doc_type')
            
        except Exception as e:
            logger.error(f"Error getting student documents: {str(e)}", exc_info=True)
            return StudentDocument.objects.none()
    
    @staticmethod
    def get_allowed_document_types() -> List[Dict]:
        """
        Get allowed document types and their validation rules
        
        Returns:
            List of document type dictionaries
        """
        return [
            {
                'value': 'PHOTO',
                'label': _('Student Photo'),
                'extensions': ['.jpg', '.jpeg', '.png', '.gif'],
                'max_size_mb': 5,
                'required': True
            },
            {
                'value': 'BIRTH_CERTIFICATE',
                'label': _('Birth Certificate'),
                'extensions': ['.pdf', '.jpg', '.jpeg', '.png'],
                'max_size_mb': 10,
                'required': True
            },
            {
                'value': 'ADDRESS_PROOF',
                'label': _('Address Proof'),
                'extensions': ['.pdf', '.jpg', '.jpeg', '.png'],
                'max_size_mb': 10,
                'required': True
            },
            {
                'value': 'AADHAR_CARD',
                'label': _('Aadhar Card'),
                'extensions': ['.pdf', '.jpg', '.jpeg', '.png'],
                'max_size_mb': 10,
                'required': False
            },
            {
                'value': 'PREVIOUS_MARKSHEET',
                'label': _('Previous Marksheet'),
                'extensions': ['.pdf', '.jpg', '.jpeg', '.png'],
                'max_size_mb': 10,
                'required': False
            },
            {
                'value': 'TRANSFER_CERTIFICATE',
                'label': _('Transfer Certificate'),
                'extensions': ['.pdf', '.jpg', '.jpeg', '.png'],
                'max_size_mb': 10,
                'required': False
            },
            {
                'value': 'MEDICAL_CERTIFICATE',
                'label': _('Medical Certificate'),
                'extensions': ['.pdf', '.jpg', '.jpeg', '.png'],
                'max_size_mb': 5,
                'required': False
            },
            {
                'value': 'OTHER',
                'label': _('Other Document'),
                'extensions': ['.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png'],
                'max_size_mb': 20,
                'required': False
            },
        ]
    
    @staticmethod
    def get_max_file_size_mb() -> int:
        """
        Get maximum file size in MB
        
        Returns:
            Maximum file size in MB
        """
        return 25  # 25MB default
    
    @staticmethod
    def validate_document_file(file, doc_type: str = None) -> Tuple[bool, List[str]]:
        """
        Validate document file
        
        Args:
            file: Uploaded file object
            doc_type: Document type for specific validation
            
        Returns:
            Tuple of (is_valid, list of errors)
        """
        errors = []
        
        try:
            # Check file exists
            if not file:
                errors.append(_("No file provided."))
                return False, errors
            
            # Get allowed types
            allowed_types = DocumentService.get_allowed_document_types()
            type_config = None
            
            if doc_type:
                for config in allowed_types:
                    if config['value'] == doc_type:
                        type_config = config
                        break
            
            # Check file size
            max_size_mb = DocumentService.get_max_file_size_mb()
            max_size_bytes = max_size_mb * 1024 * 1024
            
            if file.size > max_size_bytes:
                errors.append(_(f"File size exceeds maximum allowed size of {max_size_mb}MB."))
            
            # Check file extension
            import os
            file_name = file.name.lower()
            file_ext = os.path.splitext(file_name)[1]
            
            if type_config:
                # Type-specific validation
                if file_ext not in type_config['extensions']:
                    allowed_exts = ', '.join(type_config['extensions'])
                    errors.append(_(f"Invalid file type. Allowed types: {allowed_exts}"))
                
                type_max_size = type_config['max_size_mb'] * 1024 * 1024
                if file.size > type_max_size:
                    errors.append(_(f"File size exceeds maximum allowed size of {type_config['max_size_mb']}MB for this document type."))
            else:
                # General validation - allow common document types
                allowed_extensions = ['.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png', '.gif']
                if file_ext not in allowed_extensions:
                    errors.append(_(f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}"))
            
            # Check file name length
            if len(file.name) > 255:
                errors.append(_("File name is too long. Maximum 255 characters allowed."))
            
            # Check for malicious file names
            import re
            if re.search(r'[<>:"/\\|?*]', file.name):
                errors.append(_("File name contains invalid characters."))
            
            return len(errors) == 0, errors
            
        except Exception as e:
            logger.error(f"Error validating document file: {str(e)}", exc_info=True)
            errors.append(str(e))
            return False, errors
    
    @staticmethod
    def create_document(data: Dict, student: Student, user: User, tenant=None) -> Tuple[Optional[StudentDocument], List[str]]:
        """
        Create student document
        
        Args:
            data: Document data including file
            student: Student object
            user: User uploading document
            tenant: Tenant object
            
        Returns:
            Tuple of (document object, list of errors)
        """
        errors = []
        
        try:
            from .forms import StudentDocumentForm
            
            # Prepare form data
            form_data = data.copy()
            form_data['student'] = student
            form_data['tenant'] = tenant
            
            # Create form instance
            form = StudentDocumentForm(
                data=form_data,
                files=data.get('files'),
                student=student,
                tenant=tenant,
                user=user
            )
            
            if form.is_valid():
                # Save document
                document = form.save(commit=False)
                document.tenant = tenant
                document.uploaded_by = user
                document.save()
                
                # Create audit log
                AuditService.log_creation(
                    user=user,
                    instance=document,
                    request=None,
                    extra_data={
                        'student_id': str(student.id),
                        'document_type': document.doc_type,
                        'file_size': document.file.size if document.file else 0,
                        'created_via': 'service_api'
                    }
                )
                
                return document, errors
            else:
                # Collect form errors
                for field, field_errors in form.errors.items():
                    for error in field_errors:
                        errors.append(f"{field}: {error}")
                
                return None, errors
                
        except Exception as e:
            logger.error(f"Error creating document: {str(e)}", exc_info=True)
            errors.append(str(e))
            return None, errors
    
    @staticmethod
    def verify_document(document_id: int, user: User, remarks: str = "", tenant=None) -> Tuple[bool, List[str]]:
        """
        Verify student document
        
        Args:
            document_id: Document ID
            user: User verifying document
            remarks: Verification remarks
            tenant: Tenant object
            
        Returns:
            Tuple of (success status, list of errors)
        """
        errors = []
        
        try:
            # Get document
            document = StudentDocument.objects.get(id=document_id, tenant=tenant, is_active=True)
            
            # Check permission via student
            if not StudentService.has_student_access(user, document.student):
                raise PermissionDenied(_("You don't have permission to verify this document."))
            
            # Verify document
            document.is_verified = True
            document.verified_by = user
            document.verified_at = timezone.now()
            document.verification_remarks = remarks
            document.save()
            
            # Create audit log
            AuditService.log_update(
                user=user,
                instance=document,
                request=None,
                extra_data={
                    'action': 'document_verification',
                    'student_id': str(document.student.id),
                    'remarks': remarks
                }
            )
            
            # Send notification to student
            if hasattr(document.student, 'user') and document.student.user:
                NotificationService.send_in_app_notification(
                    recipient=document.student.user,
                    title="Document Verified",
                    message=f"Your document '{document.file_name}' has been verified.",
                    notification_type="ACADEMIC",
                    sender=user,
                    action_url=f"/students/{document.student.id}/documents/",
                    action_text="View Documents",
                    tenant=tenant
                )
            
            return True, errors
            
        except StudentDocument.DoesNotExist:
            errors.append(_("Document not found."))
            return False, errors
        except Exception as e:
            logger.error(f"Error verifying document: {str(e)}", exc_info=True)
            errors.append(str(e))
            return False, errors


class StudentImportService:
    """
    Service for importing students from various formats
    """
    
    @staticmethod
    def process_csv(file, tenant, academic_year, update_existing=False, skip_errors=False, uploaded_by=None, send_welcome_email=False):
        """
        Process CSV file for student import
        
        Returns dictionary with results
        """
        # Implementation would go here
        # This is a placeholder - you would implement the actual CSV processing logic
        return {
            'success': 'Import processed successfully (CSV placeholder)',
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'total_rows': 0
        }
    
    @staticmethod
    def process_excel(file, tenant, academic_year, update_existing=False, skip_errors=False, uploaded_by=None, send_welcome_email=False):
        """
        Process Excel file for student import
        
        Returns dictionary with results
        """
        # Implementation would go here
        # This is a placeholder - you would implement the actual Excel processing logic
        return {
            'success': 'Import processed successfully (Excel placeholder)',
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'total_rows': 0
        }
    
    @staticmethod
    def validate_upload_file(file, tenant, user):
        """
        Validate uploaded file
        
        Returns validation result dictionary
        """
        # Implementation would go here
        return {
            'valid': True,
            'row_count': 0
        }


class StudentExportService:
    """
    Service for exporting student data
    """
    
    @staticmethod
    def export_to_csv(queryset, export_type='basic'):
        """
        Export students to CSV format
        
        Returns HTTP response with CSV file
        """
        # Implementation would go here
        from django.http import HttpResponse
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Admission Number', 'Full Name', 'Email', 'Status'])
        
        for student in queryset:
            writer.writerow([
                student.admission_number,
                student.full_name,
                student.personal_email,
                student.get_status_display()
            ])
        
        return response
    
    @staticmethod
    def export_to_excel(queryset, export_type='basic'):
        """
        Export students to Excel format
        
        Returns HTTP response with Excel file
        """
        # Implementation would go here
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        headers = ['Admission Number', 'Full Name', 'Email', 'Status']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        for row_idx, student in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=student.admission_number)
            ws.cell(row=row_idx, column=2, value=student.full_name)
            ws.cell(row=row_idx, column=3, value=student.personal_email)
            ws.cell(row=row_idx, column=4, value=student.get_status_display())
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
        wb.save(response)
        
        return response


class StudentValidationService:
    """
    Service for student data validation
    """
    
    @staticmethod
    def validate_student_data(data: Dict, tenant=None) -> Tuple[bool, Dict[str, List[str]]]:
        """
        Validate student data
        
        Args:
            data: Student data dictionary
            tenant: Tenant object
            
        Returns:
            Tuple of (is_valid, dictionary of field errors)
        """
        errors = {}
        
        try:
            # Required fields
            required_fields = ['first_name', 'last_name', 'gender', 'date_of_birth']
            for field in required_fields:
                if not data.get(field):
                    if field not in errors:
                        errors[field] = []
                    errors[field].append(_("This field is required."))
            
            # Email validation
            email = data.get('personal_email')
            if email:
                from django.core.validators import validate_email
                try:
                    validate_email(email)
                    
                    # Check for duplicate email
                    if Student.objects.filter(
                        tenant=tenant,
                        personal_email=email,
                        is_active=True
                    ).exists():
                        errors.setdefault('personal_email', []).append(
                            _("A student with this email already exists.")
                        )
                except ValidationError:
                    errors.setdefault('personal_email', []).append(
                        _("Enter a valid email address.")
                    )
            
            # Phone validation
            phone = data.get('mobile_primary')
            if phone:
                from django.core.validators import RegexValidator
                phone_regex = RegexValidator(
                    regex=r'^\+?1?\d{9,15}$',
                    message=_("Phone number must be entered in the format: '+999999999'. Up to 15 digits allowed.")
                )
                try:
                    phone_regex(phone)
                except ValidationError:
                    errors.setdefault('mobile_primary', []).append(
                        _("Enter a valid phone number.")
                    )
            
            # Date of birth validation
            dob = data.get('date_of_birth')
            if dob:
                try:
                    if isinstance(dob, str):
                        from datetime import datetime
                        dob = datetime.strptime(dob, '%Y-%m-%d').date()
                    
                    # Check if student is at least 3 years old
                    from dateutil.relativedelta import relativedelta
                    age = relativedelta(timezone.now().date(), dob).years
                    if age < 3:
                        errors.setdefault('date_of_birth', []).append(
                            _("Student must be at least 3 years old.")
                        )
                    
                    # Check if student is not older than 25 years
                    if age > 25:
                        errors.setdefault('date_of_birth', []).append(
                            _("Student cannot be older than 25 years.")
                        )
                except (ValueError, TypeError):
                    errors.setdefault('date_of_birth', []).append(
                        _("Enter a valid date in YYYY-MM-DD format.")
                    )
            
            # Class validation
            class_id = data.get('current_class')
            if class_id:
                try:
                    school_class = SchoolClass.objects.get(id=class_id, tenant=tenant)
                    # Additional class validation can be added here
                except SchoolClass.DoesNotExist:
                    errors.setdefault('current_class', []).append(
                        _("Selected class does not exist.")
                    )
            
            # Academic year validation
            academic_year_id = data.get('academic_year')
            if academic_year_id:
                try:
                    academic_year = AcademicYear.objects.get(id=academic_year_id, tenant=tenant)
                    # Check if academic year is active
                    if not academic_year.is_active:
                        errors.setdefault('academic_year', []).append(
                            _("Selected academic year is not active.")
                        )
                except AcademicYear.DoesNotExist:
                    errors.setdefault('academic_year', []).append(
                        _("Selected academic year does not exist.")
                    )
            
            return len(errors) == 0, errors
            
        except Exception as e:
            logger.error(f"Error validating student data: {str(e)}", exc_info=True)
            errors['non_field_errors'] = [str(e)]
            return False, errors
    
    @staticmethod
    def validate_guardian_data(data: Dict, tenant=None) -> Tuple[bool, Dict[str, List[str]]]:
        """
        Validate guardian data
        
        Args:
            data: Guardian data dictionary
            tenant: Tenant object
            
        Returns:
            Tuple of (is_valid, dictionary of field errors)
        """
        errors = {}
        
        try:
            # Required fields
            required_fields = ['first_name', 'last_name', 'relationship']
            for field in required_fields:
                if not data.get(field):
                    if field not in errors:
                        errors[field] = []
                    errors[field].append(_("This field is required."))
            
            # Email validation
            email = data.get('email')
            if email:
                from django.core.validators import validate_email
                try:
                    validate_email(email)
                except ValidationError:
                    errors.setdefault('email', []).append(
                        _("Enter a valid email address.")
                    )
            
            # Phone validation
            phone = data.get('phone')
            if phone:
                from django.core.validators import RegexValidator
                phone_regex = RegexValidator(
                    regex=r'^\+?1?\d{9,15}$',
                    message=_("Phone number must be entered in the format: '+999999999'. Up to 15 digits allowed.")
                )
                try:
                    phone_regex(phone)
                except ValidationError:
                    errors.setdefault('phone', []).append(
                        _("Enter a valid phone number.")
                    )
            
            # Relationship validation
            relationship = data.get('relationship')
            if relationship and relationship not in dict(Guardian.RELATIONSHIP_CHOICES):
                errors.setdefault('relationship', []).append(
                    _("Invalid relationship value.")
                )
            
            return len(errors) == 0, errors
            
        except Exception as e:
            logger.error(f"Error validating guardian data: {str(e)}", exc_info=True)
            errors['non_field_errors'] = [str(e)]
            return False, errors
    
    @staticmethod
    def check_admission_number_availability(admission_number: str, tenant, exclude_student_id: int = None) -> Tuple[bool, str]:
        """
        Check if admission number is available
        
        Args:
            admission_number: Admission number to check
            tenant: Tenant object
            exclude_student_id: Student ID to exclude from check
            
        Returns:
            Tuple of (is_available, message)
        """
        try:
            # Check if admission number is empty
            if not admission_number or admission_number.strip() == '':
                return False, _("Admission number cannot be empty.")
            
            # Check format (basic validation)
            if len(admission_number) < 3:
                return False, _("Admission number is too short.")
            
            # Check for duplicates
            query = Student.objects.filter(
                tenant=tenant,
                admission_number=admission_number,
                is_active=True
            )
            
            if exclude_student_id:
                query = query.exclude(id=exclude_student_id)
            
            if query.exists():
                return False, _("Admission number already exists.")
            
            return True, _("Admission number is available.")
            
        except Exception as e:
            logger.error(f"Error checking admission number: {str(e)}", exc_info=True)
            return False, str(e)