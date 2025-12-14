
from datetime import datetime
import csv
import openpyxl
import io
import logging
import re
from django.utils import timezone
from django.conf import settings


from django.core.exceptions import ValidationError, PermissionDenied
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, IntegrityError, models
from django.db.models import Q, Count, Avg, Sum
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse, FileResponse, Http404, JsonResponse, HttpResponseForbidden
from django.urls import reverse_lazy
from django.views.decorators.csrf import csrf_protect
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.views.generic import View

logger = logging.getLogger(__name__)

# Core imports
from apps.core.views import (
    BaseView, BaseListView, BaseCreateView, 
    BaseUpdateView, BaseDeleteView, BaseDetailView
)
from apps.core.services.audit_service import AuditService
from apps.core.services.notification_service import NotificationService
from apps.core.middleware.tenant import get_current_tenant
from apps.core.permissions.mixins import (
    PermissionRequiredMixin, RoleRequiredMixin, 
    TenantAccessMixin, ObjectPermissionMixin,
    RoleBasedViewMixin, TenantRequiredMixin,RateLimitedViewMixin
)
from apps.core.permissions.decorators import (
    rate_limit_by_role,
    permission_required,
    role_required,
    min_role_level,
    require_tenant_access,
    require_https
)


# Model imports
from apps.academics.models import AcademicYear, SchoolClass, Section, Stream, Subject
from apps.finance.models import FeeStructure, Payment
from .models import Student, Guardian, StudentAddress, StudentDocument, StudentAcademicHistory
from .forms import (
    StudentForm, GuardianForm, StudentDocumentForm, 
    StudentFilterForm, StudentBulkUploadForm,
    StudentPromotionForm, StudentImportForm
)
from .services import (
    StudentService, GuardianService, DocumentService,
    StudentImportService, StudentExportService,
    StudentValidationService
)
from .idcard import StudentIDCardGenerator
from .serializers import StudentSerializer, GuardianSerializer
from .tasks import (
    process_bulk_upload_async, generate_id_cards_batch,
    send_student_welcome_email, export_student_data_async
)
from .constants import (
    StudentStatus, StudentGender, StudentCategory,
    MAX_BULK_UPLOAD_SIZE, ALLOWED_UPLOAD_FORMATS,
    REQUIRED_STUDENT_FIELDS
)


# ============================================================================
# DECORATORS & MIXINS
# ============================================================================


class StudentPermissionMixin(PermissionRequiredMixin, TenantAccessMixin):
    """Mixin for student-related permissions"""
    tenant_field = 'tenant'
    
    def get_permission_required(self):
        """Dynamically determine required permissions"""
        if hasattr(self, 'permission_required'):
            return [self.permission_required]
        return ['students.view_student']


class SecureTransactionMixin:
    """Mixin for secure database transactions"""
    
    @method_decorator(transaction.atomic)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


# ============================================================================
# DASHBOARD VIEWS
# ============================================================================

class StudentDashboardView(StudentPermissionMixin, BaseView):
    """
    Comprehensive student dashboard with real-time analytics,
    audit logging, and secure data access.
    """
    template_name = 'students/dashboard.html'
    permission_required = 'students.view_student_dashboard'
    
    @method_decorator(login_required)
    @method_decorator(require_https)
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = self.request.tenant
        user = self.request.user
        
        try:
            # Get secure student queryset
            students = StudentService.get_secure_queryset(user, tenant)
            
            # Calculate statistics with audit logging
            statistics = self.calculate_statistics(students, tenant)
            context.update(statistics)
            
            # Get performance metrics
            performance = self.get_performance_metrics(students, tenant)
            context.update(performance)
            
            # Get recent activities
            context['recent_activities'] = self.get_recent_activities(tenant)
            
            # Get upcoming events
            context['upcoming_events'] = self.get_upcoming_events(tenant)
            
            # Audit dashboard access
            AuditService.create_audit_entry(
                action='READ',
                resource_type='StudentDashboard',
                user=user,
                tenant=tenant,
                request=self.request,
                severity='INFO',
                extra_data={
                    'dashboard_type': 'student_overview',
                    'statistics_generated': True
                }
            )
            
        except Exception as e:
            logger.error(f"Dashboard error: {str(e)}", exc_info=True)
            messages.error(self.request, _("Error loading dashboard data"))
            context['error'] = True
        
        return context
    
    def calculate_statistics(self, students, tenant):
        """Calculate comprehensive student statistics"""
        total = students.count()
        
        return {
            'total_students': total,
            'active_students': students.filter(status=StudentStatus.ACTIVE).count(),
            'inactive_students': students.filter(status=StudentStatus.INACTIVE).count(),
            'graduated_students': students.filter(status=StudentStatus.GRADUATED).count(),
            'new_admissions': students.filter(
                created_at__gte=timezone.now() - timezone.timedelta(days=30)
            ).count(),
            'pending_documents': StudentDocument.objects.filter(
                student__tenant=tenant,
                is_verified=False
            ).count(),
            'class_distribution': self.get_class_distribution(students),
            'gender_distribution': self.get_gender_distribution(students),
            'category_distribution': self.get_category_distribution(students),
        }
    
    def get_class_distribution(self, students):
        """Get distribution of students by class"""
        distribution = []
        classes = SchoolClass.objects.filter(tenant=self.request.tenant)
        
        for school_class in classes:
            count = students.filter(current_class=school_class).count()
            if count > 0:
                distribution.append({
                    'name': school_class.name,
                    'count': count,
                    'percentage': (count / students.count() * 100) if students.count() > 0 else 0
                })
        
        return sorted(distribution, key=lambda x: x['count'], reverse=True)[:10]
    
    def get_gender_distribution(self, students):
        """Get gender distribution"""
        return students.values('gender').annotate(
            count=Count('id'),
            percentage=(Count('id') * 100.0 / models.Count('id', filter=models.Q(id__in=students.values('id'))))
        )
    
    def get_category_distribution(self, students):
        """Get category distribution"""
        return students.values('category').annotate(
            count=Count('id')
        ).order_by('-count')
    
    def get_performance_metrics(self, students, tenant):
        """Get student performance metrics"""
        try:
            from apps.exams.models import ExamResult
            from apps.attendance.models import Attendance
            
            # Get average performance
            avg_performance = ExamResult.objects.filter(
                student__in=students,
                is_published=True
            ).aggregate(
                avg_score=Avg('total_marks'),
                pass_rate=Avg(models.Case(
                    models.When(percentage__gte=40, then=1),
                    default=0,
                    output_field=models.FloatField()
                ))
            )
            
            # Get attendance metrics
            attendance_metrics = Attendance.objects.filter(
                student__in=students,
                date__gte=timezone.now() - timezone.timedelta(days=30)
            ).aggregate(
                avg_attendance=Avg(models.Case(
                    models.When(status='PRESENT', then=1),
                    default=0,
                    output_field=models.FloatField()
                ))
            )
            
            return {
                'avg_performance': avg_performance,
                'attendance_metrics': attendance_metrics,
                'top_performers': self.get_top_performers(students),
                'improvement_areas': self.get_improvement_areas(students),
            }
            
        except Exception as e:
            logger.error(f"Performance metrics error: {str(e)}")
            return {}
    
    def get_top_performers(self, students):
        """Get top performing students"""
        try:
            from apps.exams.models import ExamResult
            
            return ExamResult.objects.filter(
                student__in=students,
                is_published=True
            ).select_related('student').values(
                'student__id',
                'student__first_name',
                'student__last_name',
                'student__current_class__name'
            ).annotate(
                avg_score=Avg('percentage')
            ).order_by('-avg_score')[:5]
            
        except Exception:
            return []
    
    def get_improvement_areas(self, students):
        """Get areas needing improvement"""
        try:
            from apps.exams.models import ExamResult, ExamSubjectResult
            
            # Find subjects with lowest average scores
            weak_subjects = ExamSubjectResult.objects.filter(
                exam_result__student__in=students
            ).values(
                'subject__name',
                'subject__code'
            ).annotate(
                avg_score=Avg('marks_obtained'),
                total_students=Count('exam_result__student', distinct=True)
            ).filter(
                avg_score__lt=40  # Below passing threshold
            ).order_by('avg_score')[:5]
            
            return list(weak_subjects)
            
        except Exception:
            return []
    
    def get_recent_activities(self, tenant):
        """Get recent student-related activities"""
        try:
            from apps.core.models import AuditLog
            
            return AuditLog.objects.filter(
                resource_type__in=['Student', 'Guardian', 'StudentDocument'],
                tenant=tenant
            ).select_related('user').order_by('-timestamp')[:10]
            
        except Exception:
            return []
    
    def get_upcoming_events(self, tenant):
        """Get upcoming student events"""
        try:
            from apps.calendar.models import Event
            
            return Event.objects.filter(
                tenant=tenant,
                start_date__gte=timezone.now(),
                participants__student__isnull=False
            ).distinct().order_by('start_date')[:5]
            
        except Exception:
            return []


# ============================================================================
# CRUD VIEWS WITH TENANT & AUDIT INTEGRATION
# ============================================================================

class StudentListView(StudentPermissionMixin, BaseListView):
    """
    Secure student listing with advanced filtering,
    search, and export capabilities.
    """
    model = Student
    template_name = 'students/student_list.html'
    context_object_name = 'students'
    permission_required = 'students.view_student'
    paginate_by = 25
    
    @method_decorator(login_required)
    @method_decorator(require_https)
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)
    
    def get_queryset(self):
        """Get secure, filtered queryset"""
        queryset = StudentService.get_secure_queryset(
            self.request.user,
            self.request.tenant
        )
        
        # Apply filters
        queryset = self.apply_filters(queryset)
        
        # Apply sorting
        queryset = self.apply_sorting(queryset)
        
        # Prefetch related data for performance
        queryset = queryset.select_related(
            'current_class', 'section', 'academic_year', 'stream'
        ).prefetch_related(
            'guardians', 'addresses', 'documents'
        ).only(
            'id', 'admission_number', 'first_name', 'last_name',
            'personal_email', 'mobile_primary', 'status',
            'current_class__name', 'section__name',
            'academic_year__name', 'created_at'
        )
        
        return queryset
    
    def apply_filters(self, queryset):
        """Apply various filters from request"""
        filters = Q()
        
        # Search filter
        search = self.request.GET.get('search', '').strip()
        if search:
            filters &= Q(
                Q(admission_number__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(personal_email__icontains=search) |
                Q(mobile_primary__icontains=search) |
                Q(guardians__first_name__icontains=search) |
                Q(guardians__last_name__icontains=search)
            )
        
        # Status filter
        status = self.request.GET.get('status')
        if status in dict(StudentStatus.choices):
            filters &= Q(status=status)
        
        # Class filter
        class_id = self.request.GET.get('class_id')
        if class_id and class_id.isdigit():
            filters &= Q(current_class_id=int(class_id))
        
        # Section filter
        section_id = self.request.GET.get('section_id')
        if section_id and section_id.isdigit():
            filters &= Q(section_id=int(section_id))
        
        # Academic year filter
        academic_year_id = self.request.GET.get('academic_year_id')
        if academic_year_id and academic_year_id.isdigit():
            filters &= Q(academic_year_id=int(academic_year_id))
        
        # Category filter
        category = self.request.GET.get('category')
        if category in dict(StudentCategory.choices):
            filters &= Q(category=category)
        
        # Gender filter
        gender = self.request.GET.get('gender')
        if gender in dict(StudentGender.choices):
            filters &= Q(gender=gender)
        
        # Date range filters
        created_from = self.request.GET.get('created_from')
        created_to = self.request.GET.get('created_to')
        
        if created_from:
            try:
                date_from = datetime.strptime(created_from, '%Y-%m-%d').date()
                filters &= Q(created_at__date__gte=date_from)
            except ValueError:
                pass
        
        if created_to:
            try:
                date_to = datetime.strptime(created_to, '%Y-%m-%d').date()
                filters &= Q(created_at__date__lte=date_to)
            except ValueError:
                pass
        
        return queryset.filter(filters).distinct()
    
    def apply_sorting(self, queryset):
        """Apply sorting based on request parameters"""
        sort_by = self.request.GET.get('sort_by', 'created_at')
        sort_order = self.request.GET.get('sort_order', 'desc')
        
        # Validate sort field
        valid_sort_fields = [
            'admission_number', 'first_name', 'last_name',
            'current_class__name', 'created_at', 'status'
        ]
        
        if sort_by not in valid_sort_fields:
            sort_by = 'created_at'
        
        # Apply sorting
        if sort_order == 'asc':
            queryset = queryset.order_by(sort_by)
        else:
            queryset = queryset.order_by(f'-{sort_by}')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add filter form
        context['filter_form'] = StudentFilterForm(
            tenant=self.request.tenant,
            initial=self.request.GET.dict(),
            user=self.request.user
        )
        
        # Add classes for filter dropdown
        context['classes'] = SchoolClass.objects.filter(
            tenant=self.request.tenant
        ).order_by('order', 'name')
        
        # Add academic years
        context['academic_years'] = AcademicYear.objects.filter(
            tenant=self.request.tenant
        ).order_by('-is_current', '-start_date')
        
        # Add status counts for filter display
        queryset = StudentService.get_secure_queryset(
            self.request.user,
            self.request.tenant
        )
        
        context['status_counts'] = dict(
            queryset.values('status').annotate(
                count=Count('id')
            ).values_list('status', 'count')
        )
        
        # Add export formats
        context['export_formats'] = [
            {'value': 'csv', 'label': 'CSV'},
            {'value': 'excel', 'label': 'Excel'},
            {'value': 'pdf', 'label': 'PDF (with details)'},
        ]
        
        # Log list view access
        AuditService.create_audit_entry(
            action='READ',
            resource_type='StudentList',
            user=self.request.user,
            tenant=self.request.tenant,
            request=self.request,
            severity='INFO',
            extra_data={
                'total_count': context['paginator'].count if hasattr(context, 'paginator') else 0,
                'filters_applied': dict(self.request.GET),
                'page_number': self.request.GET.get('page', 1)
            }
        )
        
        return context


class StudentDetailView(StudentPermissionMixin, BaseDetailView):
    """
    Detailed student view with comprehensive information,
    audit logging, and permission checking.
    """
    template_name = 'students/student_detail.html'
    permission_required = 'students.view_student'
    
    def get_object(self, queryset=None):
        """Get student with tenant isolation and permission checking"""
        student_id = self.kwargs.get('pk')
        
        student = get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=student_id
        )
        
        # Additional permission check
        if not self.request.user.has_perm('students.view_student', student):
            raise PermissionDenied(_("You don't have permission to view this student."))
        
        return student
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        try:
            # Student information
            context['student'] = student
            
            # Guardian information
            context['guardians'] = GuardianService.get_student_guardians(
                student, self.request.user
            )
            
            # Address information
            context['addresses'] = student.addresses.filter(
                is_active=True
            ).order_by('-is_primary')
            
            # Document information
            context['documents'] = DocumentService.get_student_documents(
                student, self.request.user
            )
            
            # Academic history
            context['academic_history'] = StudentAcademicHistory.objects.filter(
                student=student
            ).order_by('-academic_year__start_date')
            
            # Performance metrics
            context.update(self.get_performance_data(student))
            
            # Fee information
            context.update(self.get_fee_information(student))
            
            # Attendance information
            context.update(self.get_attendance_data(student))
            
            # Audit the detail view access
            AuditService.log_view(
                user=self.request.user,
                instance=student,
                request=self.request,
                extra_data={'view_type': 'student_detail'}
            )
            
        except Exception as e:
            logger.error(f"Error loading student details: {str(e)}", exc_info=True)
            messages.error(self.request, _("Error loading student information"))
        
        return context
    
    def get_performance_data(self, student):
        """Get student performance data"""
        try:
            from apps.exams.models import ExamResult
            
            results = ExamResult.objects.filter(
                student=student,
                is_published=True
            ).select_related('exam').order_by('-exam__start_date')[:10]
            
            # Calculate statistics
            if results:
                total_results = results.count()
                passed_results = results.filter(percentage__gte=40).count()
                avg_percentage = results.aggregate(
                    avg=Avg('percentage')
                )['avg'] or 0
                
                return {
                    'exam_results': results,
                    'performance_stats': {
                        'total_exams': total_results,
                        'passed_exams': passed_results,
                        'pass_percentage': (passed_results / total_results * 100) if total_results > 0 else 0,
                        'avg_percentage': round(avg_percentage, 2),
                        'best_percentage': results.aggregate(
                            best=Max('percentage')
                        )['best'] or 0,
                        'worst_percentage': results.aggregate(
                            worst=Min('percentage')
                        )['worst'] or 0,
                    }
                }
            
        except Exception as e:
            logger.error(f"Performance data error: {str(e)}")
        
        return {
            'exam_results': [],
            'performance_stats': {}
        }
    
    def get_fee_information(self, student):
        """Get student fee information"""
        try:
            from apps.finance.models import FeeStructure, Payment, FeeInvoice
            
            # Get current fee structure
            fee_structure = FeeStructure.objects.filter(
                academic_year=student.academic_year,
                class_name=student.current_class,
                is_active=True
            ).first()
            
            # Get fee payments
            fee_payments = Payment.objects.filter(
                student=student,
                academic_year=student.academic_year
            ).order_by('-payment_date')
            
            # Get pending invoices
            pending_invoices = FeeInvoice.objects.filter(
                student=student,
                status='PENDING'
            ).order_by('due_date')
            
            # Calculate totals
            total_paid = fee_payments.aggregate(
                total=Sum('amount_paid')
            )['total'] or 0
            
            total_due = pending_invoices.aggregate(
                total=Sum('total_amount')
            )['total'] or 0
            
            return {
                'fee_structure': fee_structure,
                'fee_payments': fee_payments[:10],
                'pending_invoices': pending_invoices[:5],
                'fee_stats': {
                    'total_paid': total_paid,
                    'total_due': total_due,
                    'payment_count': fee_payments.count(),
                    'pending_count': pending_invoices.count(),
                }
            }
            
        except Exception as e:
            logger.error(f"Fee information error: {str(e)}")
            
        return {
            'fee_structure': None,
            'fee_payments': [],
            'pending_invoices': [],
            'fee_stats': {}
        }
    
    def get_attendance_data(self, student):
        """Get student attendance data"""
        try:
            from apps.attendance.models import Attendance
            
            # Get current month attendance
            today = timezone.now()
            first_day = today.replace(day=1)
            
            attendance_records = Attendance.objects.filter(
                student=student,
                date__gte=first_day,
                date__lte=today
            ).order_by('-date')
            
            # Calculate statistics
            total_days = attendance_records.count()
            present_days = attendance_records.filter(status='PRESENT').count()
            absent_days = attendance_records.filter(status='ABSENT').count()
            late_days = attendance_records.filter(status='LATE').count()
            
            attendance_rate = (present_days / total_days * 100) if total_days > 0 else 0
            
            return {
                'attendance_records': attendance_records[:20],
                'attendance_stats': {
                    'total_days': total_days,
                    'present_days': present_days,
                    'absent_days': absent_days,
                    'late_days': late_days,
                    'attendance_rate': round(attendance_rate, 2),
                    'current_month': today.strftime('%B %Y'),
                }
            }
            
        except Exception as e:
            logger.error(f"Attendance data error: {str(e)}")
            
        return {
            'attendance_records': [],
            'attendance_stats': {}
        }


class StudentCreateView(StudentPermissionMixin, BaseCreateView):
    """
    Secure student creation with comprehensive validation,
    audit logging, and notification.
    """
    form_class = StudentForm
    model = Student
    template_name = 'students/student_form.html'
    permission_required = 'students.add_student'
    success_url = reverse_lazy('students:student_list')
    
    @method_decorator(csrf_protect)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({
            'tenant': self.request.tenant,
            'user': self.request.user,
            'request': self.request
        })
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('Add New Student'),
            'is_create': True,
            'gender_choices': StudentGender.choices,
            'status_choices': StudentStatus.choices,
            'category_choices': StudentCategory.choices,
            'classes': SchoolClass.objects.filter(tenant=self.request.tenant),
            'academic_years': AcademicYear.objects.filter(tenant=self.request.tenant),
        })
        return context
    
    def form_valid(self, form):
        try:
            # Save student instance
            student = form.save(commit=False)
            
            # Set tenant
            student.tenant = self.request.tenant
            
            # Generate admission number if not provided
            if not student.admission_number:
                student.admission_number = StudentService.generate_admission_number(
                    self.request.tenant,
                    student.academic_year
                )
            
            # Save student
            student.save()
            form.save_m2m()
            
            # Create audit log
            AuditService.log_creation(
                user=self.request.user,
                instance=student,
                request=self.request,
                extra_data={
                    'created_via': 'web_form',
                    'admission_number': student.admission_number,
                    'class': str(student.current_class) if student.current_class else None,
                }
            )
            
            # Send notification
            NotificationService.send_student_registration_notification(
                student=student,
                created_by=self.request.user
            )
            
            # Send welcome email (async)
            send_student_welcome_email.delay(student.id)
            
            messages.success(
                self.request,
                _("Student %(name)s created successfully! Admission Number: %(adm_no)s") % {
                    'name': student.full_name,
                    'adm_no': student.admission_number
                }
            )
            
            # Redirect to next step or detail view
            if self.request.POST.get('next_step') == 'guardian':
                return redirect('students:guardian_create', student_id=student.id)
            
            return redirect(self.get_success_url())
            
        except Exception as e:
            logger.error(f"Student creation error: {str(e)}", exc_info=True)
            messages.error(
                self.request,
                _("Error creating student: %(error)s") % {'error': str(e)}
            )
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        """Handle invalid form submission"""
        messages.error(
            self.request,
            _("Please correct the errors below.")
        )
        return super().form_invalid(form)


class StudentUpdateView(StudentPermissionMixin, BaseUpdateView):
    """
    Secure student update with change tracking,
    audit logging, and permission validation.
    """
    form_class = StudentForm
    template_name = 'students/student_form.html'
    permission_required = 'students.change_student'
    
    def get_object(self, queryset=None):
        student = get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs.get('pk')
        )
        
        # Check object-level permission
        if not self.request.user.has_perm('students.change_student', student):
            raise PermissionDenied(_("You don't have permission to edit this student."))
        
        return student
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({
            'tenant': self.request.tenant,
            'user': self.request.user,
            'request': self.request,
            'instance': self.get_object()
        })
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('Edit Student'),
            'is_create': False,
            'student': self.get_object(),
        })
        return context
    
    def form_valid(self, form):
        try:
            # Get old instance for change tracking
            old_instance = Student.objects.get(id=form.instance.id)
            
            # Save changes
            response = super().form_valid(form)
            
            # Track changes
            changes = self.track_changes(old_instance, form.instance)
            
            # Create audit log with changes
            AuditService.log_update(
                user=self.request.user,
                instance=form.instance,
                old_instance=old_instance,
                request=self.request,
                extra_data={
                    'updated_via': 'web_form',
                    'changes': changes,
                    'update_reason': self.request.POST.get('update_reason', '')
                }
            )
            
            # Send update notification if significant changes
            if changes:
                NotificationService.send_student_update_notification(
                    student=form.instance,
                    updated_by=self.request.user,
                    changes=changes
                )
            
            messages.success(
                self.request,
                _("Student %(name)s updated successfully!") % {
                    'name': form.instance.full_name
                }
            )
            
            return response
            
        except Exception as e:
            logger.error(f"Student update error: {str(e)}", exc_info=True)
            messages.error(
                self.request,
                _("Error updating student: %(error)s") % {'error': str(e)}
            )
            return self.form_invalid(form)
    
    def track_changes(self, old_instance, new_instance):
        """Track changes between old and new instances"""
        changes = {}
        
        # Compare fields
        fields_to_track = [
            'first_name', 'last_name', 'date_of_birth', 'gender',
            'personal_email', 'mobile_primary', 'status',
            'current_class_id', 'section_id', 'academic_year_id',
            'category', 'blood_group', 'nationality'
        ]
        
        for field in fields_to_track:
            old_value = getattr(old_instance, field, None)
            new_value = getattr(new_instance, field, None)
            
            if old_value != new_value:
                changes[field] = {
                    'old': str(old_value) if old_value is not None else None,
                    'new': str(new_value) if new_value is not None else None
                }
        
        return changes
    
    def get_success_url(self):
        return reverse_lazy('students:student_detail', kwargs={'pk': self.object.id})


class StudentDeleteView(StudentPermissionMixin, BaseDeleteView):
    """
    Secure student deletion with soft delete,
    audit logging, and confirmation.
    """
    template_name = 'students/student_confirm_delete.html'
    permission_required = 'students.delete_student'
    
    def get_object(self, queryset=None):
        student = get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs.get('pk')
        )
        
        # Check object-level permission
        if not self.request.user.has_perm('students.delete_student', student):
            raise PermissionDenied(_("You don't have permission to delete this student."))
        
        return student
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        # Check if student can be deleted
        context['can_delete'] = self.can_delete_student(student)
        context['dependencies'] = self.get_dependencies(student)
        context['student'] = student
        
        return context
    
    def can_delete_student(self, student):
        """Check if student can be deleted"""
        # Check if student has fee payments
        from apps.finance.models import Payment
        has_payments = Payment.objects.filter(student=student).exists()
        
        # Check if student has exam results
        try:
            from apps.exams.models import ExamResult
            has_results = ExamResult.objects.filter(student=student).exists()
        except ImportError:
            has_results = False
        
        # Check if student has attendance records
        try:
            from apps.attendance.models import Attendance
            has_attendance = Attendance.objects.filter(student=student).exists()
        except ImportError:
            has_attendance = False
        
        return not (has_payments or has_results or has_attendance)
    
    def get_dependencies(self, student):
        """Get dependencies that prevent deletion"""
        dependencies = []
        
        try:
            from apps.finance.models import Payment
            if Payment.objects.filter(student=student).exists():
                dependencies.append(_("Fee payments"))
        except ImportError:
            pass
        
        try:
            from apps.exams.models import ExamResult
            if ExamResult.objects.filter(student=student).exists():
                dependencies.append(_("Exam results"))
        except ImportError:
            pass
        
        try:
            from apps.attendance.models import Attendance
            if Attendance.objects.filter(student=student).exists():
                dependencies.append(_("Attendance records"))
        except ImportError:
            pass
        
        return dependencies
    
    def delete(self, request, *args, **kwargs):
        """Override delete to use soft delete with audit"""
        student = self.get_object()
        
        try:
            # Get deletion reason
            deletion_reason = request.POST.get('deletion_reason', '').strip()
            if not deletion_reason:
                messages.error(request, _("Deletion reason is required."))
                return self.render_to_response(self.get_context_data())
            
            # Perform soft delete
            student.soft_delete(
                deleted_by=request.user,
                reason=deletion_reason,
                category=request.POST.get('deletion_category', 'ADMIN_ACTION')
            )
            
            # Log the deletion
            AuditService.log_deletion(
                user=request.user,
                instance=student,
                request=request,
                hard_delete=False,
                extra_data={
                    'deletion_reason': deletion_reason,
                    'deletion_category': request.POST.get('deletion_category'),
                    'deleted_by_user': str(request.user)
                }
            )
            
            # Send notification
            NotificationService.send_student_deletion_notification(
                student=student,
                deleted_by=request.user,
                reason=deletion_reason
            )
            
            messages.success(
                request,
                _("Student %(name)s has been deleted successfully.") % {
                    'name': student.full_name
                }
            )
            
        except ValidationError as e:
            logger.error(f"Student deletion validation error: {str(e)}")
            messages.error(request, _("Deletion failed: %(error)s") % {'error': str(e)})
            return self.render_to_response(self.get_context_data())
        except Exception as e:
            logger.error(f"Student deletion error: {str(e)}", exc_info=True)
            messages.error(request, _("Deletion failed: %(error)s") % {'error': str(e)})
            return self.render_to_response(self.get_context_data())
        
        return redirect('students:student_list')


# ============================================================================
# GUARDIAN MANAGEMENT VIEWS
# ============================================================================

class GuardianCreateView(StudentPermissionMixin, BaseCreateView):
    """Create guardian for student with comprehensive validation"""
    form_class = GuardianForm
    template_name = 'students/guardian_form.html'
    permission_required = 'students.add_student'
    
    def get_student(self):
        """Get student with permission check"""
        return get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs['student_id']
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = self.get_student()
        context['title'] = _('Add Guardian')
        context['is_create'] = True
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({
            'student': self.get_student(),
            'tenant': self.request.tenant,
            'user': self.request.user
        })
        return kwargs
    
    def form_valid(self, form):
        try:
            # Save guardian
            guardian = form.save(commit=False)
            guardian.tenant = self.request.tenant
            guardian.save()
            
            # Add to student if not already linked
            student = self.get_student()
            if guardian not in student.guardians.all():
                student.guardians.add(guardian)
            
            # Create audit log
            AuditService.log_creation(
                user=self.request.user,
                instance=guardian,
                request=self.request,
                extra_data={
                    'student_id': str(student.id),
                    'student_name': student.full_name,
                    'relationship': guardian.relationship
                }
            )
            
            # Send notification
            NotificationService.send_guardian_added_notification(
                guardian=guardian,
                student=student,
                added_by=self.request.user
            )
            
            messages.success(
                self.request,
                _("Guardian %(name)s added successfully!") % {
                    'name': guardian.full_name
                }
            )
            
            return redirect(self.get_success_url())
            
        except Exception as e:
            logger.error(f"Guardian creation error: {str(e)}", exc_info=True)
            messages.error(
                self.request,
                _("Error adding guardian: %(error)s") % {'error': str(e)}
            )
            return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('students:student_detail', kwargs={'pk': self.kwargs['student_id']})


class GuardianUpdateView(StudentPermissionMixin, BaseUpdateView):
    """Update guardian information"""
    form_class = GuardianForm
    template_name = 'students/guardian_form.html'
    permission_required = 'students.change_student'
    
    def get_object(self, queryset=None):
        return get_object_or_404(
            Guardian.objects.filter(
                tenant=self.request.tenant,
                student_id=self.kwargs['student_id']
            ),
            id=self.kwargs['guardian_id']
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs['student_id']
        )
        context['title'] = _('Edit Guardian')
        context['is_update'] = True
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({
            'tenant': self.request.tenant,
            'user': self.request.user,
            'student': get_object_or_404(
                StudentService.get_secure_queryset(self.request.user, self.request.tenant),
                id=self.kwargs['student_id']
            )
        })
        return kwargs
    
    def form_valid(self, form):
        try:
            # Track changes
            old_instance = Guardian.objects.get(id=form.instance.id)
            
            # Save changes
            response = super().form_valid(form)
            
            # Create audit log
            AuditService.log_update(
                user=self.request.user,
                instance=form.instance,
                old_instance=old_instance,
                request=self.request,
                extra_data={
                    'student_id': self.kwargs['student_id'],
                    'updated_via': 'web_form'
                }
            )
            
            messages.success(
                self.request,
                _("Guardian %(name)s updated successfully!") % {
                    'name': form.instance.full_name
                }
            )
            
            return response
            
        except Exception as e:
            logger.error(f"Guardian update error: {str(e)}", exc_info=True)
            messages.error(
                self.request,
                _("Error updating guardian: %(error)s") % {'error': str(e)}
            )
            return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('students:student_detail', kwargs={'pk': self.kwargs['student_id']})


class GuardianDeleteView(StudentPermissionMixin, BaseDeleteView):
    """Delete guardian from student"""
    template_name = 'students/guardian_confirm_delete.html'
    permission_required = 'students.delete_student'
    
    def get_object(self, queryset=None):
        guardian = get_object_or_404(
            Guardian.objects.filter(
                tenant=self.request.tenant,
                student_id=self.kwargs['student_id']
            ),
            id=self.kwargs['guardian_id']
        )
        
        # Check if guardian is primary
        if guardian.is_primary:
            raise PermissionDenied(_("Cannot delete primary guardian."))
        
        return guardian
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs['student_id']
        )
        return context
    
    def delete(self, request, *args, **kwargs):
        guardian = self.get_object()
        student_id = self.kwargs['student_id']
        
        try:
            # Remove guardian from student
            student = get_object_or_404(
                StudentService.get_secure_queryset(request.user, request.tenant),
                id=student_id
            )
            student.guardians.remove(guardian)
            
            # Create audit log
            AuditService.log_deletion(
                user=request.user,
                instance=guardian,
                request=request,
                hard_delete=False,
                extra_data={
                    'student_id': str(student_id),
                    'action': 'removed_from_student'
                }
            )
            
            messages.success(
                request,
                _("Guardian %(name)s removed successfully!") % {
                    'name': guardian.full_name
                }
            )
            
        except Exception as e:
            logger.error(f"Guardian removal error: {str(e)}", exc_info=True)
            messages.error(request, _("Error removing guardian: %(error)s") % {'error': str(e)})
        
        return redirect('students:student_detail', pk=student_id)


# ============================================================================
# DOCUMENT MANAGEMENT VIEWS
# ============================================================================

class DocumentUploadView(StudentPermissionMixin, BaseCreateView):
    """Secure document upload with validation and audit"""
    form_class = StudentDocumentForm
    template_name = 'students/document_form.html'
    permission_required = 'students.add_student'
    
    def get_student(self):
        """Get student with permission check"""
        return get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs['student_id']
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = self.get_student()
        context['title'] = _('Upload Document')
        context['allowed_types'] = DocumentService.get_allowed_document_types()
        context['max_size_mb'] = DocumentService.get_max_file_size_mb()
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({
            'student': self.get_student(),
            'tenant': self.request.tenant,
            'user': self.request.user
        })
        return kwargs
    
    def form_valid(self, form):
        try:
            # Save document
            document = form.save(commit=False)
            document.tenant = self.request.tenant
            
            # Set uploaded by
            document.uploaded_by = self.request.user
            
            # Save file
            document.save()
            
            # Create audit log
            AuditService.log_creation(
                user=self.request.user,
                instance=document,
                request=self.request,
                extra_data={
                    'student_id': str(self.get_student().id),
                    'document_type': document.doc_type,
                    'file_name': document.file_name,
                    'file_size': document.file.size if document.file else 0
                }
            )
            
            # Send notification if required document
            if document.is_required:
                NotificationService.send_document_uploaded_notification(
                    document=document,
                    uploaded_by=self.request.user
                )
            
            messages.success(
                self.request,
                _("Document %(name)s uploaded successfully!") % {
                    'name': document.file_name
                }
            )
            
            return redirect(self.get_success_url())
            
        except Exception as e:
            logger.error(f"Document upload error: {str(e)}", exc_info=True)
            messages.error(
                self.request,
                _("Error uploading document: %(error)s") % {'error': str(e)}
            )
            return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('students:student_detail', kwargs={'pk': self.kwargs['student_id']})


class DocumentDownloadView(StudentPermissionMixin, BaseView):
    """Secure document download with permission checking"""
    permission_required = 'students.view_student'
    
    @method_decorator(login_required)
    @method_decorator(require_https)
    def get(self, request, *args, **kwargs):
        document = self.get_document()
        student = document.student
        
        # Verify access permission
        if not StudentService.has_student_access(request.user, student):
            raise PermissionDenied(_("Access denied to this document."))
        
        try:
            if not document.file:
                raise Http404(_("File not found"))
            
            # Log the download
            AuditService.create_audit_entry(
                action='READ',
                resource_type='StudentDocument',
                user=request.user,
                tenant=request.tenant,
                request=request,
                instance=document,
                extra_data={
                    'student_id': str(student.id),
                    'action': 'document_download',
                    'file_name': document.file_name
                }
            )
            
            # Serve the file
            response = FileResponse(
                document.file,
                as_attachment=True,
                filename=document.file_name
            )
            
            # Set appropriate content type
            response['Content-Type'] = document.get_content_type()
            response['Content-Disposition'] = f'attachment; filename="{document.file_name}"'
            
            return response
            
        except Exception as e:
            logger.error(f"Document download error: {str(e)}", exc_info=True)
            raise Http404(_("Error downloading document"))
    
    def get_document(self):
        """Get document with permission check"""
        document = get_object_or_404(
            StudentDocument.objects.filter(tenant=self.request.tenant),
            id=self.kwargs.get('document_id')
        )
        
        # Verify student access
        if not StudentService.get_secure_queryset(
            self.request.user, self.request.tenant
        ).filter(id=document.student_id).exists():
            raise PermissionDenied(_("Access denied to this document."))
        
        return document


class DocumentDeleteView(StudentPermissionMixin, BaseDeleteView):
    """Delete student document"""
    template_name = 'students/document_confirm_delete.html'
    permission_required = 'students.delete_student'
    
    def get_object(self, queryset=None):
        document = get_object_or_404(
            StudentDocument.objects.filter(
                tenant=self.request.tenant,
                student_id=self.kwargs['student_id']
            ),
            id=self.kwargs['document_id']
        )
        
        # Check if document can be deleted
        if document.is_verified and not self.request.user.has_perm('students.delete_verified_document'):
            raise PermissionDenied(_("Cannot delete verified documents."))
        
        return document
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs['student_id']
        )
        return context
    
    def delete(self, request, *args, **kwargs):
        document = self.get_object()
        student_id = self.kwargs['student_id']
        
        try:
            # Store document info for audit
            document_info = {
                'file_name': document.file_name,
                'document_type': document.doc_type,
                'student_id': str(student_id)
            }
            
            # Delete document
            document.delete()
            
            # Create audit log
            AuditService.log_deletion(
                user=request.user,
                instance=document,
                request=request,
                hard_delete=True,
                extra_data=document_info
            )
            
            messages.success(
                request,
                _("Document %(name)s deleted successfully!") % {
                    'name': document_info['file_name']
                }
            )
            
        except Exception as e:
            logger.error(f"Document deletion error: {str(e)}", exc_info=True)
            messages.error(request, _("Error deleting document: %(error)s") % {'error': str(e)})
        
        return redirect('students:student_detail', pk=student_id)


# ============================================================================
# BULK OPERATIONS VIEWS
# ============================================================================

class StudentBulkUploadView(StudentPermissionMixin, RateLimitedViewMixin, BaseView):
    """
    Comprehensive bulk upload with validation,
    error handling, and progress tracking.
    """
    template_name = 'students/student_bulk_upload.html'
    permission_required = 'students.add_student'
    
    @method_decorator(csrf_protect)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = self.request.tenant
        
        context.update({
            'upload_form': StudentBulkUploadForm(tenant=tenant, user=self.request.user),
            'academic_years': AcademicYear.objects.filter(tenant=tenant).order_by('-is_current', '-start_date'),
            'classes': SchoolClass.objects.filter(tenant=tenant).order_by('order', 'name'),
            'streams': Stream.objects.filter(tenant=tenant).order_by('name'),
            'recent_uploads': self.get_recent_uploads(tenant),
            'max_file_size_mb': MAX_BULK_UPLOAD_SIZE // (1024 * 1024),
            'allowed_formats': ALLOWED_UPLOAD_FORMATS,
            'required_fields': REQUIRED_STUDENT_FIELDS,
        })
        
        return context
    
    def get_recent_uploads(self, tenant):
        """Get recent bulk upload history"""
        try:
            from apps.core.models import AuditLog
            
            return AuditLog.objects.filter(
                resource_type='Student',
                action='BULK_CREATE',
                tenant=tenant
            ).select_related('user').order_by('-timestamp')[:5]
            
        except Exception as e:
            logger.error(f"Error getting recent uploads: {str(e)}")
            return []
    
    def post(self, request, *args, **kwargs):
        """Handle file upload and processing"""
        form = StudentBulkUploadForm(
            request.POST,
            request.FILES,
            tenant=request.tenant,
            user=request.user
        )
        
        if not form.is_valid():
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            return self.render_to_response(self.get_context_data())
        
        try:
            # Process upload
            result = self.process_upload(form, request)
            
            # Handle result
            return self.handle_upload_result(request, result)
            
        except Exception as e:
            logger.error(f"Bulk upload error: {str(e)}", exc_info=True)
            messages.error(request, _("Error processing upload: %(error)s") % {'error': str(e)})
            return self.render_to_response(self.get_context_data())
    
    def process_upload(self, form, request):
        """Process uploaded file"""
        file = form.cleaned_data['file']
        academic_year = form.cleaned_data['academic_year']
        update_existing = form.cleaned_data['update_existing']
        skip_errors = form.cleaned_data['skip_errors']
        send_welcome_email = form.cleaned_data['send_welcome_email']
        
        # Process based on file type
        file_extension = file.name.split('.')[-1].lower()
        
        if file_extension == 'csv':
            return StudentImportService.process_csv(
                file=file,
                tenant=request.tenant,
                academic_year=academic_year,
                update_existing=update_existing,
                skip_errors=skip_errors,
                uploaded_by=request.user,
                send_welcome_email=send_welcome_email
            )
        elif file_extension in ['xls', 'xlsx']:
            return StudentImportService.process_excel(
                file=file,
                tenant=request.tenant,
                academic_year=academic_year,
                update_existing=update_existing,
                skip_errors=skip_errors,
                uploaded_by=request.user,
                send_welcome_email=send_welcome_email
            )
        else:
            raise ValidationError(_("Invalid file format. Please upload CSV or Excel file."))
    
    def handle_upload_result(self, request, result):
        """Handle upload result and display appropriate messages"""
        if 'error' in result:
            messages.error(request, result['error'])
            
            if 'errors' in result and result['errors']:
                for error in result['errors'][:5]:  # Show only first 5 errors
                    messages.warning(request, error)
                    
        elif 'success' in result:
            # Create success message
            success_msg = result['success']
            if result.get('warnings'):
                success_msg += f" ({len(result['warnings'])} warnings)"
            
            messages.success(request, success_msg)
            
            # Store details in session for summary page
            upload_summary = {
                'created': result.get('created_count', 0),
                'updated': result.get('updated_count', 0),
                'skipped': result.get('skipped_count', 0),
                'total': result.get('total_rows', 0),
                'warnings': result.get('warnings', [])[:20],
                'errors': result.get('errors', [])[:10],
                'timestamp': timezone.now().isoformat(),
                'file_name': request.FILES['file'].name if 'file' in request.FILES else '',
            }
            
            request.session['last_upload_summary'] = upload_summary
            
            # Create audit log
            AuditService.create_audit_entry(
                action='BULK_CREATE',
                resource_type='Student',
                user=request.user,
                tenant=request.tenant,
                request=request,
                severity='INFO',
                extra_data={
                    'summary': upload_summary,
                    'async_processing': result.get('async_processing', False),
                }
            )
            
            # Redirect based on warnings
            if result.get('warnings') or result.get('errors'):
                return redirect('students:bulk_upload_summary')
            else:
                return redirect('students:student_list')
        
        return self.render_to_response(self.get_context_data())


class BulkUploadSummaryView(StudentPermissionMixin, BaseView):
    """View to show detailed upload summary"""
    template_name = 'students/bulk_upload_summary.html'
    permission_required = 'students.add_student'
    
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        summary = request.session.get('last_upload_summary')
        
        if not summary:
            messages.warning(request, _("No upload summary found."))
            return redirect('students:student_bulk_upload')
        
        context = self.get_context_data(**kwargs)
        context['summary'] = summary
        context['has_warnings'] = bool(summary.get('warnings'))
        context['has_errors'] = bool(summary.get('errors'))
        context['success_rate'] = self.calculate_success_rate(summary)
        context['timestamp'] = datetime.fromisoformat(summary['timestamp']) if 'timestamp' in summary else None
        
        return self.render_to_response(context)
    
    def calculate_success_rate(self, summary):
        total = summary.get('total', 0)
        processed = summary.get('created', 0) + summary.get('updated', 0)
        return round((processed / total * 100) if total > 0 else 0, 2)


class BulkUploadSampleView(StudentPermissionMixin, BaseView):
    """Download sample templates"""
    permission_required = 'students.add_student'
    
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'csv')
        template_type = request.GET.get('type', 'simple')
        
        if format_type == 'excel':
            return self.generate_excel_sample(template_type)
        else:
            return self.generate_csv_sample(template_type)
    
    def generate_csv_sample(self, template_type):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'student_upload_{template_type}_template.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        if template_type == 'detailed':
            headers = [
                'first_name', 'last_name', 'date_of_birth', 'gender',
                'personal_email', 'mobile_primary', 'class_name', 'section_name',
                'status', 'category', 'admission_type', 'blood_group',
                'nationality', 'religion', 'caste', 'father_name',
                'mother_name', 'guardian_email', 'guardian_phone'
            ]
            writer.writerow(headers)
            
            # Sample data
            sample_data = [
                ['John', 'Doe', '2010-01-15', 'M', 
                 'john.doe@example.com', '9876543210', 'Class 1', 'A',
                 'ACTIVE', 'GENERAL', 'REGULAR', 'O+', 'Indian', 'Hindu', 'General',
                 'Robert Doe', 'Mary Doe', 'robert.doe@example.com', '9876543211'],
                ['Jane', 'Smith', '2010-05-20', 'F',
                 'jane.smith@example.com', '9876543211', 'Class 2', 'B',
                 'ACTIVE', 'OBC', 'REGULAR', 'A+', 'Indian', 'Christian', 'OBC',
                 'John Smith', 'Sarah Smith', 'john.smith@example.com', '9876543212'],
            ]
            
            for row in sample_data:
                writer.writerow(row)
        else:
            headers = ['first_name', 'last_name', 'personal_email', 'class_name']
            writer.writerow(headers)
            
            sample_data = [
                ['John', 'Doe', 'john.doe@example.com', 'Class 1'],
                ['Jane', 'Smith', 'jane.smith@example.com', 'Class 2'],
            ]
            
            for row in sample_data:
                writer.writerow(row)
        
        return response
    
    def generate_excel_sample(self, template_type):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Template"
        
        if template_type == 'detailed':
            headers = [
                'First Name', 'Last Name', 'Date of Birth (YYYY-MM-DD)', 'Gender (M/F/O)',
                'Email', 'Phone', 'Class Name', 'Section Name', 'Status (ACTIVE/INACTIVE)',
                'Category (GENERAL/OBC/SC/ST)', 'Admission Type', 'Blood Group',
                'Nationality', 'Religion', 'Caste', 'Father Name', 'Mother Name',
                'Guardian Email', 'Guardian Phone'
            ]
            
            # Add headers with formatting
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add sample data
            sample_data = [
                ['John', 'Doe', '2010-01-15', 'M', 
                 'john.doe@example.com', '9876543210', 'Class 1', 'A',
                 'ACTIVE', 'GENERAL', 'REGULAR', 'O+', 'Indian', 'Hindu', 'General',
                 'Robert Doe', 'Mary Doe', 'robert.doe@example.com', '9876543211'],
                ['Jane', 'Smith', '2010-05-20', 'F',
                 'jane.smith@example.com', '9876543211', 'Class 2', 'B',
                 'ACTIVE', 'OBC', 'REGULAR', 'A+', 'Indian', 'Christian', 'OBC',
                 'John Smith', 'Sarah Smith', 'john.smith@example.com', '9876543212'],
            ]
            
            for row_idx, row in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            headers = ['First Name', 'Last Name', 'Email', 'Class Name']
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            sample_data = [
                ['John', 'Doe', 'john.doe@example.com', 'Class 1'],
                ['Jane', 'Smith', 'jane.smith@example.com', 'Class 2'],
            ]
            
            for row_idx, row in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'student_upload_{template_type}_template.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class BulkUploadValidateView(StudentPermissionMixin, BaseView):
    """AJAX validation endpoint for bulk upload"""
    permission_required = 'students.add_student'
    
    @method_decorator(csrf_protect)
    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            file = request.FILES.get('file')
            if not file:
                return JsonResponse({
                    'valid': False,
                    'error': _('No file provided')
                }, status=400)
            
            # Validate file
            validation_result = StudentImportService.validate_upload_file(
                file=file,
                tenant=request.tenant,
                user=request.user
            )
            
            return JsonResponse(validation_result)
            
        except Exception as e:
            logger.error(f"Upload validation error: {str(e)}", exc_info=True)
            return JsonResponse({
                'valid': False,
                'error': _('Validation error: %(error)s') % {'error': str(e)}
            }, status=500)


# ============================================================================
# EXPORT VIEWS
# ============================================================================

class StudentExportView(StudentPermissionMixin, BaseView):
    """Export student data with comprehensive filtering and audit logging"""
    permission_required = 'students.export_student_data'
    
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            format_type = request.GET.get('format', 'csv')
            export_type = request.GET.get('type', 'basic')
            
            # Get filtered queryset
            queryset = self.get_filtered_queryset(request)
            
            # Validate export size
            if queryset.count() > 10000 and format_type == 'excel':
                return JsonResponse({
                    'error': _('Export size too large for Excel format. Please use CSV or reduce filters.')
                }, status=400)
            
            # Generate export
            if format_type == 'excel':
                return self.export_excel(queryset, export_type, request)
            elif format_type == 'pdf':
                return self.export_pdf(queryset, export_type, request)
            else:
                return self.export_csv(queryset, export_type, request)
                
        except Exception as e:
            logger.error(f"Export error: {str(e)}", exc_info=True)
            messages.error(request, _("Export failed: %(error)s") % {'error': str(e)})
            return redirect('students:student_list')
    
    def get_filtered_queryset(self, request):
        """Get filtered student queryset"""
        queryset = StudentService.get_secure_queryset(request.user, request.tenant)
        
        # Apply filters
        filters = Q()
        
        # Status filter
        status = request.GET.get('status')
        if status in dict(StudentStatus.choices):
            filters &= Q(status=status)
        
        # Class filter
        class_id = request.GET.get('class_id')
        if class_id and class_id.isdigit():
            filters &= Q(current_class_id=int(class_id))
        
        # Section filter
        section_id = request.GET.get('section_id')
        if section_id and section_id.isdigit():
            filters &= Q(section_id=int(section_id))
        
        # Academic year filter
        academic_year_id = request.GET.get('academic_year_id')
        if academic_year_id and academic_year_id.isdigit():
            filters &= Q(academic_year_id=int(academic_year_id))
        
        # Category filter
        category = request.GET.get('category')
        if category in dict(StudentCategory.choices):
            filters &= Q(category=category)
        
        # Gender filter
        gender = request.GET.get('gender')
        if gender in dict(StudentGender.choices):
            filters &= Q(gender=gender)
        
        # Date range filters
        created_from = request.GET.get('created_from')
        created_to = request.GET.get('created_to')
        
        if created_from:
            try:
                date_from = datetime.strptime(created_from, '%Y-%m-%d').date()
                filters &= Q(created_at__date__gte=date_from)
            except ValueError:
                pass
        
        if created_to:
            try:
                date_to = datetime.strptime(created_to, '%Y-%m-%d').date()
                filters &= Q(created_at__date__lte=date_to)
            except ValueError:
                pass
        
        return queryset.filter(filters).distinct()
    
    def export_csv(self, queryset, export_type, request):
        """Export data to CSV format"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f'students_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        if export_type == 'detailed':
            headers = [
                'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                'Email', 'Phone', 'Class', 'Section', 'Stream', 'Academic Year',
                'Status', 'Category', 'Blood Group', 'Nationality', 'Religion',
                'Caste', 'Admission Date', 'Created At', 'Updated At'
            ]
            writer.writerow(headers)
            
            for student in queryset.select_related(
                'current_class', 'section', 'stream', 'academic_year'
            ):
                writer.writerow([
                    student.admission_number,
                    student.full_name,
                    student.date_of_birth,
                    student.get_gender_display(),
                    student.personal_email,
                    student.mobile_primary,
                    student.current_class.name if student.current_class else '',
                    student.section.name if student.section else '',
                    student.stream.name if student.stream else '',
                    student.academic_year.name if student.academic_year else '',
                    student.get_status_display(),
                    student.get_category_display(),
                    student.blood_group or '',
                    student.nationality or '',
                    student.religion or '',
                    student.caste or '',
                    student.admission_date or '',
                    student.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    student.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                ])
        else:
            headers = [
                'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                'Email', 'Phone', 'Class', 'Section', 'Status', 'Category'
            ]
            writer.writerow(headers)
            
            for student in queryset.select_related('current_class', 'section'):
                writer.writerow([
                    student.admission_number,
                    student.full_name,
                    student.date_of_birth,
                    student.get_gender_display(),
                    student.personal_email,
                    student.mobile_primary,
                    student.current_class.name if student.current_class else '',
                    student.section.name if student.section else '',
                    student.get_status_display(),
                    student.get_category_display()
                ])
        
        # Log the export
        self.log_export(request, queryset, 'csv', export_type)
        
        return response
    
    def export_excel(self, queryset, export_type, request):
        """Export data to Excel format"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        if export_type == 'detailed':
            headers = [
                'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                'Email', 'Phone', 'Class', 'Section', 'Stream', 'Academic Year',
                'Status', 'Category', 'Blood Group', 'Nationality', 'Religion',
                'Caste', 'Admission Date', 'Created At', 'Updated At'
            ]
            
            # Add headers with formatting
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row_idx, student in enumerate(
                queryset.select_related('current_class', 'section', 'stream', 'academic_year'),
                start=2
            ):
                ws.cell(row=row_idx, column=1, value=student.admission_number)
                ws.cell(row=row_idx, column=2, value=student.full_name)
                ws.cell(row=row_idx, column=3, value=str(student.date_of_birth))
                ws.cell(row=row_idx, column=4, value=student.get_gender_display())
                ws.cell(row=row_idx, column=5, value=student.personal_email)
                ws.cell(row=row_idx, column=6, value=student.mobile_primary)
                ws.cell(row=row_idx, column=7, value=student.current_class.name if student.current_class else '')
                ws.cell(row=row_idx, column=8, value=student.section.name if student.section else '')
                ws.cell(row=row_idx, column=9, value=student.stream.name if student.stream else '')
                ws.cell(row=row_idx, column=10, value=student.academic_year.name if student.academic_year else '')
                ws.cell(row=row_idx, column=11, value=student.get_status_display())
                ws.cell(row=row_idx, column=12, value=student.get_category_display())
                ws.cell(row=row_idx, column=13, value=student.blood_group or '')
                ws.cell(row=row_idx, column=14, value=student.nationality or '')
                ws.cell(row=row_idx, column=15, value=student.religion or '')
                ws.cell(row=row_idx, column=16, value=student.caste or '')
                ws.cell(row=row_idx, column=17, value=str(student.admission_date) if student.admission_date else '')
                ws.cell(row=row_idx, column=18, value=student.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_idx, column=19, value=student.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        else:
            headers = [
                'Admission Number', 'Full Name', 'Date of Birth', 'Gender',
                'Email', 'Phone', 'Class', 'Section', 'Status', 'Category'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row_idx, student in enumerate(
                queryset.select_related('current_class', 'section'),
                start=2
            ):
                ws.cell(row=row_idx, column=1, value=student.admission_number)
                ws.cell(row=row_idx, column=2, value=student.full_name)
                ws.cell(row=row_idx, column=3, value=str(student.date_of_birth))
                ws.cell(row=row_idx, column=4, value=student.get_gender_display())
                ws.cell(row=row_idx, column=5, value=student.personal_email)
                ws.cell(row=row_idx, column=6, value=student.mobile_primary)
                ws.cell(row=row_idx, column=7, value=student.current_class.name if student.current_class else '')
                ws.cell(row=row_idx, column=8, value=student.section.name if student.section else '')
                ws.cell(row=row_idx, column=9, value=student.get_status_display())
                ws.cell(row=row_idx, column=10, value=student.get_category_display())
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'students_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        # Log the export
        self.log_export(request, queryset, 'excel', export_type)
        
        return response
    
    def export_pdf(self, queryset, export_type, request):
        """Export data to PDF format"""
        # This would integrate with a PDF generation library like ReportLab or WeasyPrint
        # For now, return a message to use CSV or Excel
        messages.info(request, _("PDF export is currently not available. Please use CSV or Excel format."))
        return redirect('students:student_list')
    
    def log_export(self, request, queryset, format_type, export_type):
        """Log export operation"""
        AuditService.create_audit_entry(
            action='EXPORT',
            resource_type='Student',
            user=request.user,
            tenant=request.tenant,
            request=request,
            severity='INFO',
            extra_data={
                'export_format': format_type,
                'export_type': export_type,
                'record_count': queryset.count(),
                'filters': dict(request.GET),
                'filename': f'students_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.{format_type}'
            }
        )


# ============================================================================
# SPECIALIZED VIEWS
# ============================================================================

class StudentAcademicHistoryView(StudentPermissionMixin, BaseView):
    """View student academic history with comprehensive data"""
    template_name = 'students/student_academic_history.html'
    permission_required = 'students.view_student'
    
    def get_object(self):
        return get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs.get('pk')
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        context['student'] = student
        context['academic_history'] = StudentAcademicHistory.objects.filter(
            student=student
        ).select_related(
            'academic_year', 'class_name', 'section'
        ).order_by('-academic_year__start_date')
        
        # Get exam results if available
        context['exam_results'] = self.get_exam_results(student)
        
        # Get attendance summary
        context['attendance_summary'] = self.get_attendance_summary(student)
        
        # Get fee history
        context['fee_history'] = self.get_fee_history(student)
        
        # Get awards and achievements
        context['achievements'] = self.get_achievements(student)
        
        # Log access
        AuditService.log_view(
            user=self.request.user,
            instance=student,
            request=self.request,
            extra_data={'view_type': 'academic_history'}
        )
        
        return context
    
    def get_exam_results(self, student):
        """Get student exam results"""
        try:
            from apps.exams.models import ExamResult
            
            return ExamResult.objects.filter(
                student=student
            ).select_related(
                'exam', 'exam__exam_type'
            ).order_by('-exam__start_date')[:20]
            
        except ImportError:
            return []
    
    def get_attendance_summary(self, student):
        """Get attendance summary for current academic year"""
        try:
            from apps.attendance.models import Attendance
            
            if not student.academic_year:
                return {}
            
            attendance = Attendance.objects.filter(
                student=student,
                academic_year=student.academic_year
            )
            
            return attendance.aggregate(
                total=models.Count('id'),
                present=models.Count('id', filter=models.Q(status='PRESENT')),
                absent=models.Count('id', filter=models.Q(status='ABSENT')),
                late=models.Count('id', filter=models.Q(status='LATE')),
                half_day=models.Count('id', filter=models.Q(status='HALF_DAY')),
            )
            
        except ImportError:
            return {}
    
    def get_fee_history(self, student):
        """Get student fee payment history"""
        try:
            from apps.finance.models import Payment
            
            return Payment.objects.filter(
                student=student
            ).select_related(
                'fee_structure', 'academic_year'
            ).order_by('-payment_date')[:20]
            
        except ImportError:
            return []
    
    def get_achievements(self, student):
        """Get student achievements and awards"""
        try:
            from apps.achievements.models import StudentAchievement
            
            return StudentAchievement.objects.filter(
                student=student,
                is_verified=True
            ).order_by('-achievement_date')[:10]
            
        except ImportError:
            return []


class StudentPromoteView(StudentPermissionMixin, BaseView):
    """Promote student to next class with comprehensive validation"""
    template_name = 'students/student_promote.html'
    permission_required = 'students.change_student'
    
    def get_object(self):
        return get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs.get('pk')
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        context['student'] = student
        context['promotion_form'] = StudentPromotionForm(
            tenant=self.request.tenant,
            student=student,
            user=self.request.user
        )
        
        # Get next available classes
        if student.current_class:
            context['next_classes'] = SchoolClass.objects.filter(
                tenant=self.request.tenant,
                order__gt=student.current_class.order
            ).order_by('order')
        else:
            context['next_classes'] = SchoolClass.objects.filter(
                tenant=self.request.tenant
            ).order_by('order')
        
        # Get next academic year
        if student.academic_year:
            context['next_academic_year'] = AcademicYear.objects.filter(
                tenant=self.request.tenant,
                start_date__gt=student.academic_year.start_date
            ).order_by('start_date').first()
        
        # Check if student is eligible for promotion
        context['is_eligible'] = self.check_promotion_eligibility(student)
        
        return context
    
    def check_promotion_eligibility(self, student):
        """Check if student is eligible for promotion"""
        # Check attendance requirement
        attendance_eligible = self.check_attendance_eligibility(student)
        
        # Check academic performance requirement
        performance_eligible = self.check_performance_eligibility(student)
        
        # Check fee payment requirement
        fee_eligible = self.check_fee_eligibility(student)
        
        return all([attendance_eligible, performance_eligible, fee_eligible])
    
    def check_attendance_eligibility(self, student):
        """Check attendance eligibility for promotion"""
        try:
            from apps.attendance.models import Attendance
            
            if not student.academic_year:
                return True  # Can't check without academic year
            
            attendance = Attendance.objects.filter(
                student=student,
                academic_year=student.academic_year
            )
            
            total_days = attendance.count()
            present_days = attendance.filter(status='PRESENT').count()
            
            if total_days == 0:
                return True  # No attendance records
            
            attendance_percentage = (present_days / total_days) * 100
            return attendance_percentage >= 75  # Minimum 75% attendance required
            
        except ImportError:
            return True  # Attendance module not available
    
    def check_performance_eligibility(self, student):
        """Check academic performance eligibility"""
        try:
            from apps.exams.models import ExamResult
            
            results = ExamResult.objects.filter(
                student=student,
                academic_year=student.academic_year,
                is_published=True
            )
            
            if not results.exists():
                return True  # No exam results
            
            # Check if student has passed all subjects
            for result in results:
                if result.percentage < 40:  # Below passing percentage
                    return False
            
            return True
            
        except ImportError:
            return True  # Exams module not available
    
    def check_fee_eligibility(self, student):
        """Check fee payment eligibility"""
        try:
            from apps.finance.models import FeeInvoice
            
            pending_invoices = FeeInvoice.objects.filter(
                student=student,
                academic_year=student.academic_year,
                status='PENDING',
                due_date__lt=timezone.now()
            )
            
            return not pending_invoices.exists()
            
        except ImportError:
            return True  # Finance module not available
    
    def post(self, request, *args, **kwargs):
        student = self.get_object()
        form = StudentPromotionForm(
            request.POST,
            tenant=request.tenant,
            student=student,
            user=request.user
        )
        
        if not form.is_valid():
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            return redirect('students:student_promote', pk=student.id)
        
        try:
            # Process promotion
            new_class = form.cleaned_data['new_class']
            new_academic_year = form.cleaned_data['new_academic_year']
            promotion_date = form.cleaned_data['promotion_date']
            remarks = form.cleaned_data['remarks']
            
            # Store old data for audit
            old_class = student.current_class
            old_academic_year = student.academic_year
            
            # Update student
            student.current_class = new_class
            student.academic_year = new_academic_year
            student.save()
            
            # Create academic history record
            StudentAcademicHistory.objects.create(
                student=student,
                academic_year=old_academic_year,
                class_name=old_class,
                section=student.section,
                promoted_to=new_class,
                promotion_date=promotion_date,
                remarks=remarks,
                promoted_by=request.user
            )
            
            # Create audit log
            AuditService.log_update(
                user=request.user,
                instance=student,
                request=request,
                extra_data={
                    'action': 'promotion',
                    'old_class': str(old_class),
                    'new_class': str(new_class),
                    'old_academic_year': str(old_academic_year),
                    'new_academic_year': str(new_academic_year),
                    'promotion_date': str(promotion_date),
                    'remarks': remarks
                }
            )
            
            # Send notification
            NotificationService.send_student_promotion_notification(
                student=student,
                old_class=old_class,
                new_class=new_class,
                promoted_by=request.user
            )
            
            messages.success(
                request,
                _("Student %(name)s promoted from %(old_class)s to %(new_class)s successfully!") % {
                    'name': student.full_name,
                    'old_class': old_class.name if old_class else '',
                    'new_class': new_class.name
                }
            )
            
        except Exception as e:
            logger.error(f"Promotion error: {str(e)}", exc_info=True)
            messages.error(
                request,
                _("Promotion failed: %(error)s") % {'error': str(e)}
            )
        
        return redirect('students:student_detail', pk=student.id)


class StudentReportView(StudentPermissionMixin, BaseView):
    """Generate comprehensive student report"""
    template_name = 'students/student_report.html'
    permission_required = 'students.view_student'
    
    def get_object(self):
        return get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs.get('pk')
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.get_object()
        
        context['student'] = student
        context['report_date'] = timezone.now()
        
        # Get report data
        context['academic_data'] = self.get_academic_data(student)
        context['attendance_data'] = self.get_attendance_data(student)
        context['behavior_data'] = self.get_behavior_data(student)
        context['parent_feedback'] = self.get_parent_feedback(student)
        context['teacher_comments'] = self.get_teacher_comments(student)
        
        # Log report generation
        AuditService.create_audit_entry(
            action='READ',
            resource_type='StudentReport',
            user=self.request.user,
            tenant=self.request.tenant,
            request=self.request,
            instance=student,
            extra_data={'report_type': 'comprehensive'}
        )
        
        return context
    
    def get_academic_data(self, student):
        """Get academic performance data"""
        try:
            from apps.exams.models import ExamResult, ExamSubjectResult
            
            # Get all exam results
            exam_results = ExamResult.objects.filter(
                student=student
            ).select_related('exam').order_by('exam__start_date')
            
            # Calculate subject-wise performance
            subject_performance = {}
            for result in exam_results:
                subject_results = ExamSubjectResult.objects.filter(
                    exam_result=result
                ).select_related('subject')
                
                for subj_result in subject_results:
                    subject_name = subj_result.subject.name
                    if subject_name not in subject_performance:
                        subject_performance[subject_name] = {
                            'total_marks': 0,
                            'obtained_marks': 0,
                            'count': 0
                        }
                    
                    subject_performance[subject_name]['total_marks'] += subj_result.total_marks
                    subject_performance[subject_name]['obtained_marks'] += subj_result.marks_obtained
                    subject_performance[subject_name]['count'] += 1
            
            # Calculate percentages
            for subject, data in subject_performance.items():
                if data['total_marks'] > 0:
                    data['percentage'] = round((data['obtained_marks'] / data['total_marks']) * 100, 2)
                else:
                    data['percentage'] = 0
            
            return {
                'exam_results': exam_results,
                'subject_performance': subject_performance,
                'overall_average': self.calculate_overall_average(exam_results)
            }
            
        except ImportError:
            return {}
    
    def calculate_overall_average(self, exam_results):
        """Calculate overall average percentage"""
        if not exam_results:
            return 0
        
        total_percentage = sum(result.percentage for result in exam_results)
        return round(total_percentage / exam_results.count(), 2)
    
    def get_attendance_data(self, student):
        """Get attendance data"""
        try:
            from apps.attendance.models import Attendance
            
            attendance = Attendance.objects.filter(
                student=student,
                academic_year=student.academic_year
            )
            
            monthly_data = {}
            for record in attendance:
                month_key = record.date.strftime('%Y-%m')
                if month_key not in monthly_data:
                    monthly_data[month_key] = {
                        'total': 0,
                        'present': 0,
                        'absent': 0,
                        'late': 0
                    }
                
                monthly_data[month_key]['total'] += 1
                
                if record.status == 'PRESENT':
                    monthly_data[month_key]['present'] += 1
                elif record.status == 'ABSENT':
                    monthly_data[month_key]['absent'] += 1
                elif record.status == 'LATE':
                    monthly_data[month_key]['late'] += 1
            
            # Calculate percentages
            for month, data in monthly_data.items():
                if data['total'] > 0:
                    data['attendance_percentage'] = round((data['present'] / data['total']) * 100, 2)
                else:
                    data['attendance_percentage'] = 0
            
            return {
                'monthly_data': monthly_data,
                'overall_attendance': self.calculate_overall_attendance(attendance)
            }
            
        except ImportError:
            return {}
    
    def calculate_overall_attendance(self, attendance):
        """Calculate overall attendance percentage"""
        if not attendance:
            return 0
        
        total_days = attendance.count()
        present_days = attendance.filter(status='PRESENT').count()
        
        return round((present_days / total_days) * 100, 2) if total_days > 0 else 0
    
    def get_behavior_data(self, student):
        """Get behavior and discipline data"""
        try:
            from apps.discipline.models import DisciplineRecord
            
            records = DisciplineRecord.objects.filter(
                student=student,
                academic_year=student.academic_year
            ).order_by('-incident_date')
            
            return {
                'records': records,
                'total_incidents': records.count(),
                'serious_incidents': records.filter(severity__in=['HIGH', 'CRITICAL']).count(),
                'resolved_incidents': records.filter(status='RESOLVED').count()
            }
            
        except ImportError:
            return {}
    
    def get_parent_feedback(self, student):
        """Get parent feedback"""
        try:
            from apps.feedback.models import ParentFeedback
            
            return ParentFeedback.objects.filter(
                student=student
            ).order_by('-feedback_date')[:5]
            
        except ImportError:
            return []
    
    def get_teacher_comments(self, student):
        """Get teacher comments"""
        try:
            from apps.feedback.models import TeacherComment
            
            return TeacherComment.objects.filter(
                student=student
            ).select_related('teacher').order_by('-comment_date')[:10]
            
        except ImportError:
            return []


class StudentIdCardView(StudentPermissionMixin, BaseDetailView):
    """Generate and download student ID card"""
    permission_required = 'students.view_student'
    
    def get_object(self):
        return get_object_or_404(
            StudentService.get_secure_queryset(self.request.user, self.request.tenant),
            id=self.kwargs.get('pk')
        )
    
    def get(self, request, *args, **kwargs):
        """Generate ID card on the fly"""
        student = self.get_object()
        
        try:
            # Generate ID card
            generator = StudentIDCardGenerator(student)
            response = generator.get_id_card_response()
            
            # Log ID card generation
            AuditService.create_audit_entry(
                action='READ',
                resource_type='StudentIDCard',
                user=request.user,
                tenant=request.tenant,
                request=request,
                instance=student,
                extra_data={'action': 'generated_id_card'}
            )
            
            return response
            
        except Exception as e:
            logger.error(f"ID card generation error: {str(e)}", exc_info=True)
            messages.error(request, _("Error generating ID card: %(error)s") % {'error': str(e)})
            return redirect('students:student_detail', pk=student.id)


class BatchIdCardView(StudentPermissionMixin, BaseView):
    """Generate ID cards for multiple students"""
    template_name = 'students/batch_id_cards.html'
    permission_required = 'students.view_student'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = self.request.tenant
        
        context.update({
            'classes': SchoolClass.objects.filter(tenant=tenant).order_by('order', 'name'),
            'sections': Section.objects.filter(tenant=tenant).order_by('name'),
            'academic_years': AcademicYear.objects.filter(tenant=tenant).order_by('-is_current', '-start_date'),
        })
        
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            # Get filter parameters
            class_id = request.POST.get('class_id')
            section_id = request.POST.get('section_id')
            academic_year_id = request.POST.get('academic_year_id')
            student_ids = request.POST.getlist('student_ids')
            
            # Get students based on filters
            queryset = StudentService.get_secure_queryset(request.user, request.tenant)
            
            if class_id and class_id.isdigit():
                queryset = queryset.filter(current_class_id=int(class_id))
            
            if section_id and section_id.isdigit():
                queryset = queryset.filter(section_id=int(section_id))
            
            if academic_year_id and academic_year_id.isdigit():
                queryset = queryset.filter(academic_year_id=int(academic_year_id))
            
            if student_ids:
                queryset = queryset.filter(id__in=[int(sid) for sid in student_ids if sid.isdigit()])
            
            # Limit to reasonable number
            student_count = queryset.count()
            if student_count > 100:
                messages.warning(request, _("Too many students selected. Maximum 100 allowed."))
                return redirect('students:batch_id_cards')
            
            if student_count == 0:
                messages.warning(request, _("No students found matching criteria."))
                return redirect('students:batch_id_cards')
            
            # Generate ID cards (async if many)
            if student_count > 10:
                # Start async task
                task_id = generate_id_cards_batch.delay(
                    student_ids=list(queryset.values_list('id', flat=True)),
                    user_id=request.user.id,
                    tenant_id=request.tenant.id
                )
                
                messages.info(
                    request,
                    _("ID card generation started for %(count)s students. Task ID: %(task_id)s") % {
                        'count': student_count,
                        'task_id': task_id
                    }
                )
                
                # Store task ID in session
                request.session['id_card_task_id'] = str(task_id)
                
                return redirect('students:batch_id_cards_status')
            
            else:
                # Generate synchronously
                generator = StudentIDCardGenerator(None)
                response = generator.generate_batch_id_cards(queryset)
                
                # Log batch generation
                AuditService.create_audit_entry(
                    action='BULK_OPERATION',
                    resource_type='StudentIDCard',
                    user=request.user,
                    tenant=request.tenant,
                    request=request,
                    severity='INFO',
                    extra_data={
                        'action': 'batch_id_card_generation',
                        'student_count': student_count,
                        'filters': {
                            'class_id': class_id,
                            'section_id': section_id,
                            'academic_year_id': academic_year_id,
                        }
                    }
                )
                
                return response
                
        except Exception as e:
            logger.error(f"Batch ID card error: {str(e)}", exc_info=True)
            messages.error(request, _("Error generating ID cards: %(error)s") % {'error': str(e)})
            return redirect('students:batch_id_cards')


# ============================================================================
# MULTI-STEP REGISTRATION VIEWS
# ============================================================================

class StudentRegistrationWizardView(StudentPermissionMixin, BaseView):
    """
    Multi-step student registration wizard
    Handles student, guardian, and document registration in sequence
    """
    template_name = 'students/registration/wizard.html'
    permission_required = 'students.add_student'
    
    def get(self, request, *args, **kwargs):
        step = request.GET.get('step', '1')
        
        if step == '1':
            return self.step1(request)
        elif step == '2':
            return self.step2(request)
        elif step == '3':
            return self.step3(request)
        elif step == '4':
            return self.step4(request)
        else:
            return redirect('students:registration_wizard', step='1')
    
    def step1(self, request):
        """Step 1: Student Information"""
        form = StudentForm(
            tenant=request.tenant,
            user=request.user,
            prefix='step1'
        )
        
        context = self.get_context_data()
        context.update({
            'step': 1,
            'form': form,
            'title': _('Student Information'),
            'next_step': '2'
        })
        
        return render(request, 'students/registration/step1_student.html', context)
    
    def step2(self, request):
        """Step 2: Guardian Information"""
        # Check if student data is in session
        student_data = request.session.get('registration_student_data')
        if not student_data:
            messages.warning(request, _("Please complete student information first."))
            return redirect('students:registration_wizard', step='1')
        
        form = GuardianForm(
            tenant=request.tenant,
            user=request.user,
            prefix='step2'
        )
        
        context = self.get_context_data()
        context.update({
            'step': 2,
            'form': form,
            'title': _('Guardian Information'),
            'next_step': '3',
            'prev_step': '1'
        })
        
        return render(request, 'students/registration/step2_guardian.html', context)
    
    def step3(self, request):
        """Step 3: Address Information"""
        # Check if previous steps are complete
        if not request.session.get('registration_student_data'):
            return redirect('students:registration_wizard', step='1')
        
        if not request.session.get('registration_guardian_data'):
            return redirect('students:registration_wizard', step='2')
        
        context = self.get_context_data()
        context.update({
            'step': 3,
            'title': _('Address Information'),
            'next_step': '4',
            'prev_step': '2'
        })
        
        return render(request, 'students/registration/step3_address.html', context)
    
    def step4(self, request):
        """Step 4: Document Upload"""
        # Check if previous steps are complete
        if not request.session.get('registration_student_data'):
            return redirect('students:registration_wizard', step='1')
        
        if not request.session.get('registration_guardian_data'):
            return redirect('students:registration_wizard', step='2')
        
        form = StudentDocumentForm(
            tenant=request.tenant,
            user=request.user,
            prefix='step4'
        )
        
        context = self.get_context_data()
        context.update({
            'step': 4,
            'form': form,
            'title': _('Document Upload'),
            'prev_step': '3'
        })
        
        return render(request, 'students/registration/step4_documents.html', context)
    
    def post(self, request, *args, **kwargs):
        step = request.POST.get('current_step', '1')
        
        if step == '1':
            return self.process_step1(request)
        elif step == '2':
            return self.process_step2(request)
        elif step == '3':
            return self.process_step3(request)
        elif step == '4':
            return self.process_step4(request)
        else:
            return redirect('students:registration_wizard', step='1')
    
    def process_step1(self, request):
        """Process step 1: Student Information"""
        form = StudentForm(
            request.POST,
            tenant=request.tenant,
            user=request.user,
            prefix='step1'
        )
        
        if form.is_valid():
            # Store student data in session
            request.session['registration_student_data'] = form.cleaned_data
            messages.success(request, _("Student information saved."))
            return redirect('students:registration_wizard', step='2')
        else:
            context = self.get_context_data()
            context.update({
                'step': 1,
                'form': form,
                'title': _('Student Information'),
                'next_step': '2'
            })
            return render(request, 'students/registration/step1_student.html', context)
    
    def process_step2(self, request):
        """Process step 2: Guardian Information"""
        form = GuardianForm(
            request.POST,
            tenant=request.tenant,
            user=request.user,
            prefix='step2'
        )
        
        if form.is_valid():
            # Store guardian data in session
            request.session['registration_guardian_data'] = form.cleaned_data
            messages.success(request, _("Guardian information saved."))
            return redirect('students:registration_wizard', step='3')
        else:
            context = self.get_context_data()
            context.update({
                'step': 2,
                'form': form,
                'title': _('Guardian Information'),
                'next_step': '3',
                'prev_step': '1'
            })
            return render(request, 'students/registration/step2_guardian.html', context)
    
    def process_step3(self, request):
        """Process step 3: Address Information"""
        # Store address data in session
        request.session['registration_address_data'] = {
            'address_line1': request.POST.get('address_line1', ''),
            'address_line2': request.POST.get('address_line2', ''),
            'city': request.POST.get('city', ''),
            'state': request.POST.get('state', ''),
            'pincode': request.POST.get('pincode', ''),
            'country': request.POST.get('country', ''),
            'address_type': request.POST.get('address_type', 'PERMANENT'),
        }
        
        messages.success(request, _("Address information saved."))
        return redirect('students:registration_wizard', step='4')
    
    @method_decorator(transaction.atomic)
    def process_step4(self, request):
        """Process step 4: Complete Registration"""
        try:
            # Get all stored data
            student_data = request.session.get('registration_student_data', {})
            guardian_data = request.session.get('registration_guardian_data', {})
            address_data = request.session.get('registration_address_data', {})
            
            if not student_data:
                messages.error(request, _("Registration data expired. Please start again."))
                return redirect('students:registration_wizard', step='1')
            
            # Create student
            student_form = StudentForm(
                data=student_data,
                tenant=request.tenant,
                user=request.user
            )
            
            if not student_form.is_valid():
                messages.error(request, _("Invalid student data. Please start again."))
                return redirect('students:registration_wizard', step='1')
            
            student = student_form.save(commit=False)
            student.tenant = request.tenant
            
            # Generate admission number
            if not student.admission_number:
                student.admission_number = StudentService.generate_admission_number(
                    request.tenant,
                    student.academic_year
                )
            
            student.save()
            
            # Create guardian
            guardian_form = GuardianForm(
                data=guardian_data,
                tenant=request.tenant,
                user=request.user,
                student=student
            )
            
            if guardian_form.is_valid():
                guardian = guardian_form.save(commit=False)
                guardian.tenant = request.tenant
                guardian.save()
                student.guardians.add(guardian)
            
            # Create address
            if address_data:
                StudentAddress.objects.create(
                    student=student,
                    tenant=request.tenant,
                    **address_data
                )
            
            # Handle document upload
            doc_form = StudentDocumentForm(
                request.POST,
                request.FILES,
                tenant=request.tenant,
                user=request.user,
                student=student
            )
            
            if doc_form.is_valid():
                document = doc_form.save(commit=False)
                document.tenant = request.tenant
                document.uploaded_by = request.user
                document.save()
            
            # Create user account
            try:
                student.create_user_account()
            except Exception as e:
                logger.warning(f"User account creation failed: {e}")
            
            # Clear session data
            self.clear_registration_session(request)
            
            # Create comprehensive audit log
            AuditService.create_audit_entry(
                action='CREATE',
                resource_type='StudentRegistration',
                user=request.user,
                tenant=request.tenant,
                request=request,
                severity='INFO',
                instance=student,
                extra_data={
                    'registration_type': 'wizard',
                    'has_guardian': bool(guardian_data),
                    'has_address': bool(address_data),
                    'has_document': 'document' in request.FILES,
                    'admission_number': student.admission_number
                }
            )
            
            # Send notifications
            NotificationService.send_student_registration_notification(
                student=student,
                created_by=request.user
            )
            
            # Send welcome email
            send_student_welcome_email.delay(student.id)
            
            messages.success(
                request,
                _("Student %(name)s registered successfully! Admission Number: %(adm_no)s") % {
                    'name': student.full_name,
                    'adm_no': student.admission_number
                }
            )
            
            return redirect('students:student_detail', pk=student.id)
            
        except Exception as e:
            logger.error(f"Registration error: {str(e)}", exc_info=True)
            messages.error(
                request,
                _("Registration failed: %(error)s. Please try again.") % {'error': str(e)}
            )
            return redirect('students:registration_wizard', step='1')
    
    def clear_registration_session(self, request):
        """Clear registration data from session"""
        keys_to_remove = [
            'registration_student_data',
            'registration_guardian_data',
            'registration_address_data'
        ]
        
        for key in keys_to_remove:
            if key in request.session:
                del request.session[key]
    
    def get_success_url(self):
        return reverse_lazy('students:student_list')


# ============================================================================
# API VIEWS (For AJAX calls)
# ============================================================================

class StudentSearchAPIView(StudentPermissionMixin, BaseView):
    """AJAX endpoint for student search"""
    permission_required = 'students.view_student'
    
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            search_term = request.GET.get('q', '').strip()
            
            if len(search_term) < 2:
                return JsonResponse({'results': []})
            
            queryset = StudentService.get_secure_queryset(request.user, request.tenant)
            
            # Search in multiple fields
            students = queryset.filter(
                Q(admission_number__icontains=search_term) |
                Q(first_name__icontains=search_term) |
                Q(last_name__icontains=search_term) |
                Q(personal_email__icontains=search_term) |
                Q(mobile_primary__icontains=search_term)
            ).select_related('current_class', 'section')[:10]
            
            results = []
            for student in students:
                results.append({
                    'id': student.id,
                    'admission_number': student.admission_number,
                    'full_name': student.full_name,
                    'email': student.personal_email,
                    'phone': student.mobile_primary,
                    'class': student.current_class.name if student.current_class else '',
                    'section': student.section.name if student.section else '',
                    'status': student.get_status_display(),
                    'detail_url': reverse_lazy('students:student_detail', kwargs={'pk': student.id})
                })
            
            return JsonResponse({'results': results})
            
        except Exception as e:
            logger.error(f"Student search error: {str(e)}", exc_info=True)
            return JsonResponse({'error': str(e)}, status=500)


class StudentStatsAPIView(StudentPermissionMixin, BaseView):
    """AJAX endpoint for student statistics"""
    permission_required = 'students.view_student'
    
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            stats_type = request.GET.get('type', 'overview')
            tenant = request.tenant
            
            if stats_type == 'overview':
                stats = self.get_overview_stats(tenant)
            elif stats_type == 'monthly':
                stats = self.get_monthly_stats(tenant)
            elif stats_type == 'class_wise':
                stats = self.get_class_wise_stats(tenant)
            else:
                stats = {}
            
            return JsonResponse({'stats': stats})
            
        except Exception as e:
            logger.error(f"Stats API error: {str(e)}", exc_info=True)
            return JsonResponse({'error': str(e)}, status=500)
    
    def get_overview_stats(self, tenant):
        """Get overview statistics"""
        students = Student.objects.filter(tenant=tenant, is_active=True)
        
        return {
            'total': students.count(),
            'active': students.filter(status=StudentStatus.ACTIVE).count(),
            'inactive': students.filter(status=StudentStatus.INACTIVE).count(),
            'graduated': students.filter(status=StudentStatus.GRADUATED).count(),
            'male': students.filter(gender=StudentGender.MALE).count(),
            'female': students.filter(gender=StudentGender.FEMALE).count(),
            'other': students.filter(gender=StudentGender.OTHER).count(),
        }
    
    def get_monthly_stats(self, tenant):
        """Get monthly admission statistics for current year"""
        current_year = timezone.now().year
        
        monthly_stats = []
        for month in range(1, 13):
            count = Student.objects.filter(
                tenant=tenant,
                admission_date__year=current_year,
                admission_date__month=month
            ).count()
            
            monthly_stats.append({
                'month': month,
                'count': count
            })
        
        return monthly_stats
    
    def get_class_wise_stats(self, tenant):
        """Get class-wise student distribution"""
        classes = SchoolClass.objects.filter(tenant=tenant)
        
        class_stats = []
        for school_class in classes:
            count = Student.objects.filter(
                tenant=tenant,
                current_class=school_class,
                status=StudentStatus.ACTIVE
            ).count()
            
            class_stats.append({
                'class_name': school_class.name,
                'count': count
            })
        
        return class_stats


# ============================================================================
# ERROR HANDLING VIEWS
# ============================================================================

class StudentErrorView(BaseView):
    """Handle student-related errors"""
    template_name = 'students/error.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        error_type = self.kwargs.get('error_type', 'generic')
        
        error_messages = {
            'permission_denied': {
                'title': _('Permission Denied'),
                'message': _('You do not have permission to access this resource.'),
                'code': 403
            },
            'not_found': {
                'title': _('Student Not Found'),
                'message': _('The requested student could not be found.'),
                'code': 404
            },
            'validation_error': {
                'title': _('Validation Error'),
                'message': _('There was an error validating the student data.'),
                'code': 400
            },
            'server_error': {
                'title': _('Server Error'),
                'message': _('An internal server error occurred.'),
                'code': 500
            },
            'bulk_upload_error': {
                'title': _('Bulk Upload Error'),
                'message': _('There was an error processing the bulk upload.'),
                'code': 400
            },
        }
        
        context.update(error_messages.get(error_type, error_messages['generic']))
        return context


